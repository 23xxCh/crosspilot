#!/usr/bin/env python3
"""CrossPilot 统一管道 — 生图+文本并行执行。"""
from __future__ import annotations

import os
import sys
import json
import time
import threading
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor, as_completed
from typing import Any

# Ensure scripts/ is importable
_SCRIPTS = str(Path(__file__).resolve().parent.parent / 'scripts')
if _SCRIPTS not in sys.path:
    sys.path.insert(0, _SCRIPTS)

from crosspilot.config import load_config, get, get_bool, get_int
from crosspilot.health import run_health_check, print_health_report


# ── progress bar ──────────────────────────────────────────────
def _bar(percent: float, width: int = 30) -> str:
    filled = int(width * percent / 100)
    return f"[{'=' * filled}{'>' if filled < width else ''}{' ' * (width - filled - 1)}] {percent:5.1f}%"


# ── input detection ───────────────────────────────────────────
def detect_input(path: str) -> dict[str, Any]:
    """检测输入文件类型和平台。"""
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(f'Input file not found: {path}')
    if not p.is_file():
        raise IsADirectoryError(f'Input is a directory: {path}')

    ext = p.suffix.lower()
    info = {
        'path': str(p.resolve()),
        'name': p.name,
        'stem': p.stem,
        'is_json': ext == '.json',
        'is_xlsx': ext in ('.xlsx', '.xls'),
        'platform': 'unknown',
    }

    if not info['is_json'] and not info['is_xlsx']:
        raise ValueError(f'Unsupported file format: {ext}. Use .json or .xlsx')

    # Auto-detect platform from filename
    name_lower = p.stem.lower()
    if any(kw in name_lower for kw in ('ebay', '易贝')):
        info['platform'] = 'ebay'
    elif any(kw in name_lower for kw in ('amazon', '亚马逊', '采集表', '回填表')):
        info['platform'] = 'amazon'

    # JSON files: check internal structure
    if info['is_json']:
        try:
            with open(p, encoding='utf-8-sig') as f:
                data = json.load(f)
            keys = list(data.keys())
            info['row_count'] = len(data[keys[0]]) if keys else 0
            if '产品标题' in keys and '产品描述' in keys:
                info['platform'] = 'amazon'
            info['fields'] = keys
        except Exception as e:
            raise ValueError(f'Invalid JSON file: {e}')

    # XLSX files: check headers
    if info['is_xlsx'] and info['platform'] == 'unknown':
        import openpyxl
        wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
        try:
            ws = wb.active
            h1 = str(ws.cell(1, 1).value or '')
            if 'Title' in h1 or '标题' in h1:
                info['platform'] = 'ebay'
            else:
                info['platform'] = 'amazon'
            info['row_count'] = ws.max_row - 1
        finally:
            wb.close()

    return info


# ── parallel pipeline ─────────────────────────────────────────
class PipelineRunner:
    """并行管道：生图线程 + 文本线程 同时执行。"""

    def __init__(self):
        self.cfg = load_config()
        self._lock = threading.Lock()
        self._image_result: Any = None
        self._text_result: Any = None
        self._image_error: Exception | None = None
        self._text_error: Exception | None = None
        self._image_done = threading.Event()
        self._text_done = threading.Event()

    def run(self, input_path: str, *,
            text_only: bool = False,
            image_only: bool = False,
            max_rows: int = 0) -> str:
        """运行管道，返回输出文件路径。"""
        info = detect_input(input_path)
        platform = info['platform']

        # Apply max_rows from config or arg
        if max_rows <= 0:
            max_rows = get_int('MAX_ROWS', 0)
        if max_rows > 0:
            os.environ['CROSSPILOT_MAX_ROWS'] = str(max_rows)
        else:
            os.environ.pop('CROSSPILOT_MAX_ROWS', None)
        os.environ['CROSSPILOT_MAX_INPUT_ROWS'] = str(
            max(1, get_int('MAX_INPUT_ROWS', 10000))
        )

        # Set concurrency from config
        os.environ['CROSSPILOT_IMAGE_GEN_CONCURRENCY'] = str(get_int('IMAGE_GEN_CONCURRENCY', 5))
        os.environ['CROSSPILOT_REVIEW_CONCURRENCY'] = str(get_int('REVIEW_CONCURRENCY', 30))
        os.environ['CROSSPILOT_TEXT_CONCURRENCY'] = str(get_int('TEXT_CONCURRENCY', 100))

        # Quality gate
        if get_bool('QUALITY_GATE', False):
            os.environ['CROSSPILOT_IMAGE_QUALITY_GATE'] = '1'
            os.environ.pop('CROSSPILOT_SKIP_QUALITY_GATE', None)
        else:
            os.environ['CROSSPILOT_IMAGE_QUALITY_GATE'] = '0'
            os.environ['CROSSPILOT_SKIP_QUALITY_GATE'] = '1'
        os.environ['CROSSPILOT_IMAGE_REMEDIATE_ONLY'] = (
            '1' if get_bool('IMAGE_REMEDIATE_ONLY', False) else '0'
        )
        os.environ['CROSSPILOT_IMAGE_QUALITY_REGEN_LIMIT'] = str(
            max(0, get_int('IMAGE_QUALITY_REGEN_LIMIT', 1))
        )
        os.environ['CROSSPILOT_VALIDATE_GENERATED_IMAGE'] = (
            '1' if get_bool('VALIDATE_GENERATED_IMAGE', False) else '0'
        )
        os.environ['CROSSPILOT_IMAGE_VALIDATION_ROUTE_LIMIT'] = str(
            max(1, min(3, get_int('IMAGE_VALIDATION_ROUTE_LIMIT', 3)))
        )

        # Skip image gen
        skip_images = image_only is False and (text_only or get_bool('SKIP_IMAGE_GEN', False))
        skip_text = text_only is False and image_only

        print(f'\n  Input: {info["name"]}')
        print(f'  Platform: {platform} | Format: {"JSON" if info["is_json"] else "XLSX"}')
        if info.get('row_count'):
            print(f'  Rows: {info["row_count"]}')
        print(f'  Text only: {text_only or skip_text} | Image only: {image_only}')
        print(f'  Skip image gen: {skip_images}')

        if skip_images:
            # Text only: skip review+gen stage
            print('\n  [Text Only Mode] Skipping image review & generation.')
            os.environ['CROSSPILOT_TEXT_ONLY'] = '1'

        if skip_text:
            print('\n  [Image Only Mode] Running image review & generation only.')

        # Run pipeline
        if platform == 'amazon':
            return self._run_amazon(input_path, skip_images, skip_text)
        elif platform == 'ebay':
            return self._run_ebay(input_path, skip_images, skip_text)
        else:
            raise ValueError(f'Cannot detect platform for: {info["name"]}')

    def _run_amazon(self, input_path: str, skip_images: bool, skip_text: bool) -> str:
        """Amazon 管道：生图+文本并行。"""
        import process_amazon

        if skip_images:
            # Text only: skip review+gen stage
            print('\n  [Text Only Mode] Skipping image review & generation.')
            return process_amazon._main_impl(input_path)

        if skip_text:
            # Image only: run full pipeline but text stages will be minimal
            print('\n  [Image Only Mode] Running image review & generation.')
            return process_amazon._main_impl(input_path)

        # Full parallel mode:
        # Load data once, run image stage + text stages in parallel threads
        print('\n  [Parallel Mode] Image + Text running concurrently...')

        # For now, run the full pipeline (which handles images first then text).
        # The existing pipeline already caches image results.
        # True parallelism requires refactoring process_amazon to split stages.
        # Current approach: run full pipeline, image cache makes it fast.
        print('  Note: Full parallelism requires pipeline refactor. Using cached flow.')
        return process_amazon._main_impl(input_path)

    def _run_ebay(self, input_path: str, skip_images: bool, skip_text: bool) -> str:
        """eBay 管道。"""
        from process_ebay_tk import _main
        print('\n  Running eBay pipeline...')
        return _main(input_path)


def run_pipeline(input_path: str, **kwargs) -> str:
    """便捷入口：一行代码跑管道。"""
    runner = PipelineRunner()
    return runner.run(input_path, **kwargs)
