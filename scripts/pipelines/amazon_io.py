#!/usr/bin/env python3
"""Amazon input parsing, output writing, and delivery validation."""
from __future__ import annotations

import os
import re
import time

import openpyxl

from ..services.amazon_json import load_columnar_json, write_output_json
from .amazon_quality import (
    add_quality_issue as _add_quality_issue,
    validate_amazon_rows as _validate_amazon_rows,
)


def _row_limits():
    """Return (requested output rows, hard input safety limit)."""
    requested = max(0, int(os.environ.get('CROSSPILOT_MAX_ROWS', '0') or 0))
    safety = max(
        1,
        int(os.environ.get('CROSSPILOT_MAX_INPUT_ROWS', '10000') or 10000),
    )
    return requested, safety


def _stage_read_json(ws, tp, progress=None):
    """读 JSON 格式采集表（列名: [值列表]）。"""
    requested_rows, safety_rows = _row_limits()
    raw = load_columnar_json(tp, max_rows=safety_rows)
    titles = raw.get('产品标题', [])
    descs = raw.get('产品描述', [])
    img_urls_list = raw.get('产品图片链接', [])
    var_urls_list = raw.get('变种图片链接', [])
    ids = raw.get('商品id', [])
    n = min(len(titles), requested_rows) if requested_rows else len(titles)
    print(f'读取 JSON: {n} 行', flush=True)
    data = []
    for i in range(n):
        img_urls = [
            url.strip() for url in img_urls_list[i]
            if url.strip().startswith(('http://', 'https://'))
        ]
        var_urls = [
            url.strip() for url in var_urls_list[i]
            if url.strip().startswith(('http://', 'https://'))
        ]
        data.append({
            'id': str(ids[i] or ''),
            'title': str(titles[i] or '').strip() if i < len(titles) else '',
            'desc': str(descs[i] or '') if i < len(descs) else '',
            'main_img': img_urls[0] if img_urls else '',
            'extra_imgs': img_urls[1:] if len(img_urls) > 1 else [],
            'var_imgs': var_urls,
            'var_img': var_urls[0] if var_urls else '',
        })
        if progress:
            progress(i + 1, n)
    return data


def _stage_read(ws, adapter, progress=None):
    """读采集表数据到内存，处理多 URL 图片列（换行分隔）。"""
    requested_rows, _ = _row_limits()
    available = max(0, ws.max_row - 1)
    total = min(available, requested_rows) if requested_rows else available
    print(f"读取 {total} 行...", flush=True)
    data = []
    for r in range(2, total + 2):
        img_raw = str(ws.cell(r, adapter.cols['main_image']).value or '').strip()
        var_raw = str(ws.cell(r, adapter.cols['variant']).value or '').strip()
        # 多 URL 用换行分隔 → 拆成列表
        img_urls = [u.strip() for u in img_raw.replace('\r', '').split('\n') if u.strip().startswith('http')]
        var_urls = [u.strip() for u in var_raw.replace('\r', '').split('\n') if u.strip().startswith('http')]
        main_img = img_urls[0] if img_urls else ''
        extra_imgs = img_urls[1:] if len(img_urls) > 1 else []
        data.append({
            'id': str(ws.cell(r, 1).value or '').strip(),
            'title': str(ws.cell(r, adapter.cols['title']).value or '').strip(),
            'desc': str(ws.cell(r, adapter.cols['desc']).value or ''),
            'main_img': main_img,
            'var_imgs': var_urls,
            'var_img': var_urls[0] if var_urls else '',
            'extra_imgs': extra_imgs,
        })
        if progress:
            progress(r - 1, total)
    return data


def _validate_amazon_input(data):
    """拒绝空表、超大表和缺少核心商品字段的数据。"""
    if not data:
        raise ValueError('Amazon 输入表没有商品数据')
    _, safety_rows = _row_limits()
    if len(data) > safety_rows:
        raise ValueError(
            f'Amazon 输入表有 {len(data)} 行，超过安全上限 {safety_rows} 行'
        )

    issues = []
    fatal_issue_count = 0
    for index, row in enumerate(data, 1):
        if not str(row.get('title', '')).strip():
            fatal_issue_count += 1
            if len(issues) < 20:
                issues.append(f'第 {index} 行缺少产品标题')
        if not str(row.get('desc', '')).strip():
            _add_quality_issue(
                row,
                'missing_source_description',
                '源产品描述为空，已保留空值并标记人工复核',
            )
        main_img = str(row.get('main_img', '')).strip()
        if not re.match(r'^https?://', main_img, re.IGNORECASE):
            fatal_issue_count += 1
            if len(issues) < 20:
                issues.append(f'第 {index} 行缺少有效的主图 URL')
    if issues:
        suffix = '（仅显示前 20 项）' if fatal_issue_count > len(issues) else ''
        raise ValueError('Amazon 输入质量检查失败' + suffix + '：' + '；'.join(issues))

def _validate_amazon_output(output, expected_rows, extra_issues=None):
    """重新打开输出文件，返回结构化交付质量结果。"""
    wb = openpyxl.load_workbook(output, read_only=True, data_only=True)
    try:
        ws = wb.active
        actual_rows = max(0, ws.max_row - 2)
        if actual_rows != expected_rows:
            raise RuntimeError(
                f'Amazon 输出行数不一致：预期 {expected_rows} 行，实际 {actual_rows} 行'
            )

        rows = []
        for row_number in range(3, ws.max_row + 1):
            rows.append({
                'title': ws.cell(row_number, 1).value,
                'desc': ws.cell(row_number, 2).value,
                'main_img': ws.cell(row_number, 3).value,
                'bullets': [
                    ws.cell(row_number, column).value
                    for column in range(18, 23)
                ],
                'keywords': ws.cell(row_number, 23).value,
            })
        validation = _validate_amazon_rows(rows, extra_issues=extra_issues, row_offset=2)
        if validation['issues']:
            suffix = '（仅显示前 20 项）' if validation.get('truncated') else ''
            print(
                f'\n[WARN] Amazon 输出质量检查：{suffix}'
                + '；'.join(validation['issues']),
                flush=True,
            )
        return validation
    finally:
        wb.close()


def _stage_write_output(data, input_path, progress=None):
    """按输入格式写回填 JSON 或 24 列 XLSX。"""
    if input_path.lower().endswith('.json'):
        output = write_output_json(data, input_path)
        if progress:
            progress(len(data), len(data))
        print(f"完成! 保存 JSON: {output}", flush=True)
        return output

    output = os.path.splitext(input_path)[0] + '_回填.xlsx'
    if os.path.exists(output):
        output = os.path.splitext(input_path)[0] + time.strftime('_回填_%H%M%S_') + \
            str(int(time.time() * 1000) % 1000).zfill(3) + '.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Sheet1'

    # Row 1: Headers (exact match to template)
    headers = [
        '产品标题', '产品描述', '产品图片(本地地址)', '变体图片(本地地址)',
        '制造商', 'Model Number(型号)', 'Model Name(型号名称)',
        'Item Package Length(包装长度)', 'Package Length Unit(包装长度单位)',
        'Item Package Width(包装宽度)', 'Package Width Unit(包装宽度单位)',
        'Item Package Height(包装高度)', 'Package Height Unit(包装高度单位)',
        'Package Weight(包装重量)', 'Package Weight Unit(包装重量单位)',
        'MPN', '促销价 (USD)', 'Bullet Point1', 'Bullet Point2',
        'Bullet Point3', 'Bullet Point4', 'Bullet Point5',
        '关键词信息', 'UPC豁免:'
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(1, c).value = h

    # Row 2: Default values (template reference row)
    defaults = [
        '同事提供', '同事提供', '同事提供', '同事提供',
        'Generic', '随机生成数值', '随机生成',
        '0.1', 'Inches(英寸)', '0.1', 'Inches(英寸)',
        '0.1', 'Inches(英寸)', '0.1', 'Kilograms(公斤）',
        '保存与SKU一致', '数值清空', '同事提供', '同事提供',
        '同事提供', '同事提供', '同事提供',
        '同事提供', '是'
    ]
    for c, v in enumerate(defaults, 1):
        ws.cell(2, c).value = v

    # Data rows start from row 3
    for i, row in enumerate(data):
        r = i + 3
        ws.cell(r, 1).value = row['title']
        ws.cell(r, 2).value = row['desc']
        # 主图 → Col3, 所有附图 → Col4（换行分隔）
        ws.cell(r, 3).value = row['main_img']
        image_urls = list(dict.fromkeys(row.get('extra_imgs', []) + row.get('var_imgs', [])))
        ws.cell(r, 4).value = '\n'.join(image_urls) if image_urls else ''
        # Cols 5-17: left as template defaults (manufacturer, package, MPN, etc.)
        for bi in range(5):
            ws.cell(r, 18 + bi).value = row.get('bullets', [''] * 5)[bi] if bi < len(row.get('bullets', [])) else ''
        ws.cell(r, 23).value = row.get('keywords', '')
        if progress:
            progress(i + 1, len(data))

    # Column widths (all 24 columns)
    widths = {1: 50, 2: 80, 3: 60, 4: 60, 5: 12, 6: 18, 7: 22, 8: 21, 9: 25,
              10: 22, 11: 27, 12: 25, 13: 23, 14: 23, 15: 30, 16: 18, 17: 13,
              18: 60, 19: 60, 20: 60, 21: 60, 22: 60, 23: 50, 24: 12}
    wrap_cols = {1, 2, 3, 4, 18, 19, 20, 21, 22, 23}  # text-heavy columns
    for col, w in widths.items():
        letter = openpyxl.utils.get_column_letter(col)
        ws.column_dimensions[letter].width = w
    # Text wrap for content columns
    from openpyxl.styles import Alignment
    wrap_align = Alignment(wrap_text=True, vertical='top')
    for r in range(1, ws.max_row + 1):
        for c in wrap_cols:
            ws.cell(r, c).alignment = wrap_align
    # Header row bold
    from openpyxl.styles import Font
    for c in range(1, 25):
        ws.cell(1, c).font = Font(bold=True)

    wb.save(output)
    wb.close()
    print(f"完成! 保存: {output}", flush=True)
    return output
