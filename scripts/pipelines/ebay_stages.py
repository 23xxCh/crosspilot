"""eBay 管道阶段函数 + 编排。从 ebay_shared 导入共享状态。

使用 model_provider 进行所有 AI 调用，与具体模型解耦。
"""
import json
import os
import re
import sys
import time
import hashlib
import glob
from concurrent.futures import ThreadPoolExecutor, as_completed

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import openpyxl
from crosspilot.prompt_registry import (
    build_runtime_signature,
    get_prompt_registry,
)
from adapters import detect_adapter
from pipeline_log import log as _log, new_request_id, PipelineMetrics
from services.constants import IMAGE_POLICY_VERSION
from model_provider import ProviderQuotaError
from concurrency import adaptive_map
from pipelines.ebay_shared import (
    # sessions & keys
    reload_credentials, get_provider,
    # constants
    GEN_CONCURRENCY, TEXT_CONCURRENCY, REVIEW_CONCURRENCY,
    TMP_DIR,
    # column mapping
    _cols, _init_col_defaults, _apply_adapter_cols,
    # status
    StatusReporter, _DASHBOARD_HOOK,
    # text helpers
    _translate_svc, _strip_code_fence, clean_text_ai, DESC_PROMPT,
    # translate helpers
    translate_text, TITLE_TRANSLATE_PROMPT, TRANSLATE_PROMPT,
    _select_prompt, _CHINESE_RE, _BRAND_PATTERN,
    # image helpers
    review_single, rule_strip_brands, _gen_image,
    IMG_TAG_RE, embed_new_images_in_desc,
    # batch
    batch_translate_texts, batch_clean_texts,
)


def _is_valid_url(val):
    """检查值是否为有效的 HTTP(S) URL。"""
    if isinstance(val, str):
        return val.startswith('http://') or val.startswith('https://')
    return False


TEXT_CACHE_POLICY_VERSION = 'ebay_text_cache_v1'
_prompts = get_prompt_registry()


def _text_model_signature():
    """Return provider/model identity without exposing API keys."""
    from crosspilot.config import load_config

    cfg = load_config()
    configured = cfg.get('TEXT_PROVIDER', 'deepseek')
    models = (
        [
            cfg.get('AGNES_TEXT_MODEL', ''),
        ]
        if configured == 'agnes'
        else [
            cfg.get('DEEPSEEK_TEXT_MODEL', ''),
            cfg.get('DEEPSEEK_TEXT_FALLBACK_MODEL', ''),
        ]
    )
    return {
        'provider': configured,
        'models': [model for model in models if model],
    }


def _current_text_cache_version():
    return build_runtime_signature(
        TEXT_CACHE_POLICY_VERSION,
        "translation.title",
        "ebay.description_clean",
        "translation.text",
    )


def _current_image_cache_version():
    return build_runtime_signature(
        IMAGE_POLICY_VERSION,
        "images.review",
        "images.main_product",
        "images.variant",
    )


def _record_cache_stat(cache, name, hits=0, misses=0):
    stats = cache.setdefault('cache_stats', {})
    item = stats.setdefault(name, {'hits': 0, 'misses': 0})
    item['hits'] += max(0, int(hits or 0))
    item['misses'] += max(0, int(misses or 0))


def _record_concurrency_stat(cache, name, stats):
    if not isinstance(stats, dict):
        return
    cache.setdefault('concurrency_stats', {})[name] = stats


class PipelineContext:
    """Pipeline shared state — passed through all stages."""
    __slots__ = ('tp','total_rows','titles_mem','descs_mem','mains_mem','atts_mem',
                 'variants_mem','videos_mem','url_map','row_images','all_urls',
                 'cache','review_results','gen_results','cleared_att_urls',
                 'to_delete_att','adapter','status','output_path')


def _run_stage(name, fn, status, metrics, *args, item_count=1):
    """包装阶段函数：计时 + 错误处理 + 写入 status.json。"""
    t0 = time.time()
    try:
        result = fn(*args)
        metrics.record_stage(name, time.time() - t0, max(0, item_count))
        return result
    except Exception as e:
        _log.error(f"阶段 [{name}] 失败", error=str(e), exc_info=True)
        try:
            with open(status.status_path, 'w', encoding='utf-8') as f:
                json.dump({'stage': '错误', 'error': f'{name}: {e}',
                           'request_id': new_request_id()}, f, ensure_ascii=False)
        except Exception as se:
            _log.error("状态文件写入失败", error=str(se)[:100])
        raise


def _setup_cache(tp):
    """初始化全局哈希缓存（断点续跑）。返回 (cache_dict, save_fn)。"""
    data_dir = os.environ.get('CROSSPILOT_DATA_DIR') or os.path.join(
        os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))), 'data')
    _CACHE_DIR = os.path.join(data_dir, 'cache')
    os.makedirs(_CACHE_DIR, exist_ok=True)
    cache_version = 2
    text_cache_version = _current_text_cache_version()
    image_cache_version = _current_image_cache_version()
    _now = time.time()
    for _cf in glob.glob(os.path.join(_CACHE_DIR, '*.json')):
        try:
            if os.path.getmtime(_cf) < _now - 30 * 86400:
                os.remove(_cf)
        except OSError as e:
            _log.warn("缓存清理失败", file=_cf, error=str(e)[:100])
    with open(tp, 'rb') as _hash_f:
        _FILE_HASH = hashlib.file_digest(_hash_f, 'sha256').hexdigest()[:16] \
            if hasattr(hashlib, 'file_digest') else hashlib.sha256(_hash_f.read()).hexdigest()[:16]
    CACHE_PATH = os.path.join(_CACHE_DIR, f'{_FILE_HASH}.json')

    def _load():
        try:
            with open(CACHE_PATH, encoding='utf-8') as _cache_f:
                c = json.load(_cache_f)
            if c.get('version') != cache_version:
                raise ValueError("缓存版本已过期")
            if (
                c.get('image_policy_version') != IMAGE_POLICY_VERSION
                or c.get('image_cache_version') != image_cache_version
            ):
                c['review_results'] = {}
                c['gen_results'] = {}
                c['image_policy_version'] = IMAGE_POLICY_VERSION
                c['image_cache_version'] = image_cache_version
                print("图片策略已更新，旧图审和生图缓存已失效", flush=True)
            if c.get('text_cache_version') != text_cache_version:
                c['title_translations'] = {}
                c['desc_cleaned'] = {}
                c['desc_translations'] = {}
                c['text_cache_version'] = text_cache_version
                print("文本策略/模型已更新，旧标题/描述文本缓存已失效", flush=True)
            c.setdefault('cache_stats', {})
            print(f"缓存命中: {len(c.get('review_results',{}))} 图审 + {len(c.get('gen_results',{}))} 生图 + "
                  f"{len(c.get('title_translations',{}))} 标题翻译 + {len(c.get('desc_cleaned',{}))} 描述清洗 + "
                  f"{len(c.get('desc_translations',{}))} 描述翻译", flush=True)
            return c
        except Exception:
            return {'version': cache_version, 'image_policy_version': IMAGE_POLICY_VERSION,
                    'image_cache_version': image_cache_version,
                    'text_cache_version': text_cache_version, 'cache_stats': {},
                    'review_results':{},'gen_results':{},
                    'title_translations':{},'desc_cleaned':{},'desc_translations':{}}

    def _save(c):
        try:
            c['version'] = cache_version
            c['image_policy_version'] = IMAGE_POLICY_VERSION
            c['text_cache_version'] = text_cache_version
            tmp = CACHE_PATH + f'.{os.getpid()}.{id(c)}.tmp'
            with open(tmp, 'w', encoding='utf-8') as f:
                json.dump(c, f, ensure_ascii=False)
            os.replace(tmp, CACHE_PATH)
        except Exception as e:
            _log.warn("缓存保存失败", error=str(e))

    return _load(), _save


# === Pipeline stage functions ===

def _stage_review(status, all_urls, url_map, cache, _save_cache):
    """检测水印、品牌覆盖和人物，按图片角色决定重生或删除。"""
    cached_review = cache.get('review_results', {})
    to_review = list(all_urls)
    review_results = {u: cached_review[u] for u in to_review if u in cached_review}
    to_review_new = [u for u in to_review if u not in cached_review]
    _record_cache_stat(cache, 'review_results', len(review_results), len(to_review_new))
    if review_results:
        print(f"图审缓存命中: {len(review_results)}/{len(to_review)} 张，剩余 {len(to_review_new)} 张待审", flush=True)
    status.start_stage('Agnes图审', len(to_review_new))

    t0 = time.time(); reviewed = 0

    def _review_done(url, result):
        nonlocal reviewed
        review_results[url] = None if isinstance(result, Exception) else result
        reviewed += 1
        status.update(reviewed)

    _review_batch, review_stats = adaptive_map(
        to_review_new,
        review_single,
        operation='review',
        initial_workers=REVIEW_CONCURRENCY,
        min_workers=2,
        is_success=lambda result: result is not None and not isinstance(result, Exception),
        on_result=_review_done,
        terminal_exceptions=(ProviderQuotaError,),
        backoff_s=2,
        max_backoff_s=15,
    )
    _record_concurrency_stat(cache, 'review', review_stats)
    if review_stats.get('reductions'):
        print(
            f"图审并发自适应降级: {review_stats.get('initial_workers')} → "
            f"{review_stats.get('final_workers')} ({review_stats.get('reductions')} 次)",
            flush=True,
        )

    failed_urls = [u for u, r in review_results.items() if r is None]
    if failed_urls:
        print(f"\n[{time.strftime('%H:%M:%S')}] 第一轮失败 {len(failed_urls)} 张，第二轮低并发重审...", flush=True)
        _retry_batch, retry_stats = adaptive_map(
            failed_urls,
            review_single,
            operation='review_retry',
            initial_workers=10,
            min_workers=2,
            is_success=lambda result: result is not None and not isinstance(result, Exception),
            on_result=lambda url, result: (
                review_results.__setitem__(url, result)
                if result is not None and not isinstance(result, Exception)
                else None
            ),
            terminal_exceptions=(ProviderQuotaError,),
            backoff_s=2,
            max_backoff_s=15,
        )
        _record_concurrency_stat(cache, 'review_retry', retry_stats)

    cache['review_results'].update({u: r for u, r in review_results.items() if r is not None})
    _save_cache(cache)

    unreviewed = [u for u, r in review_results.items() if r is None]
    if unreviewed:
        raise RuntimeError(f"图审仍有 {len(unreviewed)} 张失败，为避免漏检水印或人物已停止输出")
    to_regen = [u for u, r in review_results.items() if r
                and (url_map.get(u, {}).get('main') or url_map.get(u, {}).get('variant'))]
    to_delete_att = [u for u, r in review_results.items() if r and url_map.get(u, {}).get('att')]
    n_var = sum(1 for u in to_regen if url_map.get(u, {}).get('variant'))
    print(
        f"Agnes图审: {time.time()-t0:.1f}s | 主图/变种需重生: "
        f"{len(to_regen)}(含变种{n_var}) | 附图删除: {len(to_delete_att)}",
        flush=True,
    )
    return review_results, unreviewed, to_regen, to_delete_att


def _stage_generate(status, to_regen, url_map, cache, _save_cache, mains_mem, variants_mem):
    """按主图/变种角色重生，去除水印、品牌覆盖和全部人物。"""
    cached_gen = cache.get('gen_results', {})
    gen_results = {u: cached_gen[u] for u in to_regen if u in cached_gen}
    to_gen_new = [u for u in to_regen if u not in cached_gen]
    _record_cache_stat(cache, 'gen_results', len(gen_results), len(to_gen_new))
    if gen_results:
        print(f"生图缓存命中: {len(gen_results)}/{len(to_regen)} 张，剩余 {len(to_gen_new)} 张待生成", flush=True)
    if to_gen_new:
        t0 = time.time()
        status.start_stage('图生图', len(to_gen_new))
        gen_done = 0

        def _generate_one(url):
            is_main = bool(url_map.get(url, {}).get('main'))
            return _gen_image(url, is_variant=not is_main)

        def _gen_done(url, result):
            nonlocal gen_done
            gen_done += 1
            if result and not isinstance(result, Exception):
                gen_results[url] = result
            status.update(gen_done, force=True)

        _gen_batch, gen_stats = adaptive_map(
            to_gen_new,
            _generate_one,
            operation='image_gen',
            initial_workers=GEN_CONCURRENCY,
            min_workers=2,
            is_success=lambda result: bool(result) and not isinstance(result, Exception),
            on_result=_gen_done,
            terminal_exceptions=(ProviderQuotaError,),
            backoff_s=2,
            max_backoff_s=15,
        )
        _record_concurrency_stat(cache, 'image_gen', gen_stats)
        if gen_stats.get('reductions'):
            print(
                f"图生图并发自适应降级: {gen_stats.get('initial_workers')} → "
                f"{gen_stats.get('final_workers')} ({gen_stats.get('reductions')} 次)",
                flush=True,
            )
        print(f"图生图: {time.time()-t0:.1f}s | 成功: {len(gen_results)}/{len(to_regen)}", flush=True)
        cache['gen_results'].update({u: gen_results[u] for u in to_gen_new if u in gen_results})
        _save_cache(cache)
    missing = [u for u in to_regen if not gen_results.get(u)]
    if missing:
        raise RuntimeError(f"有 {len(missing)} 张需整改的主图/变种图生成失败，已停止输出")
    main_replaced = variant_replaced = 0
    for old_url, new_url in gen_results.items():
        for row_idx in url_map.get(old_url, {}).get('main', []):
            mains_mem[row_idx - 2] = new_url; main_replaced += 1
        for row_idx in url_map.get(old_url, {}).get('variant', []):
            variants_mem[row_idx - 2] = new_url; variant_replaced += 1
    print(f"主图替换: {main_replaced} | 变种图替换: {variant_replaced}")
    return gen_results


def _stage_clear_attachments(status, to_delete_att, url_map, total_rows, atts_mem):
    """清空含水印、品牌覆盖或人物的附图单元格。"""
    status.start_stage('附图清空', len(to_delete_att))
    att_del = 0
    for url in to_delete_att:
        for row_idx, col in url_map.get(url, {}).get('att', []):
            off = row_idx - 2
            if off < total_rows:
                for ai, ac in enumerate(_cols.att):
                    if ac == col:
                        atts_mem[off][ai] = ''
                        att_del += 1
                        break
    print(f"附图清空: {att_del} 个单元格")
    return att_del


def _stage_clear_video(status, total_rows, videos_mem):
    """清空视频连接列。"""
    status.start_stage('视频+模板图清理', total_rows)
    vid_del = 0
    for off in range(total_rows):
        if videos_mem[off]:
            videos_mem[off] = ''
            vid_del += 1
    status.update(total_rows)
    print(f"视频连接清空: {vid_del} 行", flush=True)
    return vid_del


def _stage_clear_templates(total_rows, descs_mem):
    """删除 pushauction/ibay365 模板图。"""
    template_pattern = re.compile(
        r'\s*(?:<br\s*/?>\s*)?<img[^>]*src=["\'][^"\']*(?:pushauction|ibay365)[^"\']*["\'][^>]*>(?:\s*<br\s*/?>)?',
        re.IGNORECASE)
    template_del = 0
    for off in range(total_rows):
        dv = descs_mem[off]
        if dv:
            new_dv = template_pattern.sub('', dv)
            if new_dv != dv:
                descs_mem[off] = new_dv
                template_del += 1
    print(f"模板图清除(pushauction/ibay365): {template_del} 行", flush=True)
    return template_del


def _stage_strip_brands(total_rows, titles_mem):
    """规则清洗标题中的品牌名/平台名。"""
    cnt = 0
    for off in range(total_rows):
        old = titles_mem[off]
        new = rule_strip_brands(old)
        if new != old:
            titles_mem[off] = new
            cnt += 1
    print(f"标题规则清洗(品牌): {cnt} 行", flush=True)
    return cnt


def _stage_embed_images(status, total_rows, descs_mem, row_images, gen_results, cleared_att_urls):
    """替换描述中 __IMG__ 占位符为真正的 <img> 标签。"""
    status.start_stage('嵌入+注入图片', total_rows)
    embed_cnt = 0
    for off in range(total_rows):
        v = descs_mem[off]
        if v and '__IMG__' in v:
            r = off + 2
            urls = row_images.get(r, [])
            if urls:
                new_html = embed_new_images_in_desc(v, urls, gen_results, cleared_att_urls)
                if new_html != v:
                    descs_mem[off] = new_html
                    embed_cnt += 1
    print(f"描述嵌入新图URL: {embed_cnt} 行", flush=True)
    return embed_cnt


def _stage_inject_images(status, total_rows, mains_mem, atts_mem, descs_mem,
                         gen_results, cleared_att_urls):
    """注入主图+附图新 URL 到描述，去重正文中已有 img。"""
    inject_cnt = dedup_cnt = 0
    for off in range(total_rows):
        r = off + 2
        main_url = mains_mem[off]
        if not main_url or not str(main_url).startswith('http'):
            continue
        att_urls = [str(atts_mem[off][ai]).strip()
                    for ai in range(len(_cols.att))
                    if atts_mem[off][ai] and str(atts_mem[off][ai]).startswith('http')]
        final_main = gen_results.get(str(main_url).strip(), str(main_url).strip())
        final_atts = [gen_results.get(u, u) for u in att_urls if u not in cleared_att_urls]
        att_old_to_new = {u: gen_results[u] for u in att_urls if u in gen_results}
        orig_main = str(main_url).strip()
        main_old_to_new = {orig_main: gen_results[orig_main]} if orig_main in gen_results else {}
        desc_v = descs_mem[off]
        for old_url, new_url in {**main_old_to_new, **att_old_to_new}.items():
            desc_v = re.sub(
                r'(<img[^>]*\ssrc=["\'])' + re.escape(old_url) + r'(["\'][^>]*>)',
                r'\1' + new_url + r'\2',
                desc_v, flags=re.IGNORECASE)
        urls_to_remove = {orig_main} | set(att_urls)
        urls_to_remove.update(gen_results.get(u, u) for u in att_urls if u in gen_results)
        if orig_main in gen_results:
            urls_to_remove.add(gen_results[orig_main])
        if urls_to_remove:
            url_alt = '|'.join(re.escape(u) for u in urls_to_remove)
            new_desc = re.sub(
                r'\s*(?:<br\s*/?>\s*)?<img[^>]*src=["\'](' + url_alt + r')["\'][^>]*>(?:\s*<br\s*/?>)?',
                '', desc_v, flags=re.IGNORECASE)
            if new_desc != desc_v: dedup_cnt += 1
            desc_v = new_desc
        img_block = f'<img src="{final_main}"/>' + ''.join(f'<img src="{u}"/>' for u in final_atts)
        descs_mem[off] = img_block + desc_v
        inject_cnt += 1
        if r % 40 == 0: status.update(off + 1)
    status.update(total_rows)
    print(f"描述注入主图+附图URL: {inject_cnt} 行 | 去重删除正文重复img: {dedup_cnt} 处", flush=True)
    return inject_cnt, dedup_cnt


def _stage_translate_titles(status, total_rows, titles_mem, cache, _save_cache):
    """标题翻译：批量翻译 + 个体 fallback + 中文校验。"""
    title_map = {}
    for off in range(total_rows):
        v = titles_mem[off]
        if v: title_map.setdefault(v, []).append(off + 2)
    status.start_stage('标题清洗+翻译', len(title_map))
    t0 = time.time()
    cached_titles = cache.get('title_translations', {})
    new_titles = [t for t in title_map if t not in cached_titles]
    _record_cache_stat(
        cache,
        'title_translations',
        len(title_map) - len(new_titles),
        len(new_titles),
    )
    if new_titles:
        new_results = batch_translate_texts(new_titles)
        cached_titles.update(new_results)
        print(f"标题翻译(batch): {time.time()-t0:.1f}s | {len(new_titles)} 新 + {len(title_map)-len(new_titles)} 缓存命中", flush=True)
    untranslated = [t for t in title_map if t not in cached_titles or _CHINESE_RE.search(cached_titles.get(t, ''))]
    if untranslated:
        print(f"标题翻译(个体fallback): {len(untranslated)} 条重试单条翻译({TEXT_CONCURRENCY}并发)...", flush=True)
        done = 0
        with ThreadPoolExecutor(max_workers=min(TEXT_CONCURRENCY, len(untranslated))) as pool:
            title_prompt = _prompts.get("translation.title")
            futures = {
                pool.submit(translate_text, t, title_prompt): t
                for t in untranslated
            }
            for future in as_completed(futures):
                t = futures[future]
                try:
                    vn = future.result()
                    if vn and vn != t and not _CHINESE_RE.search(vn):
                        cached_titles[t] = vn
                except ProviderQuotaError:
                    for pending in futures:
                        pending.cancel()
                    raise
                except Exception as e:
                    _log.warn("标题个体fallback 异常", error=str(e))
                done += 1
                if done % 20 == 0: print(f"  个体fallback: {done}/{len(untranslated)}", flush=True)
    elif not new_titles:
        print(f"标题翻译: 全部 {len(title_map)} 缓存命中", flush=True)
    cache['title_translations'] = cached_titles
    _save_cache(cache)
    t_changed = 0
    for src, rows in title_map.items():
        translated = cached_titles.get(src, src)
        if translated != src:
            for r in rows:
                titles_mem[r - 2] = translated
                t_changed += 1
    print(f"标题翻译: {time.time()-t0:.1f}s | 翻译: {t_changed} 行 ({len(title_map)} 唯一)", flush=True)
    return t_changed


def _stage_clean_descs(status, total_rows, descs_mem, cache, _save_cache):
    """描述 AI 清洗：去品牌名/退货政策/运费，保留产品特性。"""
    desc_map = {}
    for off in range(total_rows):
        dv = descs_mem[off]
        if dv and dv.strip(): desc_map.setdefault(dv.strip(), []).append(off + 2)
    status.start_stage('描述AI清洗', len(desc_map))
    t0 = time.time()
    cached_desc = cache.get('desc_cleaned', {})
    new_descs = [t for t in desc_map if t not in cached_desc]
    _record_cache_stat(
        cache,
        'desc_cleaned',
        len(desc_map) - len(new_descs),
        len(new_descs),
    )
    if new_descs:
        new_results = batch_clean_texts(new_descs)
        cached_desc.update(new_results)
        print(f"描述清洗(新): {time.time()-t0:.1f}s | {len(new_descs)} 新 + {len(desc_map)-len(new_descs)} 缓存命中", flush=True)
    else:
        print(f"描述清洗: 全部 {len(desc_map)} 缓存命中", flush=True)
    missing = [text for text in desc_map if not cached_desc.get(text)]
    if missing:
        print(f"描述清洗(单条fallback): {len(missing)} 条...", flush=True)
        with ThreadPoolExecutor(max_workers=min(TEXT_CONCURRENCY, len(missing))) as pool:
            futures = {
                pool.submit(
                    clean_text_ai,
                    text,
                    _prompts.get("ebay.description_clean"),
                ): text
                for text in missing
            }
            for future in as_completed(futures):
                text = futures[future]
                try:
                    cleaned = future.result()
                    if cleaned:
                        cached_desc[text] = cleaned
                except ProviderQuotaError:
                    for pending in futures:
                        pending.cancel()
                    raise
                except Exception as e:
                    _log.warn("描述清洗单条fallback异常", error=str(e))
    cache['desc_cleaned'] = cached_desc
    _save_cache(cache)
    changed = 0
    for src, rows in desc_map.items():
        cleaned = cached_desc.get(src, src)
        if cleaned != src:
            for r in rows:
                descs_mem[r - 2] = cleaned
                changed += 1
    print(f"描述清洗: {time.time()-t0:.1f}s | 修改: {changed} 行 ({len(desc_map)} 唯一)", flush=True)
    return changed


def _stage_translate_descs(status, total_rows, descs_mem, cache, _save_cache):
    """描述翻译越南语：批量翻译 + 个体 fallback。"""
    trans_map = {}
    for off in range(total_rows):
        v = descs_mem[off]
        if v and v.strip(): trans_map.setdefault(v.strip(), []).append(off + 2)
    status.start_stage('描述翻译', len(trans_map))
    t0 = time.time()
    cached_trans = cache.get('desc_translations', {})
    new_trans = [t for t in trans_map if t not in cached_trans]
    _record_cache_stat(
        cache,
        'desc_translations',
        len(trans_map) - len(new_trans),
        len(new_trans),
    )
    if new_trans:
        new_results = batch_translate_texts(new_trans)
        cached_trans.update(new_results)
        print(f"描述翻译(batch): {time.time()-t0:.1f}s | {len(new_trans)} 新 + {len(trans_map)-len(new_trans)} 缓存命中", flush=True)
    else:
        print(f"描述翻译: 全部 {len(trans_map)} 缓存命中", flush=True)
    missed = [t for t in trans_map if t not in cached_trans]
    if missed:
        print(f"描述翻译(个体fallback): {len(missed)} 条重试...", flush=True)
        done = 0
        with ThreadPoolExecutor(max_workers=min(TEXT_CONCURRENCY, len(missed))) as pool:
            text_prompt = _prompts.get("translation.text")
            futures = {
                pool.submit(translate_text, t, text_prompt): t
                for t in missed
            }
            for future in as_completed(futures):
                t = futures[future]
                try:
                    vn = future.result()
                    if vn and vn != t: cached_trans[t] = vn
                except ProviderQuotaError:
                    for pending in futures:
                        pending.cancel()
                    raise
                except Exception as e:
                    _log.warn("描述个体fallback 异常", error=str(e))
                done += 1
                if done % 20 == 0: print(f"  个体fallback: {done}/{len(missed)}", flush=True)
    cache['desc_translations'] = cached_trans
    _save_cache(cache)
    tr_changed = 0
    for src, rows in trans_map.items():
        translated = cached_trans.get(src, src)
        if translated != src:
            for r in rows:
                descs_mem[r - 2] = translated
                tr_changed += 1
    print(f"描述翻译: {time.time()-t0:.1f}s | 翻译: {tr_changed} 行", flush=True)
    return tr_changed


def _stage_finalize(status, tp, wb, adapter, total_rows, titles_mem, descs_mem,
                    mains_mem, atts_mem, variants_mem, videos_mem, cache):
    """价格列改名 + 删除本地展示价列 + 写回 Excel + 校验。返回输出路径。"""
    status.start_stage('价格列+保存', 1)
    ws2 = wb[adapter.sheet_name] if adapter and adapter.sheet_name and adapter.sheet_name in wb.sheetnames else wb.active
    ws2.cell(1, _cols.price).value = '本地展示价'
    print(f"价格列(col{_cols.price})改名'本地展示价'；删除原本地展示价列(col{_cols.local_price})；库存不动", flush=True)
    print("写回 Excel...", flush=True)
    for off in range(total_rows):
        r = off + 2
        ws2.cell(r, _cols.title).value = titles_mem[off]
        ws2.cell(r, _cols.desc).value = descs_mem[off]
        ws2.cell(r, _cols.main).value = mains_mem[off]
        ws2.cell(r, _cols.variant).value = variants_mem[off]
        ws2.cell(r, _cols.video).value = videos_mem[off]
        for ai, ac in enumerate(_cols.att):
            ws2.cell(r, ac).value = atts_mem[off][ai]
    ws2.delete_cols(_cols.local_price)
    save = os.path.splitext(tp)[0] + '_cleaned.xlsx'
    if os.path.exists(save):
        save = os.path.splitext(tp)[0] + time.strftime('_cleaned_%H%M%S_') + str(int(time.time() * 1000) % 1000).zfill(3) + '.xlsx'
        print(f"⚠️ 输出文件已存在，改存: {save}", flush=True)
    try:
        wb.save(save)
    finally:
        wb.close()
    validation = _validate_output(save, adapter, cache)
    if not validation.get('passed'):
        try:
            with open(status.status_path, 'w', encoding='utf-8') as f:
                json.dump({'status': 'failed', 'stage': '错误',
                           'error': '输出校验失败: ' + '; '.join(validation.get('warnings', [])),
                           'validation': validation}, f, ensure_ascii=False, indent=2)
        except OSError:
            pass
        raise RuntimeError('输出校验失败: ' + '; '.join(validation.get('warnings', [])))
    status.finish(save)
    if validation:
        try:
            with open(status.status_path, 'r', encoding='utf-8') as f:
                st = json.load(f)
            st['validation'] = validation
            with open(status.status_path, 'w', encoding='utf-8') as f:
                json.dump(st, f, ensure_ascii=False, indent=2)
        except Exception as e:
            _log.warn("校验结果写入 status 失败", error=str(e))
    print(f"完成! 保存: {save}", flush=True)
    return save


def _validate_output(output_path, adapter, cache):
    """输出校验：打开输出文件检查关键列。"""
    try:
        wb = openpyxl.load_workbook(output_path, data_only=True)
        ws = wb[adapter.sheet_name] if (
            adapter and adapter.sheet_name and adapter.sheet_name in wb.sheetnames
        ) else wb.active
    except Exception:
        return {"passed": False, "warnings": ["无法打开输出文件进行校验"]}
    warnings = []
    cols = adapter.cols if adapter else {}
    deleted_col = cols.get('local_price', 16)
    shifted = lambda col: col - 1 if col > deleted_col else col
    title_col = shifted(cols.get('title', 2))
    main_col = shifted(cols.get('main_image', 18))
    price_col = shifted(cols.get('price', 15))
    video_col = shifted(cols.get('video', 27))
    total = ws.max_row - 1
    cn_count = sum(1 for r in range(2, ws.max_row + 1)
                   if _CHINESE_RE.search(str(ws.cell(r, title_col).value or '')))
    cn_rate = cn_count / max(total, 1)
    if cn_rate > 0.05:
        warnings.append(f"标题中文残留 {cn_count}/{total} ({cn_rate:.0%})，阈值 5%")
    elif cn_count > 0:
        warnings.append(f"标题中文残留 {cn_count}/{total} ({cn_rate:.0%})，在容忍范围内")
    empty_main = sum(1 for r in range(2, ws.max_row + 1)
                     if not str(ws.cell(r, main_col).value or '').startswith('http'))
    if empty_main > total * 0.1:
        warnings.append(f"主图列 {empty_main}/{total} 行无有效 URL")
    price_header = str(ws.cell(1, price_col).value or '')
    if '展示价' not in price_header:
        warnings.append(f"价格列未改名为本地展示价，当前: {price_header}")
    video_filled = sum(1 for r in range(2, ws.max_row + 1)
                       if str(ws.cell(r, video_col).value or '').strip())
    if video_filled > 0:
        warnings.append(f"视频列未清空，{video_filled} 行仍有内容")
    wb.close()
    passed = len([w for w in warnings if '容忍' not in w]) == 0
    result = {"passed": passed, "warnings": warnings}
    if warnings:
        print(f"\n{'='*50}\n输出校验 {'✅ 通过' if passed else '⚠️ 发现问题'}:", flush=True)
        for w in warnings:
            print(f"  {'⚠️' if '容忍' not in w else 'ℹ️'} {w}", flush=True)
        print('='*50, flush=True)
    return result


# === Main pipeline orchestration ===

def _main(table_path=None, _TABLE_PATH=None):
    """处理单个 xlsx。返回输出文件路径（成功）或抛异常（失败）。"""
    tp = table_path or _TABLE_PATH
    if not tp:
        raise ValueError("缺少输入文件路径")
    if hasattr(sys.stdout, 'reconfigure'):
        sys.stdout.reconfigure(encoding='utf-8')
    rid = new_request_id()
    _log.info("管道启动", request_id=rid, file=os.path.basename(tp))
    print(f"=== eBay→TikTok 清洗 (ebay-tk 定制版) === [rid={rid}]")
    print(f"输入: {tp}")
    _init_col_defaults()
    reload_credentials()
    # model_provider 会在首次调用时自动检查配置
    try:
        provider = get_provider()
    except ValueError as e:
        raise ValueError(f"配置错误: {e}")
    wb = openpyxl.load_workbook(tp)

    adapter = detect_adapter(wb.active)
    ws = wb[adapter.sheet_name] if adapter and adapter.sheet_name and adapter.sheet_name in wb.sheetnames else wb.active
    if adapter and adapter.sheet_name and ws.title != adapter.sheet_name:
        print(f"⚠️ 未找到 {adapter.sheet_name} 工作表，使用活动表: {ws.title}", flush=True)
    if adapter is None:
        headers = [str(ws.cell(1, c).value or '') for c in range(1, ws.max_column + 1)]
        print(f"\n❌ 不认识的表格格式（没有适配器能识别）。表头如下：", flush=True)
        for i, h in enumerate(headers, 1):
            if h: print(f"    列{i}: {h}", flush=True)
        raise ValueError("不认识的表格格式，无匹配适配器")
    _apply_adapter_cols(adapter.cols)
    print(f"表格格式: {adapter.name} (适配器注入列映射)", flush=True)
    print(f"工作表: {ws.title}, 行数: {ws.max_row-1}, 列数: {ws.max_column}")

    EXPECTED = {_cols.title: '标题', _cols.desc: '描述', _cols.price: '价格',
                _cols.local_price: '展示价', _cols.stock: '库存', _cols.main: '主图',
                _cols.video: '视频', _cols.variant: '变种'}
    bad = [(c, kw, str(ws.cell(1, c).value or '')) for c, kw in EXPECTED.items()
           if kw not in str(ws.cell(1, c).value or '')]
    if bad:
        print(f"\n❌ 表头结构与预期不符（TikTok 可能改了导出格式）：", flush=True)
        for c, kw, hv in bad:
            print(f"    列{c}: 期望含'{kw}'，实际'{hv}'", flush=True)
        raise ValueError("表头结构与预期不符，已跳过")

    status = StatusReporter(tp)
    total_rows = ws.max_row - 1
    max_rows = max(1, int(os.environ.get('CROSSPILOT_MAX_ROWS', '10000')))
    if total_rows <= 0:
        wb.close()
        raise ValueError("表格没有可处理的数据行")
    if total_rows > max_rows:
        wb.close()
        raise ValueError(f"数据行数 {total_rows} 超过安全上限 {max_rows}，请拆分文件")
    print(f"读取 {total_rows} 行到内存...", flush=True)
    titles_mem = [''] * total_rows; descs_mem = [''] * total_rows
    mains_mem = [''] * total_rows; variants_mem = [''] * total_rows
    videos_mem = [''] * total_rows
    atts_mem = [[''] * len(_cols.att) for _ in range(total_rows)]
    url_map = {}; row_images = {}
    status.start_stage('提取图片URL', total_rows)
    for off, r in enumerate(range(2, ws.max_row + 1)):
        titles_mem[off] = str(ws.cell(r, _cols.title).value or '').strip()
        descs_mem[off] = str(ws.cell(r, _cols.desc).value or '')
        mains_mem[off] = str(ws.cell(r, _cols.main).value or '').strip()
        for ai, ac in enumerate(_cols.att):
            atts_mem[off][ai] = str(ws.cell(r, ac).value or '').strip()
        variants_mem[off] = str(ws.cell(r, _cols.variant).value or '').strip()
        videos_mem[off] = str(ws.cell(r, _cols.video).value or '')
        if '<img' in descs_mem[off]:
            urls = [m.group(1) for m in IMG_TAG_RE.finditer(descs_mem[off])]
            if urls: row_images[r] = urls
        for val, kind in [(mains_mem[off], 'main'), (variants_mem[off], 'variant')]:
            if val and val.startswith('http'):
                url_map.setdefault(val, {'main': [], 'att': [], 'variant': []})[kind].append(r)
        for ai, val in enumerate(atts_mem[off]):
            if val and val.startswith('http'):
                url_map.setdefault(val, {'main': [], 'att': [], 'variant': []})['att'].append((r, _cols.att[ai]))
        status.update(off + 1)
    all_urls = list(url_map.keys())
    print(f"唯一图片(单元格): {len(all_urls)}, 数据行: {total_rows}, 描述含img行: {len(row_images)}", flush=True)

    metrics = PipelineMetrics()

    cache, _save_cache = _run_stage(
        '缓存初始化', _setup_cache, status, metrics, tp, item_count=1
    )
    review_results, unreviewed, to_regen, to_delete_att = _run_stage(
        '图审',
        _stage_review,
        status,
        metrics,
        status,
        all_urls,
        url_map,
        cache,
        _save_cache,
        item_count=len(all_urls),
    )
    gen_results = _run_stage('图生图', _stage_generate, status, metrics, status, to_regen, url_map, cache, _save_cache,
                             mains_mem, variants_mem, item_count=len(to_regen))

    cleared_att_urls = set(to_delete_att)
    _run_stage('附图清空', _stage_clear_attachments, status, metrics, status, to_delete_att, url_map, total_rows, atts_mem, item_count=len(to_delete_att))
    _run_stage('品牌清洗', _stage_strip_brands, status, metrics, total_rows, titles_mem, item_count=total_rows)
    _run_stage('标题翻译', _stage_translate_titles, status, metrics, status, total_rows, titles_mem, cache, _save_cache, item_count=total_rows)
    _run_stage('描述清洗', _stage_clean_descs, status, metrics, status, total_rows, descs_mem, cache, _save_cache, item_count=total_rows)
    _run_stage('描述翻译', _stage_translate_descs, status, metrics, status, total_rows, descs_mem, cache, _save_cache, item_count=total_rows)
    _run_stage('嵌入图片', _stage_embed_images, status, metrics, status, total_rows, descs_mem, row_images, gen_results, cleared_att_urls, item_count=total_rows)
    _run_stage('注入图片', _stage_inject_images, status, metrics, status, total_rows, mains_mem, atts_mem, descs_mem, gen_results, cleared_att_urls, item_count=total_rows)
    _run_stage('视频清空', _stage_clear_video, status, metrics, status, total_rows, videos_mem, item_count=total_rows)
    _run_stage('模板清除', _stage_clear_templates, status, metrics, total_rows, descs_mem, item_count=total_rows)
    output_path = _run_stage('保存', _stage_finalize, status, metrics, status, tp, wb, adapter, total_rows,
                             titles_mem, descs_mem, mains_mem, atts_mem, variants_mem, videos_mem, cache,
                             item_count=total_rows)

    try:
        provider = get_provider()
        if hasattr(provider, 'metrics_snapshot'):
            metrics.set_provider_metrics(provider.metrics_snapshot())
        metrics.set_cache_metrics(cache.get('cache_stats'))
        metrics.set_concurrency_metrics(cache.get('concurrency_stats'))
        with open(status.status_path, 'r', encoding='utf-8') as f:
            st = json.load(f)
        st['metrics'] = metrics.to_dict()
        with open(status.status_path, 'w', encoding='utf-8') as f:
            json.dump(st, f, ensure_ascii=False, indent=2)
    except Exception as e:
        _log.warn("metrics 写入 status 失败", error=str(e)[:100])

    return output_path
