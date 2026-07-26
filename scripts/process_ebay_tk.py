#!/usr/bin/env python3
"""eBay→TikTok Shop 表格清洗 (ebay-tk 定制版) — DMXAPI + Agnes

API 分工：
  - DMXAPI 中转站：MiMo V2.5 图审 + deepseek-v4-flash 文本翻译/描述清洗（不限流）
  - Agnes (agnes-image-2.0-flash)：图生图去水印（返 URL）

流程：
  MiMo图审 -> 主图(含水印)Agnes图生图替换 ->
  附图清空 -> 标题规则清洗 + 翻译越南语 ->
  描述 AI清洗 + 翻译越南语 ->
  嵌入新图URL + 注入主图附图URL到描述(去重) -> 清空视频 -> 删列改名 -> 保存

进度：运行中在输入文件同目录写 <文件名>_status.json（阶段/进度/ETA）
"""
import openpyxl, json, os, re, sys, time, tempfile, hashlib, glob, threading
from concurrent.futures import ThreadPoolExecutor, as_completed
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from adapters import detect_adapter
from pipeline_log import log as _log, new_request_id, PipelineMetrics
from dmx_client import (_post_json, _extract_image_url, gen_wan, gen_doubao,
                        generate_image as _gen_image, WAN_MODEL, DOUBAO_MODEL, GEN_PROMPT)
from services import ImageReviewService, TranslationService, ImageGenService

_RUN_DIRECTLY = __name__ == '__main__' or sys.argv[0].endswith(('process_ebay_tk.py', 'runner.py'))
TABLE_PATH = sys.argv[1] if _RUN_DIRECTLY and len(sys.argv) > 1 else None
if _RUN_DIRECTLY and not TABLE_PATH:
    print("用法: uv run python -u scripts/process_ebay_tk.py \"<输入文件.xlsx>\"")
    sys.exit(1)
# payload 临时文件统一放系统 temp，避免污染工作目录
TMP_DIR = tempfile.gettempdir()
GEN_CONCURRENCY = 20   # 图生图并发数
TEXT_CONCURRENCY = 20  # 文本批次并发数
MAX_RETRIES = 8

# ===== API keys（keys.json 外置，gitignore 保护；兼容旧 agnes_key.txt）=====
_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

def _load_keys():
    keys = {}
    kf = os.environ.get('CROSSPILOT_KEYS_PATH') or os.path.join(_ROOT, 'keys.json')
    if os.path.exists(kf):
        try:
            with open(kf, encoding='utf-8') as f:
                keys = json.load(f)
        except Exception as e:
            _log.warn("keys.json 读取失败", error=str(e))
    return keys

_KEYS = _load_keys()

def _get_dmx_key():
    """实时读取 keys.json（支持 Web 界面改 key 后无需重启）。"""
    return _load_keys().get('dmx_key', '')

# ===== DMXAPI 中转站（图审 + 文本翻译/清洗，不限流）=====
DMX_BASE = "https://www.dmxapi.cn"
import requests as _requests
from requests.adapters import HTTPAdapter as _HTTPAdapter
_http = _requests.Session()
_adapter = _HTTPAdapter(pool_connections=100, pool_maxsize=100, max_retries=0)
_http.mount('https://', _adapter)
_http.mount('http://', _adapter)

def _reload_http_auth():
    """每次 pipeline 启动时刷新 Authorization header（支持 key 热更新）。"""
    _http.headers.update({
        'Authorization': f'Bearer {_get_dmx_key()}',
        'Content-Type': 'application/json'
    })

# 初始加载
DMX_KEY = _KEYS.get('dmx_key', '')
_http.headers.update({
    'Authorization': f'Bearer {DMX_KEY}',
    'Content-Type': 'application/json'
})
MIMO_MODEL = "mimo-v2.5"          # 图审
TEXT_MODEL = "mimo-v2.5"          # 文本翻译/描述清洗（主模型）
TEXT_FALLBACK_MODELS = ["deepseek-v4-flash", "hy3", "step-3.5-flash"]  # 备用模型
MIMO_CONCURRENCY = 100  # DMXAPI 不限流，无限并发

print(f"DMXAPI: {MIMO_MODEL}(图审) + {TEXT_MODEL}(文本) @ {DMX_BASE}")
# Service instances (shared across pipeline runs)
_review_svc = ImageReviewService(_http)
_translate_svc = TranslationService(_http)
_gen_svc = ImageGenService(_http)

# 列映射：threading.local() 保证 exe 线程模式和 dev 子进程模式都安全
_cols = threading.local()

def _init_col_defaults():
    """重置当前线程的列映射为 eBay→TikTok 默认值"""
    _cols.main = 18
    _cols.att = [19,20,21,22,23,24,25,26,28]
    _cols.variant = 29
    _cols.title = 2
    _cols.desc = 3
    _cols.price = 15
    _cols.local_price = 16
    _cols.stock = 17
    _cols.video = 27

_init_col_defaults()  # 模块加载时初始化主线程

# 批量模式看板 hook
_DASHBOARD_HOOK = None


def _apply_adapter_cols(cols):
    """把适配器的列映射写入当前线程的 thread-local 存储"""
    _cols.main = cols['main_image']
    _cols.att = list(cols['attachments'])
    _cols.variant = cols['variant']
    _cols.title = cols['title']
    _cols.desc = cols['desc']
    _cols.price = cols['price']
    _cols.local_price = cols['local_price']
    _cols.stock = cols['stock']
    _cols.video = cols['video']


# ===== 进度状态上报（用户/agent 可随时查看进度）=====
class StatusReporter:
    """在输入文件同目录写 <文件名>_status.json，记录当前阶段/进度/ETA。
    同时打印带时间戳的换行日志（替代 \r 进度条，后台日志可读）。"""

    STAGES = ['提取图片URL', 'MiMo图审', '图生图', '附图清空', '标题清洗+翻译',
              '描述AI清洗', '描述翻译', '嵌入+注入图片', '视频+模板图清理', '价格列+保存']

    def __init__(self, table_path):
        self.status_path = os.path.splitext(table_path)[0] + '_status.json'
        self.t_start = time.time()
        self.stage_idx = 0
        self.stage_t0 = self.t_start
        self.total = 0

    def _write(self, current):
        elapsed = time.time() - self.stage_t0
        total_elapsed = time.time() - self.t_start
        eta = int(elapsed / current * (self.total - current)) if current > 0 and self.total > 0 else 0
        pct = int(current / self.total * 100) if self.total > 0 else 0
        data = {
            'stage': self.STAGES[self.stage_idx],
            'stage_index': self.stage_idx + 1,
            'stage_total': len(self.STAGES),
            'current': current,
            'total': self.total,
            'percent': pct,
            'elapsed_s': int(elapsed),
            'eta_s': eta if eta > 0 else int(total_elapsed / (self.stage_idx + 1) * (len(self.STAGES) - self.stage_idx - 1)) if self.stage_idx < len(self.STAGES) - 1 else 0,
            'total_elapsed_s': int(total_elapsed),
            'updated_at': time.strftime('%Y-%m-%d %H:%M:%S'),
        }
        try:
            with open(self.status_path, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
        except Exception as e:
            _log.warn("status.json 写入失败", error=str(e))
            pass
        return data

    def start_stage(self, name, total=0):
        """进入新阶段（按名称查找索引，不再硬编码编号）。"""
        try:
            self.stage_idx = self.STAGES.index(name)
        except ValueError:
            self.stage_idx = len(self.STAGES) - 1  # fallback to last
        self.total = total
        self.stage_t0 = time.time()
        ts = time.strftime('%H:%M:%S')
        print(f"[{ts}] ===== 阶段 {self.stage_idx+1}/{len(self.STAGES)}: {name} ({total} 项) =====", flush=True)
        self._write(0)

    def update(self, current, force=False):
        """更新进度：每 10% 或每 20 项写盘+打印日志（节流，减少 I/O）。"""
        should_log = force or current == self.total or current % 20 == 0 or (
            self.total > 0 and current % max(1, self.total // 10) == 0)
        if should_log:
            d = self._write(current)
            print(f"[{time.strftime('%H:%M:%S')}] {d['stage']} {current}/{self.total} "
                  f"({d['percent']}%) | 已用 {d['elapsed_s']}s | 预计剩余 {d['eta_s']}s", flush=True)
            if _DASHBOARD_HOOK:
                try: _DASHBOARD_HOOK()
                except Exception: pass

    def finish(self, output_path):
        try:
            with open(self.status_path, 'w', encoding='utf-8') as f:
                json.dump({'stage': '完成', 'output': output_path,
                           'total_elapsed_s': int(time.time() - self.t_start),
                           'updated_at': time.strftime('%Y-%m-%d %H:%M:%S')},
                          f, ensure_ascii=False, indent=2)
        except Exception as e:
            _log.warn("status.json 写入失败", error=str(e))
            pass


# ===== 通用 DMXAPI chat completions 调用（文本翻译/描述清洗）=====
def _err_msg(obj):
    """从 API 错误响应取 message（委托 TranslationService）。"""
    return _translate_svc._err_msg(obj)

def dmx_call(payload, max_tokens_override=None):
    """调用 DMXAPI chat completions（委托 TranslationService）。"""
    return _translate_svc.dmx_call(payload, max_tokens_override)


# ===== 图审：四级 fallback 链 =====
REVIEW_PROMPT = ("Look at this image. Answer YES only if it has a SELLER WATERMARK: semi-transparent/faint text "
"(like a store ID, e.g. 'liazh-93', 'Constituen78') repeated or tiled across the image, or a car brand logo "
"(BMW/Toyota/Honda) overlaid on the photo.\n\n"
"Answer NO for these cases:\n"
"- Product spec text: '2PCS', '1Set', '1PC', dimensions like '48*30CM' (solid text, single instance, in a corner)\n"
"- Product feature callouts or marketing banners: '7 COLOR', 'Soft and Comfortable', feature icons\n"
"- Text that is part of the product itself: emblems/badges like 'SPORT', '4x4', 'LIMITED EDITION' printed ON the product\n"
"- Clean images with no text overlay\n\n"
"Answer YES or NO only.")

FALLBACK_MODEL = "gemini-3.1-flash-lite-image"  # DMXAPI 实测可用的备用视觉模型

def _vision_call(model, image_url, timeout=25):
    """单次视觉判定调用。直接传 URL，不下载图片。
    返回 True=有水印, False=无水印, None=调用失败。"""
    try:
        img_part = {"type": "image_url", "image_url": {"url": image_url}}
        payload = {
            "model": model,
            "messages": [{"role": "user", "content": [
                {"type": "text", "text": REVIEW_PROMPT}, img_part]}],
            "max_completion_tokens": 200
        }
        # 用 requests 替代 curl subprocess——Windows 上 subprocess 超时无法杀死孤儿 curl.exe
        r = _http.post(f'{DMX_BASE}/v1/chat/completions', json=payload,
                      timeout=(min(5, timeout // 3), timeout))
        if not r.ok:
            return None
        obj = r.json()
        if 'choices' not in obj:
            return None
        content_text = obj['choices'][0].get('message', {}).get('content', '')
        if not content_text:
            return None
        return content_text.strip().upper().startswith('YES')
    except Exception as e:
        _log.warn("vision_call 异常", error=str(e))
        return None

_REVIEW_DEADLINE_S = 45  # 单张图审总超时（秒），避免个别 URL 卡死整个阶段


def review_single(image_url):
    """单张图审（委托 ImageReviewService 三级 fallback）。"""
    return _review_svc.review(image_url)

BRANDS = ['bmw','porsche','toyota','honda','mercedes','audi','vw','ford','hyundai',
          'nissan','kia','mazda','lexus','benz','xiaomi','redmi','joyon','shopee','lazada','xmen','diy',
          'smiling','htghtg','yrbwd','lemontree','jojo','zaofahua',
          '宝马','保时捷','小米','红米','丰田','本田','奔驰','奥迪','大众','福特','现代','日产','起亚','马自达','雷克萨斯']
_BRAND_PATTERN = re.compile('|'.join(re.escape(b) for b in BRANDS), re.IGNORECASE)

def rule_strip_brands(t):
    if not t: return t
    t = _BRAND_PATTERN.sub('', str(t))
    return re.sub(r'\s+', ' ', t).strip()


# ===== 描述 AI 清洗（eBay 定制：只留产品特性，img → 占位符） =====
DESC_PROMPT = ("Clean the following HTML product description. Requirements:\n"
"1) Remove ALL third-party brand names and store names (car brands like BMW/Toyota/Honda/Mercedes/Porsche/Audi, "
"platform names like Shopee/Lazada, any other brand/trademark names — English and Chinese both).\n"
"2) Remove ALL content unrelated to the product itself: return policy, payment methods, shipping info, FAQ, "
"store introduction, follow-store prompts, contact info, promotional campaigns, QR codes, any policy boilerplate text.\n"
"3) Keep ONLY product information: product name, features, material, dimensions, specifications, color, quantity, "
"use cases, applicable scenarios/models, installation instructions.\n"
"4) Replace every <img ...> tag with a placeholder __IMG__ (so image positions are preserved).\n"
"5) Output clean HTML with __IMG__ placeholders and <p>/<ul>/<li>/<br> text. No explanation, no prefix/suffix.\n\n"
"Description: {}")


_CHINESE_RE = re.compile(r'[一-鿿]')


def _strip_code_fence(content):
    """去除 markdown 代码围栏和首尾引号（委托 TranslationService）。"""
    return _translate_svc._strip_code_fence(content)


def clean_text_ai(text, prompt):
    """AI 清洗文本（dmx_call 内部已有重试+模型 fallback，无需外层循环）。"""
    if not text or not str(text).strip(): return text
    text = str(text)
    try:
        payload = {"model": TEXT_MODEL,
                   "messages": [{"role": "user", "content": prompt.format(text)}],
                   "max_completion_tokens": 2048}
        content = dmx_call(payload, max_tokens_override=2048)
        if content is not None:
            content = _strip_code_fence(content)
            if content:
                return rule_strip_brands(content)
    except Exception as e:
        _log.warn("clean_text_ai 异常", error=str(e))
    return rule_strip_brands(text)


# ===== 英文/中文 → 越南语翻译（保留产品信息+产品关键词） =====
TRANSLATE_PROMPT = ("Translate the following product text to Vietnamese.\n"
"Rules:\n"
"- Preserve ALL product information: material, dimensions, specifications, color, quantity, type, model numbers, "
"product names, technical specs — keep numbers/units/model codes UNCHANGED.\n"
"- Keep the original formatting, HTML tags (p, ul, li, br, img) and punctuation intact.\n"
"- IMPORTANT: Keep the placeholder __IMG__ EXACTLY as-is (do not translate or remove it).\n"
"- If already in Vietnamese, return unchanged.\n"
"Output Vietnamese translation directly, no explanation or prefix/suffix.\n\n"
"Text: {}")


# ===== 标题翻译（英→越，保留产品关键词不译） =====
TITLE_TRANSLATE_PROMPT = ("Translate the following product title to Vietnamese.\n"
"Rules:\n"
"- KEEP product model numbers, brand-compatible vehicle names/years, part numbers, and technical codes in original form (do not translate proper nouns/specs).\n"
"- Translate descriptive words, connectors, prepositions, and general terms to Vietnamese naturally.\n"
"- Do NOT add, remove, or rewrite any original product meaning.\n"
"Output only the translated title, no explanation or quotes.\n\n"
"Title: {}")


def _select_prompt(text, default_prompt):
    """中文输入用专用 prompt（委托 TranslationService）。"""
    return _translate_svc._select_prompt(text, default_prompt)


def translate_text(text, prompt):
    """翻译为越南语（委托 TranslationService，主模型+全部fallback模型）。"""
    return _translate_svc.translate_text(text, prompt)

def embed_new_images_in_desc(html, img_queue, gen_results, cleared_att_urls):
    """把描述中的 __IMG__ 占位符替换为真正的 <img> 标签。
    img_queue: 该行原始描述中按顺序出现的图片 URL 列表。
    优先级: 图生图新 URL > 原始 URL > 跳过(附图已清空)。
    """
    if not html: return html
    it = iter(img_queue)

    def _replace(_m):
        try:
            src = next(it)
        except StopIteration:
            return ''
        if src in gen_results:
            return f'<img src="{gen_results[src]}"/>'
        if src in cleared_att_urls:
            return ''
        return f'<img src="{src}"/>'

    result = re.sub(r'__IMG__', _replace, html)
    return result


class PipelineContext:
    """Pipeline shared state — passed through all stages."""
    __slots__ = ('tp','total_rows','titles_mem','descs_mem','mains_mem','atts_mem',
                 'variants_mem','videos_mem','url_map','row_images','all_urls',
                 'cache','review_results','gen_results','cleared_att_urls',
                 'to_delete_att','adapter','status','output_path',
                 'review_svc','translate_svc','gen_svc')


def _main(table_path=None):
    """处理单个 xlsx。table_path 为 None 时用全局 TABLE_PATH（命令行模式）。
    返回输出文件路径（成功）或抛异常（失败），供批量队列判断。"""
    tp = table_path or TABLE_PATH
    if hasattr(sys.stdout, 'reconfigure'):
        sys.stdout.reconfigure(encoding='utf-8')
    rid = new_request_id()
    _log.info("管道启动", request_id=rid, file=os.path.basename(tp))
    print(f"=== eBay→TikTok 清洗 (ebay-tk 定制版) === [rid={rid}]")
    print(f"输入: {tp}")
    _init_col_defaults()  # 线程安全：重置当前线程的列映射为默认值
    _reload_http_auth()  # 实时读取 key（支持热更新）
    if not _get_dmx_key():
        raise ValueError("缺少 DMXAPI key：请在 keys.json 配置 dmx_key（参考 keys.example.json）")
    wb = openpyxl.load_workbook(tp, data_only=True)

    # ---- 表格格式识别（适配器）：先识别来源格式，再按适配器找对应工作表 ----
    adapter = detect_adapter(wb.active)
    # 按适配器指定的工作表名查找，找不到则用活动表
    ws = wb[adapter.sheet_name] if adapter and adapter.sheet_name and adapter.sheet_name in wb.sheetnames else wb.active
    if adapter and adapter.sheet_name and ws.title != adapter.sheet_name:
        print(f"⚠️ 未找到 {adapter.sheet_name} 工作表，使用活动表: {ws.title}", flush=True)

    if adapter is None:
        headers = [str(ws.cell(1, c).value or '') for c in range(1, ws.max_column + 1)]
        print(f"\n❌ 不认识的表格格式（没有适配器能识别）。表头如下：", flush=True)
        for i, h in enumerate(headers, 1):
            if h: print(f"    列{i}: {h}", flush=True)
        print("请在 scripts/adapters/ 下照 ebay_tk.py 模板新增一个适配器（填 detect + cols），", flush=True)
        print("或确认这是否为支持的来源表格。", flush=True)
        raise ValueError("不认识的表格格式，无匹配适配器")
    _apply_adapter_cols(adapter.cols)
    print(f"表格格式: {adapter.name} (适配器注入列映射)", flush=True)
    print(f"工作表: {ws.title}, 行数: {ws.max_row-1}, 列数: {ws.max_column}")

    # ---- 0. 表头结构校验（防止 TikTok 改格式后删错列/写错列）----
    EXPECTED = {_cols.title: '标题', _cols.desc: '描述', _cols.price: '价格',
                _cols.local_price: '展示价', _cols.stock: '库存', _cols.main: '主图',
                _cols.video: '视频', _cols.variant: '变种'}
    bad = [(c, kw, str(ws.cell(1, c).value or '')) for c, kw in EXPECTED.items()
           if kw not in str(ws.cell(1, c).value or '')]
    if bad:
        print(f"\n❌ 表头结构与预期不符（TikTok 可能改了导出格式）：", flush=True)
        for c, kw, hv in bad:
            print(f"    列{c}: 期望含'{kw}'，实际'{hv}'", flush=True)
        print("已跳过该文件以防止处理错列。请检查表格或更新脚本列号映射。", flush=True)
        raise ValueError("表头结构与预期不符，已跳过")  # 批量模式下跳过该文件继续队列

    status = StatusReporter(tp)

    # ---- 一次性读全部数据到内存（避免后续重复 openpyxl 遍历）----
    total_rows = ws.max_row - 1
    print(f"读取 {total_rows} 行到内存...", flush=True)
    titles_mem = [''] * total_rows; descs_mem = [''] * total_rows
    mains_mem = [''] * total_rows; variants_mem = [''] * total_rows
    videos_mem = [''] * total_rows
    atts_mem = [[''] * len(_cols.att) for _ in range(total_rows)]
    url_map = {}; row_images = {}
    for off, r in enumerate(range(2, ws.max_row + 1)):
        titles_mem[off] = str(ws.cell(r, _cols.title).value or '').strip()
        descs_mem[off] = str(ws.cell(r, _cols.desc).value or '')
        mains_mem[off] = str(ws.cell(r, _cols.main).value or '').strip()
        for ai, ac in enumerate(_cols.att):
            atts_mem[off][ai] = str(ws.cell(r, ac).value or '').strip()
        variants_mem[off] = str(ws.cell(r, _cols.variant).value or '').strip()
        videos_mem[off] = str(ws.cell(r, _cols.video).value or '')
        if '<img' in descs_mem[off]:
            urls = [m.group(1) for m in IMG_TAG_RE.finditer(descs_mem[off])]
            if urls: row_images[r] = urls
        for val, kind in [(mains_mem[off],'main'),(variants_mem[off],'variant')]:
            if val and val.startswith('http'):
                url_map.setdefault(val,{'main':[],'att':[],'variant':[]})[kind].append(r)
        for ai, val in enumerate(atts_mem[off]):
            if val and val.startswith('http'):
                url_map.setdefault(val,{'main':[],'att':[],'variant':[]})['att'].append((r,_cols.att[ai]))
    all_urls = list(url_map.keys())
    print(f"唯一图片(单元格): {len(all_urls)}, 数据行: {total_rows}, 描述含img行: {len(row_images)}", flush=True)
    wb.close(); del wb

    # ---- Pipeline stages with error isolation ----
    metrics = PipelineMetrics()

    def _run_stage(name, fn, *args):
        t0 = time.time()
        try:
            result = fn(*args)
            metrics.record_stage(name, time.time() - t0, 1)
            return result
        except Exception as e:
            _log.error(f"阶段 [{name}] 失败", error=str(e), exc_info=True)
            # Write error to status before re-raising
            try:
                with open(status.status_path, 'w', encoding='utf-8') as f:
                    json.dump({'stage': '错误', 'error': f'{name}: {e}',
                               'request_id': _get_request_id()}, f, ensure_ascii=False)
            except Exception:
                pass
            raise

    cache, _save_cache = _run_stage('缓存初始化', _setup_cache, tp)
    review_results, unreviewed, to_regen, to_delete_att = \
        _run_stage('图审', _stage_review, status, all_urls, url_map, cache, _save_cache)
    gen_results = _run_stage('图生图', _stage_generate, status, to_regen, url_map, cache, _save_cache,
                             mains_mem, variants_mem)

    cleared_att_urls = set(to_delete_att)
    att_del = _run_stage('附图清空', _stage_clear_attachments, status, to_delete_att, url_map, total_rows, atts_mem)
    title_rule_cnt = _run_stage('品牌清洗', _stage_strip_brands, total_rows, titles_mem)
    t_changed = _run_stage('标题翻译', _stage_translate_titles, status, total_rows, titles_mem, cache, _save_cache)
    changed = _run_stage('描述清洗', _stage_clean_descs, status, total_rows, descs_mem, cache, _save_cache)
    tr_changed = _run_stage('描述翻译', _stage_translate_descs, status, total_rows, descs_mem, cache, _save_cache)
    embed_cnt = _run_stage('嵌入图片', _stage_embed_images, status, total_rows, descs_mem, row_images, gen_results, cleared_att_urls)
    inject_cnt, dedup_cnt = _run_stage('注入图片', _stage_inject_images, status, total_rows,
                                        mains_mem, atts_mem, descs_mem, gen_results, cleared_att_urls)
    vid_del = _run_stage('视频清空', _stage_clear_video, status, total_rows, videos_mem)
    template_del = _run_stage('模板清除', _stage_clear_templates, total_rows, descs_mem)
    output_path = _run_stage('保存', _stage_finalize, status, tp, adapter, total_rows,
                             titles_mem, descs_mem, mains_mem, atts_mem, variants_mem, videos_mem, cache)

    # Write metrics to status
    try:
        with open(status.status_path, 'r', encoding='utf-8') as f:
            st = json.load(f)
        st['metrics'] = metrics.to_dict()
        with open(status.status_path, 'w', encoding='utf-8') as f:
            json.dump(st, f, ensure_ascii=False, indent=2)
    except Exception:
        pass

    return output_path


def _validate_output(output_path, adapter, cache):
    """输出校验：打开输出文件检查关键列，返回 {"passed": bool, "warnings": [...]}。"""
    try:
        wb = openpyxl.load_workbook(output_path, data_only=True)
        ws = wb.active
    except Exception:
        return {"passed": False, "warnings": ["无法打开输出文件进行校验"]}
    warnings = []
    cols = adapter.cols if adapter else {}
    title_col = cols.get('title', 2)
    main_col = cols.get('main_image', 18)
    price_col = cols.get('price', 15)
    video_col = cols.get('video', 27)
    total = ws.max_row - 1

    # 标题：中文残留率 ≤ 5%
    cn_count = 0
    for r in range(2, ws.max_row + 1):
        t = str(ws.cell(r, title_col).value or '')
        if _CHINESE_RE.search(t):
            cn_count += 1
    cn_rate = cn_count / max(total, 1)
    if cn_rate > 0.05:
        warnings.append(f"标题中文残留 {cn_count}/{total} ({cn_rate:.0%})，阈值 5%")
    elif cn_count > 0:
        warnings.append(f"标题中文残留 {cn_count}/{total} ({cn_rate:.0%})，在容忍范围内")

    # 主图：检查是否有 URL（非空）
    empty_main = sum(1 for r in range(2, ws.max_row + 1)
                     if not str(ws.cell(r, main_col).value or '').startswith('http'))
    if empty_main > total * 0.1:
        warnings.append(f"主图列 {empty_main}/{total} 行无有效 URL")

    # 价格列：检查是否改名为"本地展示价"
    price_header = str(ws.cell(1, price_col).value or '')
    if '展示价' not in price_header:
        warnings.append(f"价格列未改名为本地展示价，当前: {price_header}")

    # 视频列：应为空
    video_filled = sum(1 for r in range(2, ws.max_row + 1)
                       if str(ws.cell(r, video_col).value or '').strip())
    if video_filled > 0:
        warnings.append(f"视频列未清空，{video_filled} 行仍有内容")

    wb.close()

    passed = len([w for w in warnings if '容忍' not in w]) == 0
    result = {"passed": passed, "warnings": warnings}
    if warnings:
        print(f"\n{'='*50}\n输出校验 {'✅ 通过' if passed else '⚠️ 发现问题'}:", flush=True)
        for w in warnings:
            print(f"  {'⚠️' if '容忍' not in w else 'ℹ️'} {w}", flush=True)
        print('='*50, flush=True)
    return result


# === Pipeline stage functions (extracted from _main) ===

def _stage_clear_attachments(status, to_delete_att, url_map, total_rows, atts_mem):
    """清空水印附图单元格。返回删除计数。"""
    status.start_stage('附图清空', len(to_delete_att))
    att_del = 0
    for url in to_delete_att:
        for row_idx, col in url_map.get(url, {}).get('att', []):
            off = row_idx - 2
            if off < total_rows:
                for ai, ac in enumerate(_cols.att):
                    if ac == col:
                        atts_mem[off][ai] = ''
                        att_del += 1
                        break
    print(f"附图清空: {att_del} 个单元格")
    return att_del


def _stage_clear_video(status, total_rows, videos_mem):
    """清空视频连接列。返回清空计数。"""
    status.start_stage('视频+模板图清理', total_rows)
    vid_del = 0
    for off in range(total_rows):
        if videos_mem[off]:
            videos_mem[off] = ''
            vid_del += 1
    status.update(total_rows)
    print(f"视频连接清空: {vid_del} 行", flush=True)
    return vid_del


def _stage_clear_templates(total_rows, descs_mem):
    """删除 pushauction/ibay365 模板图。返回删除计数。"""
    import re
    template_pattern = re.compile(
        r'\s*(?:<br\s*/?>\s*)?<img[^>]*src=["\'][^"\']*(?:pushauction|ibay365)[^"\']*["\'][^>]*>(?:\s*<br\s*/?>)?',
        re.IGNORECASE
    )
    template_del = 0
    for off in range(total_rows):
        dv = descs_mem[off]
        if dv:
            new_dv = template_pattern.sub('', dv)
            if new_dv != dv:
                descs_mem[off] = new_dv
                template_del += 1
    print(f"模板图清除(pushauction/ibay365): {template_del} 行", flush=True)
    return template_del


def _stage_strip_brands(total_rows, titles_mem):
    """规则清洗标题中的品牌名/平台名。返回修改计数。"""
    cnt = 0
    for off in range(total_rows):
        old = titles_mem[off]
        new = rule_strip_brands(old)
        if new != old:
            titles_mem[off] = new
            cnt += 1
    print(f"标题规则清洗(品牌): {cnt} 行", flush=True)
    return cnt


def _stage_embed_images(status, total_rows, descs_mem, row_images, gen_results, cleared_att_urls):
    """替换描述中 __IMG__ 占位符为真正的 <img> 标签。"""
    status.start_stage('嵌入+注入图片', total_rows)
    embed_cnt = 0
    for off in range(total_rows):
        v = descs_mem[off]
        if v and '__IMG__' in v:
            r = off + 2
            urls = row_images.get(r, [])
            if urls:
                new_html = embed_new_images_in_desc(v, urls, gen_results, cleared_att_urls)
                if new_html != v:
                    descs_mem[off] = new_html
                    embed_cnt += 1
    print(f"描述嵌入新图URL: {embed_cnt} 行", flush=True)
    return embed_cnt


def _stage_finalize(status, tp, adapter, total_rows,
                    titles_mem, descs_mem, mains_mem, atts_mem,
                    variants_mem, videos_mem, cache):
    """价格列改名 + 删除本地展示价列 + 写回 Excel + 保存 + 校验。返回输出路径。"""
    status.start_stage('价格列+保存', 1)
    wb2 = openpyxl.load_workbook(tp)
    ws2 = wb2[adapter.sheet_name] if adapter and adapter.sheet_name and adapter.sheet_name in wb2.sheetnames else wb2.active

    ws2.cell(1, _cols.price).value = '本地展示价'
    ws2.delete_cols(_cols.local_price)
    print(f"价格列(col{_cols.price})改名'本地展示价'；删除原本地展示价列(col{_cols.local_price})；库存不动", flush=True)

    print("写回 Excel...", flush=True)
    for off in range(total_rows):
        r = off + 2
        ws2.cell(r, _cols.title).value = titles_mem[off]
        ws2.cell(r, _cols.desc).value = descs_mem[off]
        ws2.cell(r, _cols.main).value = mains_mem[off]
        ws2.cell(r, _cols.variant).value = variants_mem[off]
        ws2.cell(r, _cols.video).value = videos_mem[off]
        for ai, ac in enumerate(_cols.att):
            ws2.cell(r, ac).value = atts_mem[off][ai]

    save = os.path.splitext(tp)[0] + '_cleaned.xlsx'
    if os.path.exists(save):
        save = os.path.splitext(tp)[0] + time.strftime('_cleaned_%H%M%S.xlsx')
        print(f"⚠️ 输出文件已存在，改存: {save}", flush=True)
    wb2.save(save)
    wb2.close()

    validation = _validate_output(save, adapter, cache)
    status.finish(save)
    if validation:
        try:
            with open(status.status_path, 'r', encoding='utf-8') as f:
                st = json.load(f)
            st['validation'] = validation
            with open(status.status_path, 'w', encoding='utf-8') as f:
                json.dump(st, f, ensure_ascii=False, indent=2)
        except Exception as e:
            _log.warn("校验结果写入 status 失败", error=str(e))
    print(f"完成! 保存: {save}", flush=True)
    return save


def _stage_inject_images(status, total_rows, mains_mem, atts_mem, descs_mem,
                         gen_results, cleared_att_urls):
    """注入主图+附图新 URL 到描述（图生图新 URL 优先），去重正文中已有 img。"""
    inject_cnt = 0; dedup_cnt = 0
    for off in range(total_rows):
        r = off + 2
        main_url = mains_mem[off]
        if not main_url or not str(main_url).startswith('http'):
            continue
        att_urls = [str(atts_mem[off][ai]).strip()
                    for ai in range(len(_cols.att))
                    if atts_mem[off][ai] and str(atts_mem[off][ai]).startswith('http')]
        final_main = gen_results.get(str(main_url).strip(), str(main_url).strip())
        final_atts = [gen_results.get(u, u) for u in att_urls if u not in cleared_att_urls]

        # 构建旧→新 URL 映射，正则替换避免部分匹配
        att_old_to_new = {u: gen_results[u] for u in att_urls if u in gen_results}
        orig_main = str(main_url).strip()
        main_old_to_new = {orig_main: gen_results[orig_main]} if orig_main in gen_results else {}
        desc_v = descs_mem[off]
        for old_url, new_url in {**main_old_to_new, **att_old_to_new}.items():
            desc_v = re.sub(re.escape(old_url), new_url, desc_v)

        # 去重删除正文中已有的产品图 img 标签
        urls_to_remove = {orig_main} | set(att_urls)
        urls_to_remove.update(gen_results.get(u, u) for u in att_urls if u in gen_results)
        if orig_main in gen_results:
            urls_to_remove.add(gen_results[orig_main])
        if urls_to_remove:
            url_alt = '|'.join(re.escape(u) for u in urls_to_remove)
            pattern = re.compile(
                r'\s*(?:<br\s*/?>\s*)?<img[^>]*src=["\'](' + url_alt + r')["\'][^>]*>(?:\s*<br\s*/?>)?',
                re.IGNORECASE)
            new_desc = pattern.sub('', desc_v)
            if new_desc != desc_v:
                dedup_cnt += 1
            desc_v = new_desc

        img_block = f'<img src="{final_main}"/>' + ''.join(f'<img src="{u}"/>' for u in final_atts)
        descs_mem[off] = img_block + desc_v
        inject_cnt += 1
        if r % 40 == 0:
            status.update(off + 1)
    status.update(total_rows)
    print(f"描述注入主图+附图URL: {inject_cnt} 行 | 去重删除正文重复img: {dedup_cnt} 处", flush=True)
    return inject_cnt, dedup_cnt


def _stage_translate_titles(status, total_rows, titles_mem, cache, _save_cache):
    """标题翻译：批量翻译 + 个体 fallback + 中文校验。返回修改行数。"""
    title_map = {}
    for off in range(total_rows):
        v = titles_mem[off]
        if v:
            title_map.setdefault(v, []).append(off + 2)
    status.start_stage('标题清洗+翻译', len(title_map))
    t0 = time.time()
    cached_titles = cache.get('title_translations', {})
    new_titles = [t for t in title_map if t not in cached_titles]
    if new_titles:
        new_results = batch_translate_texts(new_titles)
        cached_titles.update(new_results)
        print(f"标题翻译(batch): {time.time()-t0:.1f}s | {len(new_titles)} 新 + {len(title_map)-len(new_titles)} 缓存命中", flush=True)
    untranslated = [t for t in title_map if t not in cached_titles or _CHINESE_RE.search(cached_titles.get(t, ''))]
    if untranslated:
        print(f"标题翻译(个体fallback): {len(untranslated)} 条重试单条翻译({TEXT_CONCURRENCY}并发)...", flush=True)
        done = 0
        with ThreadPoolExecutor(max_workers=min(TEXT_CONCURRENCY, len(untranslated))) as pool:
            futures = {pool.submit(translate_text, t, TITLE_TRANSLATE_PROMPT): t for t in untranslated}
            for future in as_completed(futures):
                t = futures[future]
                try:
                    vn = future.result()
                    if vn and vn != t and not _CHINESE_RE.search(vn):
                        cached_titles[t] = vn
                except Exception as e:
                    _log.warn("标题个体fallback 异常", error=str(e))
                done += 1
                if done % 20 == 0:
                    print(f"  个体fallback: {done}/{len(untranslated)}", flush=True)
    elif not new_titles:
        print(f"标题翻译: 全部 {len(title_map)} 缓存命中", flush=True)
    cache['title_translations'] = cached_titles
    _save_cache(cache)
    t_changed = 0
    for src, rows in title_map.items():
        translated = cached_titles.get(src, src)
        if translated != src:
            for r in rows:
                titles_mem[r - 2] = translated
                t_changed += 1
    print(f"标题翻译: {time.time()-t0:.1f}s | 翻译: {t_changed} 行 ({len(title_map)} 唯一)", flush=True)
    return t_changed


def _stage_clean_descs(status, total_rows, descs_mem, cache, _save_cache):
    """描述 AI 清洗：去品牌名/退货政策/运费，保留产品特性。返回修改行数。"""
    desc_map = {}
    for off in range(total_rows):
        dv = descs_mem[off]
        if dv and dv.strip():
            desc_map.setdefault(dv.strip(), []).append(off + 2)
    status.start_stage('描述AI清洗', len(desc_map))
    t0 = time.time()
    cached_desc = cache.get('desc_cleaned', {})
    new_descs = [t for t in desc_map if t not in cached_desc]
    if new_descs:
        new_results = batch_clean_texts(new_descs)
        cached_desc.update(new_results)
        print(f"描述清洗(新): {time.time()-t0:.1f}s | {len(new_descs)} 新 + {len(desc_map)-len(new_descs)} 缓存命中", flush=True)
    else:
        print(f"描述清洗: 全部 {len(desc_map)} 缓存命中", flush=True)
    cache['desc_cleaned'] = cached_desc
    _save_cache(cache)
    changed = 0
    for src, rows in desc_map.items():
        cleaned = cached_desc.get(src, src)
        if cleaned != src:
            for r in rows:
                descs_mem[r - 2] = cleaned
                changed += 1
    print(f"描述清洗: {time.time()-t0:.1f}s | 修改: {changed} 行 ({len(desc_map)} 唯一)", flush=True)
    return changed


def _stage_translate_descs(status, total_rows, descs_mem, cache, _save_cache):
    """描述翻译越南语：批量翻译 + 个体 fallback。返回修改行数。"""
    trans_map = {}
    for off in range(total_rows):
        v = descs_mem[off]
        if v and v.strip():
            trans_map.setdefault(v.strip(), []).append(off + 2)
    status.start_stage('描述翻译', len(trans_map))
    t0 = time.time()
    cached_trans = cache.get('desc_translations', {})
    new_trans = [t for t in trans_map if t not in cached_trans]
    if new_trans:
        new_results = batch_translate_texts(new_trans)
        cached_trans.update(new_results)
        print(f"描述翻译(batch): {time.time()-t0:.1f}s | {len(new_trans)} 新 + {len(trans_map)-len(new_trans)} 缓存命中", flush=True)
    else:
        print(f"描述翻译: 全部 {len(trans_map)} 缓存命中", flush=True)
    missed = [t for t in trans_map if t not in cached_trans]
    if missed:
        print(f"描述翻译(个体fallback): {len(missed)} 条重试...", flush=True)
        done = 0
        with ThreadPoolExecutor(max_workers=min(TEXT_CONCURRENCY, len(missed))) as pool:
            futures = {pool.submit(translate_text, t, TRANSLATE_PROMPT): t for t in missed}
            for future in as_completed(futures):
                t = futures[future]
                try:
                    vn = future.result()
                    if vn and vn != t:
                        cached_trans[t] = vn
                except Exception as e:
                    _log.warn("描述个体fallback 异常", error=str(e))
                done += 1
                if done % 20 == 0:
                    print(f"  个体fallback: {done}/{len(missed)}", flush=True)
    cache['desc_translations'] = cached_trans
    _save_cache(cache)
    tr_changed = 0
    for src, rows in trans_map.items():
        translated = cached_trans.get(src, src)
        if translated != src:
            for r in rows:
                descs_mem[r - 2] = translated
                tr_changed += 1
    print(f"描述翻译: {time.time()-t0:.1f}s | 翻译: {tr_changed} 行", flush=True)
    return tr_changed


def _setup_cache(tp):
    """初始化全局哈希缓存（断点续跑）。返回 (cache_dict, save_fn)。"""
    _CACHE_DIR = os.path.join(os.environ.get('CROSSPILOT_DATA_DIR',
        os.path.dirname(os.path.dirname(os.path.abspath(__file__)))), 'data', 'cache')
    os.makedirs(_CACHE_DIR, exist_ok=True)
    _now = time.time()
    for _cf in glob.glob(os.path.join(_CACHE_DIR, '*.json')):
        try:
            if os.path.getmtime(_cf) < _now - 30 * 86400:
                os.remove(_cf)
        except OSError:
            pass
    _FILE_HASH = hashlib.file_digest(open(tp, 'rb'), 'sha256').hexdigest()[:16] \
        if hasattr(hashlib, 'file_digest') else hashlib.sha256(open(tp, 'rb').read()).hexdigest()[:16]
    CACHE_PATH = os.path.join(_CACHE_DIR, f'{_FILE_HASH}.json')

    def _load():
        try:
            c = json.load(open(CACHE_PATH, encoding='utf-8'))
            print(f"缓存命中: {len(c.get('review_results',{}))} 图审 + {len(c.get('gen_results',{}))} 生图 + "
                  f"{len(c.get('title_translations',{}))} 标题翻译 + {len(c.get('desc_cleaned',{}))} 描述清洗 + "
                  f"{len(c.get('desc_translations',{}))} 描述翻译", flush=True)
            return c
        except Exception:
            return {'review_results':{},'gen_results':{},'title_translations':{},'desc_cleaned':{},'desc_translations':{}}

    def _save(c):
        try:
            tmp = CACHE_PATH + '.tmp'
            with open(tmp, 'w', encoding='utf-8') as f:
                json.dump(c, f, ensure_ascii=False)
            os.replace(tmp, CACHE_PATH)
        except Exception as e:
            _log.warn("缓存保存失败", error=str(e))

    return _load(), _save


def _stage_review(status, all_urls, url_map, cache, _save_cache):
    """图审：两轮并发（100→10），三级 fallback 链。返回 (review_results, unreviewed, to_regen, to_delete_att)。"""
    cached_review = cache.get('review_results', {})
    to_review = [u for u in all_urls if url_map.get(u, {}).get('main')
                 or url_map.get(u, {}).get('variant') or url_map.get(u, {}).get('att')]
    review_results = {u: cached_review[u] for u in to_review if u in cached_review}
    to_review_new = [u for u in to_review if u not in cached_review]
    if review_results:
        print(f"图审缓存命中: {len(review_results)}/{len(to_review)} 张，剩余 {len(to_review_new)} 张待审", flush=True)
    status.start_stage('MiMo图审', len(to_review_new))

    t0 = time.time(); reviewed = 0
    with ThreadPoolExecutor(max_workers=MIMO_CONCURRENCY) as pool:
        futures = {pool.submit(review_single, url): url for url in to_review_new}
        for future in as_completed(futures):
            review_results[futures[future]] = future.result()
            reviewed += 1; status.update(reviewed)

    failed_urls = [u for u, r in review_results.items() if r is None]
    if failed_urls:
        print(f"\n[{time.strftime('%H:%M:%S')}] 第一轮失败 {len(failed_urls)} 张，第二轮低并发重审...", flush=True)
        with ThreadPoolExecutor(max_workers=10) as pool:
            futures = {pool.submit(review_single, url): url for url in failed_urls}
            for future in as_completed(futures):
                url = futures[future]; r = future.result()
                if r is not None: review_results[url] = r

    cache['review_results'].update({u: r for u, r in review_results.items() if r is not None})
    _save_cache(cache)

    unreviewed = [u for u, r in review_results.items() if r is None]
    for u in unreviewed: review_results[u] = False
    if unreviewed:
        print(f"⚠️ 四级链后仍 {len(unreviewed)} 张未审出（保留原图）", flush=True)
    unreviewed_rate = len(unreviewed) / len(to_review) if to_review else 0
    if unreviewed_rate > 0.1:
        print(f"\n⚠️ 图审失败率 {len(unreviewed)}/{len(to_review)} ({unreviewed_rate*100:.0f}%)，未审图片保留原图", flush=True)
    try:
        with open(status.status_path, 'r', encoding='utf-8') as f:
            _st = json.load(f)
        _st['unreviewed'] = unreviewed[:100]
        _st['unreviewed_count'] = len(unreviewed)
        _st['unreviewed_rate'] = round(unreviewed_rate, 3)
        with open(status.status_path, 'w', encoding='utf-8') as f:
            json.dump(_st, f, ensure_ascii=False, indent=2)
    except Exception as e:
        _log.warn("未审清单写入 status 失败", error=str(e))

    to_regen = [u for u, r in review_results.items() if r
                and (url_map.get(u, {}).get('main') or url_map.get(u, {}).get('variant'))]
    to_delete_att = [u for u, r in review_results.items() if r and url_map.get(u, {}).get('att')]
    n_var = sum(1 for u in to_regen if url_map.get(u, {}).get('variant'))
    print(f"MiMo审: {time.time()-t0:.1f}s | 水印: {len(to_regen)}(含变种{n_var}) | 附图清: {len(to_delete_att)}", flush=True)
    return review_results, unreviewed, to_regen, to_delete_att


def _stage_generate(status, to_regen, url_map, cache, _save_cache, mains_mem, variants_mem):
    """图生图：wan→doubao 二级 fallback，缓存复用。返回 gen_results dict。"""
    cached_gen = cache.get('gen_results', {})
    gen_results = {u: cached_gen[u] for u in to_regen if u in cached_gen}
    to_gen_new = [u for u in to_regen if u not in cached_gen]
    if gen_results:
        print(f"生图缓存命中: {len(gen_results)}/{len(to_regen)} 张，剩余 {len(to_gen_new)} 张待生成", flush=True)
    if to_gen_new:
        t0 = time.time()
        status.start_stage('图生图', len(to_gen_new))
        gen_done = 0
        with ThreadPoolExecutor(max_workers=GEN_CONCURRENCY) as pool:
            futures = {pool.submit(_gen_image, _http, url): url for url in to_gen_new}
            for future in as_completed(futures):
                url = futures[future]; new_url = future.result()
                gen_done += 1
                if new_url: gen_results[url] = new_url
                status.update(gen_done, force=True)
        print(f"图生图: {time.time()-t0:.1f}s | 成功: {len(gen_results)}/{len(to_regen)}", flush=True)
        cache['gen_results'].update(gen_results)
        _save_cache(cache)
    main_replaced = variant_replaced = 0
    for old_url, new_url in gen_results.items():
        for row_idx in url_map.get(old_url, {}).get('main', []):
            mains_mem[row_idx - 2] = new_url; main_replaced += 1
        for row_idx in url_map.get(old_url, {}).get('variant', []):
            variants_mem[row_idx - 2] = new_url; variant_replaced += 1
    print(f"主图替换: {main_replaced} | 变种图替换: {variant_replaced}")
    return gen_results


def main():
    """异常保护：崩溃时在 status.json 写错误状态，避免像还在跑"""
    try:
        _main()
    except SystemExit:
        raise
    except Exception as e:
        try:
            status = StatusReporter(TABLE_PATH)
            with open(status.status_path, 'w', encoding='utf-8') as f:
                json.dump({'stage': '错误', 'error': f'{type(e).__name__}: {e}',
                           'updated_at': time.strftime('%Y-%m-%d %H:%M:%S')},
                          f, ensure_ascii=False, indent=2)
        except Exception as e:
            _log.warn("status.json 写入失败", error=str(e))
            pass
        print(f"\n❌ 运行失败: {type(e).__name__}: {e}", flush=True)
        raise


# ===== 批量文本处理：一次 API 调用处理多个文本，大幅减少请求数 =====
_BATCH_SIZE = 25  # 每批处理的文本数


def _parse_batch_response(raw, expected_count):
    """解析批量 API 返回的 JSON 数组（委托 TranslationService）。"""
    return _translate_svc._parse_batch_response(raw)

def _process_batch(batch, prompt_template, label):
    """处理一个批次（委托 TranslationService）。"""
    return _translate_svc._process_batch(batch, prompt_template, label)

def _batch_process(texts, prompt_template, label):
    """通用批量文本处理（委托 TranslationService）。"""
    return _translate_svc.batch_process(texts, prompt_template, label)

def batch_translate_texts(texts):
    """批量翻译文本（委托 TranslationService）。"""
    return _translate_svc.batch_translate(texts)

def batch_clean_texts(texts):
    """批量清洗文本（委托 TranslationService）。"""
    return _translate_svc.batch_clean(texts)


if __name__ == '__main__':
    main()
