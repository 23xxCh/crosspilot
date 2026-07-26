"""Test fixtures: shared across all pipeline tests."""
import atexit
import os, sys, json, shutil, tempfile
import pytest

# Tests must never touch the real task database, uploads, cache, or API keys.
_TEST_RUNTIME_DIR = tempfile.mkdtemp(prefix='crosspilot-tests-')
_TEST_KEYS_PATH = os.path.join(_TEST_RUNTIME_DIR, 'keys.json')
os.environ['CROSSPILOT_DATA_DIR'] = _TEST_RUNTIME_DIR
os.environ['CROSSPILOT_KEYS_PATH'] = _TEST_KEYS_PATH
with open(_TEST_KEYS_PATH, 'w', encoding='utf-8') as _keys_file:
    json.dump({
        'text_provider': 'deepseek',
        'vision_provider': 'agnes',
        'image_gen_provider': 'agnes',
        'deepseek_key': 'test-deepseek-key',
        'agnes_key': 'test-agnes-key'
    }, _keys_file)
atexit.register(lambda: shutil.rmtree(_TEST_RUNTIME_DIR, ignore_errors=True))

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'scripts'))
import process_ebay_tk
import openpyxl


@pytest.fixture
def sample_xlsx_path():
    """复制迷你表到临时路径，测试完清理所有 cache/status/output 副产物。"""
    import glob
    src = os.path.join(os.path.dirname(__file__), '..', 'data', 'sample_small.xlsx')
    dst = os.path.join(os.path.dirname(__file__), '..', 'data', '_test_sample.xlsx')
    shutil.copy2(src, dst)
    yield dst
    for pattern in ['_cache.json', '_status.json', '_cleaned.xlsx', '_cleaned_*']:
        for f in glob.glob(os.path.splitext(dst)[0] + '_cleaned_*'):
            os.remove(f) if os.path.exists(f) else None
        p = os.path.splitext(dst)[0] + '_cache.json'
        if os.path.exists(p): os.remove(p)
        p = os.path.splitext(dst)[0] + '_status.json'
        if os.path.exists(p): os.remove(p)
        p = os.path.splitext(dst)[0] + '_cleaned.xlsx'
        if os.path.exists(p): os.remove(p)
    if os.path.exists(dst):
        os.remove(dst)


@pytest.fixture
def invalid_xlsx_path():
    """创一个不识格式的 xlsx（表头全陌生）。"""
    dst = os.path.join(os.path.dirname(__file__), '..', 'data', '_test_invalid.xlsx')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Sheet1'
    ws.cell(1, 1, 'Foo')
    ws.cell(1, 2, 'Bar')
    wb.save(dst)
    yield dst
    if os.path.exists(dst):
        os.remove(dst)


@pytest.fixture
def header_mismatch_xlsx_path():
    """创一个能识别格式但 header 不匹配的 xlsx（col 15 不是"价格"）。"""
    dst = os.path.join(os.path.dirname(__file__), '..', 'data', '_test_hdr_bad.xlsx')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'tiktok_chanpin_'
    # 放正确的识别 header 使 adapter 能 detect
    ws.cell(1, 2, '产品标题')
    ws.cell(1, 3, 'Tiktok产品描述')
    ws.cell(1, 18, '主图(url)地址')
    ws.cell(1, 36, '来源Url')
    # col 15 放错（应是"价格"）
    ws.cell(1, 15, 'NOT_A_PRICE_COLUMN')
    wb.save(dst)
    yield dst
    if os.path.exists(dst):
        os.remove(dst)
