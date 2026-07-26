"""从全表截前 N 行造迷你测试表，省 API 额度。

用法：uv run python -u web/make_sample.py [行数] [全表路径]
默认 5 行，从项目根第一个 tiktok_chanpin_*.xlsx 截。
输出 data/sample_small.xlsx。
"""
import os, sys, glob
import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))


def main():
    n = int(sys.argv[1]) if len(sys.argv) > 1 else 5
    src = sys.argv[2] if len(sys.argv) > 2 else (
        sorted(glob.glob(os.path.join(ROOT, 'tiktok_chanpin_*.xlsx')))[0])
    out = os.path.join(ROOT, 'data', 'sample_small.xlsx')

    src_wb = openpyxl.load_workbook(src, data_only=True)
    ws = src_wb.active
    out_wb = openpyxl.Workbook()
    out_wb.remove(out_wb.active)
    out_ws = out_wb.create_sheet('tiktok_chanpin_')

    # 复制表头 + 前 n 数据行（含所有列、样式）
    max_col = ws.max_column
    for r in range(1, n + 2):  # 1 表头 + n 行
        for c in range(1, max_col + 1):
            src_cell = ws.cell(r, c)
            dst = out_ws.cell(r, c, src_cell.value)
            if src_cell.has_style:
                dst.font = src_cell.font.copy()
                dst.alignment = src_cell.alignment.copy()

    out_wb.save(out)
    print(f"OK 截表 {n} 行 x {max_col} 列 -> {out}")


if __name__ == '__main__':
    main()
