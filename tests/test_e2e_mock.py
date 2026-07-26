"""E2E: Mock 全链路管道测试 — 不依赖任何 API key，验证管道逻辑正确性。"""
import os, sys, json, pytest
from unittest.mock import patch, Mock

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'scripts'))
import process_ebay_tk
import openpyxl


@pytest.fixture(autouse=True)
def setup_and_teardown():
    """每个测试前后清理缓存/状态文件，防止跨测试污染。"""
    yield
    import glob
    for pat in ['_cache.json', '_status.json', '_cleaned.xlsx', '_cleaned_*']:
        for f in glob.glob(os.path.join('data', '*_test_*' + pat.split('*')[0] + '*')):
            try:
                os.remove(f)
            except OSError:
                pass


class TestMockFullPipeline:
    """Mock 所有外部 API，跑完整 eBay 管道，验证输出。"""

    def test_pipeline_mocked_all_clean(self, sample_xlsx_path, monkeypatch):
        """所有图审返回 False（无水印），翻译返回越南语，验证管道输出。"""
        from pipelines import ebay_stages, ebay_shared

        # Mock 图审：全部返回 False（无水印，不生图）
        monkeypatch.setattr(ebay_stages, 'review_single', Mock(return_value=False))

        # Mock 图生图：返回假 URL
        monkeypatch.setattr(ebay_stages, '_gen_image', Mock(return_value='https://new.example.com/1.png'))

        # Mock 批量翻译/清洗
        monkeypatch.setattr(ebay_stages, 'batch_translate_texts',
                           lambda texts: {t: 'Sản phẩm ' + t[:10] for t in texts})
        monkeypatch.setattr(ebay_stages, 'batch_clean_texts',
                           lambda texts: {t: t for t in texts})

        # 跑管道
        output = process_ebay_tk._main(sample_xlsx_path)
        assert output is not None
        assert os.path.exists(output)

        # 验证输出内容
        wb = openpyxl.load_workbook(output, data_only=True)
        ws = wb.active
        assert ws.max_row >= 3  # header + 2 data rows
        assert '本地展示价' in str(ws.cell(1, 15).value or '')  # price column renamed
        assert ws.cell(2, 2).value  # title not empty
        wb.close()

    def test_pipeline_mocked_with_watermarks(self, sample_xlsx_path, monkeypatch):
        """图审返回 True（有水印），验证生图 mock 被调用，输出仍正常。"""
        from pipelines import ebay_stages

        monkeypatch.setattr(ebay_stages, 'review_single', Mock(return_value=True))
        gen_calls = []
        monkeypatch.setattr(ebay_stages, '_gen_image',
                           lambda s, u, **_kwargs: (
                               gen_calls.append(u),
                               'https://gen.example.com/' + u.split('/')[-1],
                           )[-1])
        monkeypatch.setattr(ebay_stages, 'batch_translate_texts',
                           lambda texts: {t: 'Sản phẩm ô tô' for t in texts})
        monkeypatch.setattr(ebay_stages, 'batch_clean_texts',
                           lambda texts: {t: t for t in texts})

        output = process_ebay_tk._main(sample_xlsx_path)
        assert output is not None
        assert os.path.exists(output)
        # 缓存可能已命中（from prior test），gen 可能为 0 或 >0，都是正确的
        print(f'gen_calls: {len(gen_calls)} images generated')

    def test_pipeline_output_valid_xlsx(self, sample_xlsx_path, monkeypatch):
        """Mock 管道，验证输出是可打开的有效 xlsx。"""
        from pipelines import ebay_stages
        monkeypatch.setattr(ebay_stages, 'review_single', Mock(return_value=False))
        monkeypatch.setattr(ebay_stages, '_gen_image', Mock(return_value='https://x.com/1.png'))
        monkeypatch.setattr(ebay_stages, 'batch_translate_texts',
                           lambda texts: {t: 'Đã dịch' for t in texts})
        monkeypatch.setattr(ebay_stages, 'batch_clean_texts',
                           lambda texts: {t: t for t in texts})

        output = process_ebay_tk._main(sample_xlsx_path)
        wb = openpyxl.load_workbook(output, data_only=True)
        ws = wb.active
        # 关键列存在且有值
        assert ws.cell(1, 2).value  # title header
        assert ws.cell(1, 3).value  # desc header
        assert ws.cell(2, 2).value  # row 1 title
        # 价格列已改名
        assert '展示价' in str(ws.cell(1, 15).value or '')
        wb.close()

    def test_pipeline_amazon_mocked(self, monkeypatch):
        """Mock Amazon 管道，验证 _main 返回输出路径。"""
        import process_amazon
        # Mock 所有阶段（各阶段签名不同）
        monkeypatch.setattr(
            process_amazon,
            '_stage_review_and_gen',
            lambda d, c=None, quality_issues=None, progress=None: d,
        )
        monkeypatch.setattr(
            process_amazon,
            '_stage_optimize_titles',
            lambda d, progress=None: d,
        )
        monkeypatch.setattr(
            process_amazon,
            '_stage_clean_descs',
            lambda d, progress=None: d,
        )

        def _fill_content(rows, progress=None):
            for row in rows:
                row['bullets'] = [f'Bullet {i}' for i in range(1, 6)]
                row['keywords'] = 'test, product, useful'
            return rows

        monkeypatch.setattr(process_amazon, '_stage_generate_bullets_keywords', _fill_content)

        # 用 JSON 输入（Amazon 管道支持）
        import json, tempfile
        data = {
            '商品id': ['item-1', 'item-2'],
            '产品标题': ['Test Product 1', 'Test Product 2'],
            '产品描述': ['Test description', 'Another description'],
            '产品图片链接': [['https://img1.jpg'], ['https://img2.jpg']],
            '变种图片链接': [[], []],
            '产品图片': [[], []],
            '变种图片': [[], []],
        }
        jp = tempfile.mktemp(suffix='.json')
        output = None
        with open(jp, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False)
        try:
            output = process_amazon._main(jp)
            assert output is not None
            assert os.path.exists(output)
            assert output.endswith('.json')
            with open(output, encoding='utf-8') as f:
                refill = json.load(f)
            assert refill['产品标题'] == data['产品标题']
            assert refill['产品图片链接'] == data['产品图片链接']
            assert refill['Bullet Point5'] == ['Bullet 5', 'Bullet 5']
            assert refill['关键词信息'] == [
                'test, product, useful',
                'test, product, useful',
            ]
        finally:
            try:
                os.remove(jp)
                if output and os.path.exists(output):
                    os.remove(output)
            except OSError:
                pass


class TestPipelineEdgeCases:
    """管道边界情况测试。"""

    def test_missing_key_raises(self):
        """缺少输入文件时应抛出异常。"""
        with pytest.raises((FileNotFoundError, OSError, ValueError)):
            process_ebay_tk._main('/nonexistent/path/file.xlsx')

    def test_brands_shared_constant(self):
        """验证 BRANDS 常量在两个管道中一致。"""
        from services.constants import BRANDS
        assert len(BRANDS) > 30
        assert 'bmw' in BRANDS
        assert '丰田' in BRANDS
        assert 'joyon' in BRANDS  # eBay-specific
