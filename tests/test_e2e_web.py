"""E2E: FastAPI TestClient 测试 Web 上传→查询→下载全流程。"""
import os, json, time, pytest
import base64
import importlib
import shutil
import subprocess

from fastapi.testclient import TestClient
from web.app import app

client = TestClient(app)


@pytest.fixture(autouse=True)
def reset_store(monkeypatch, tmp_path):
    """每个测试隔离后台队列，并清理临时 SQLite 数据。"""
    from web import jobs
    env_path = tmp_path / '.env'
    env_path.write_text('', encoding='utf-8')
    monkeypatch.setenv('CROSSPILOT_ENV', str(env_path))
    monkeypatch.setenv(
        'CROSSPILOT_PROMPT_DIR',
        str(tmp_path / 'prompts'),
    )
    monkeypatch.setenv(
        'CROSSPILOT_PROMPT_HISTORY_DIR',
        str(tmp_path / 'prompt_history'),
    )
    monkeypatch.delenv('CROSSPILOT_PROMPT_PROFILE', raising=False)
    from crosspilot.config import reload_config
    reload_config()
    from crosspilot.prompt_registry import reload_prompt_registry
    reload_prompt_registry()
    monkeypatch.setattr(jobs, 'enqueue', lambda *_args, **_kwargs: True)
    yield
    try:
        from web import store
        conn = store._get_conn()
        conn.execute("DELETE FROM tasks")
        conn.commit()
    except Exception:
        pass


class TestWebAPI:
    """Web API 端点测试（不跑真实管道）。"""

    def test_dashboard_returns_stats(self):
        resp = client.get('/api/dashboard')
        assert resp.status_code == 200
        data = resp.json()
        assert 'today_count' in data
        assert 'running_count' in data

    def test_dashboard_get_has_no_retention_side_effect(self, monkeypatch):
        from web import store

        called = []
        monkeypatch.setattr(
            store,
            'cleanup_old_tasks',
            lambda _days: called.append(True),
        )

        resp = client.get('/api/dashboard')

        assert resp.status_code == 200
        assert called == []

    def test_analytics_includes_quality_operations_summary(self, tmp_path):
        from web import store

        source = tmp_path / 'input.xlsx'
        output = tmp_path / 'out.xlsx'
        store.create('quality-pass1', 'pass.xlsx', str(source))
        store.mark_done('quality-pass1', str(output), {
            'validation': {'passed': True, 'issues': []},
            'metrics': {
                'api_success_rate': 1.0,
                'http_retries': 0,
                'quality': {'issue_count': 0},
                'concurrency': {'reductions': 0},
            },
        })
        store.create('quality-samp1', 'sample.xlsx', str(source))
        store.mark_done('quality-samp1', str(output), {
            'validation': {'passed': True, 'issues': []},
            'metrics': {
                'api_success_rate': 1.0,
                'http_retries': 4,
                'quality': {'issue_count': 0},
                'concurrency': {'reductions': 0},
            },
        })
        store.create('quality-rev01', 'review.xlsx', str(source))
        store.mark_needs_review('quality-rev01', str(output), {
            'validation': {'passed': False, 'issues': ['第 3 行 Bullet 重复']},
            'metrics': {
                'api_success_rate': 0.9,
                'http_retries': 2,
                'quality': {'issue_count': 1},
                'concurrency': {'reductions': 1},
            },
        })
        store.create('quality-fail1', 'failed.xlsx', str(source))
        store.mark_failed('quality-fail1', '模拟失败')

        resp = client.get('/api/analytics')

        assert resp.status_code == 200
        quality = resp.json()['quality']
        distribution = {item['grade']: item['count'] for item in quality['distribution']}
        reason_codes = {item['code'] for item in quality['top_reasons']}

        assert quality['scored_count'] == 4
        assert quality['usable_count'] == 1
        assert quality['needs_sample_count'] == 1
        assert quality['low_quality_count'] == 2
        assert distribution['pass'] == 1
        assert distribution['sample'] == 1
        assert distribution['review'] == 1
        assert distribution['critical'] == 1
        assert quality['average_score'] < 100
        assert quality['daily_trends'][-1]['scored'] == 4
        assert {'http_retries', 'validation_issues', 'failed'} <= reason_codes

    def test_analytics_tolerates_malformed_quality_score_json(self, tmp_path):
        from web import store

        source = tmp_path / 'input.xlsx'
        output = tmp_path / 'out.xlsx'
        store.create('bad-quality1', 'bad-quality.xlsx', str(source))
        store.mark_done('bad-quality1', str(output), {
            'validation': {'passed': True, 'issues': []},
            'metrics': {'quality': {'issue_count': 0}},
        })
        conn = store._get_conn()
        conn.execute(
            "UPDATE tasks SET quality_score_json='{bad-json' WHERE id=?",
            ('bad-quality1',),
        )
        conn.commit()

        resp = client.get('/api/analytics')

        assert resp.status_code == 200
        assert resp.json()['quality']['scored_count'] == 1

    def test_tasks_list_paginated(self):
        resp = client.get('/api/tasks?page=1&limit=5')
        assert resp.status_code == 200
        data = resp.json()
        assert 'tasks' in data
        assert 'total' in data
        assert 'page' in data
        assert 'pages' in data

    def test_tasks_list_default_page(self):
        """默认 page=1, limit=20。"""
        resp = client.get('/api/tasks')
        assert resp.status_code == 200
        data = resp.json()
        assert data['page'] == 1

    def test_tasks_list_status_filter(self, tmp_path):
        from web import store

        source = tmp_path / 'input.xlsx'
        output = tmp_path / 'out.xlsx'
        store.create('done00000001', 'done.xlsx', str(source))
        store.mark_done('done00000001', str(output))
        store.create('review000001', 'review.xlsx', str(source))
        store.mark_needs_review('review000001', str(output), {
            'validation': {'passed': False, 'issues': ['第 3 行 Bullet 不足 5 条']},
        })

        resp = client.get('/api/tasks?filter=needs_review')

        assert resp.status_code == 200
        data = resp.json()
        assert data['filter'] == 'needs_review'
        assert data['total'] == 1
        assert data['tasks'][0]['id'] == 'review000001'

    def test_tasks_list_high_risk_filter_includes_metric_risk(self, tmp_path):
        from web import store

        source = tmp_path / 'input.xlsx'
        output = tmp_path / 'out.xlsx'
        store.create('clean0000001', 'clean.xlsx', str(source))
        store.mark_done('clean0000001', str(output), {
            'metrics': {
                'quality': {'issue_count': 0},
                'http_retries': 0,
                'circuit_open': 0,
                'concurrency': {'reductions': 0},
            },
        })
        store.create('retry0000001', 'retry.xlsx', str(source))
        store.mark_done('retry0000001', str(output), {
            'metrics': {
                'quality': {'issue_count': 0},
                'http_retries': 2,
                'circuit_open': 0,
                'concurrency': {'reductions': 0},
            },
        })
        store.create('issue0000001', 'issue.xlsx', str(source))
        store.mark_done('issue0000001', str(output), {
            'validation': {'passed': False, 'issues': ['第 4 行标题丢失关键事实']},
            'metrics': {
                'quality': {'issue_count': 0},
                'http_retries': 0,
                'circuit_open': 0,
                'concurrency': {'reductions': 0},
            },
        })

        resp = client.get('/api/tasks?filter=high_risk')

        assert resp.status_code == 200
        data = resp.json()
        assert data['filter'] == 'high_risk'
        assert data['total'] == 2
        assert {task['id'] for task in data['tasks']} == {'retry0000001', 'issue0000001'}

    def test_tasks_list_high_risk_tolerates_malformed_legacy_stats(self, tmp_path):
        from web import store

        source = tmp_path / 'input.xlsx'
        store.create('legacybad001', 'legacy.xlsx', str(source))
        conn = store._get_conn()
        conn.execute(
            """UPDATE tasks
               SET status='done',
                   stats_json='{bad-json',
                   quality_score_value=100,
                   quality_grade='pass',
                   quality_score_json=?
               WHERE id=?""",
            (
                json.dumps({
                    'score': 100,
                    'grade': 'pass',
                    'label': '可用',
                    'severity': 'ok',
                    'reasons': [],
                }, ensure_ascii=False),
                'legacybad001',
            ),
        )
        conn.commit()

        resp = client.get('/api/tasks?filter=high_risk')

        assert resp.status_code == 200
        assert resp.json()['total'] == 0

    def test_tasks_list_includes_quality_score(self, tmp_path):
        from web import store

        source = tmp_path / 'input.xlsx'
        output = tmp_path / 'out.xlsx'
        store.create('abcdef123456', 'done.xlsx', str(source))
        store.mark_done('abcdef123456', str(output), {
            'validation': {'passed': True, 'issues': []},
            'metrics': {
                'api_success_rate': 1.0,
                'http_retries': 0,
                'quality': {'issue_count': 0},
                'concurrency': {'reductions': 0, 'by_operation': {}},
            },
        })

        resp = client.get('/api/tasks')

        assert resp.status_code == 200
        task = resp.json()['tasks'][0]
        assert task['quality_score']['score'] == 100
        assert task['quality_score']['grade'] == 'pass'

    def test_store_persists_quality_score_for_indexing(self, tmp_path):
        from web import store

        source = tmp_path / 'input.xlsx'
        output = tmp_path / 'out.xlsx'
        store.create('abcdef123456', 'done.xlsx', str(source))
        store.mark_done('abcdef123456', str(output), {
            'validation': {'passed': True, 'issues': []},
            'metrics': {
                'api_success_rate': 1.0,
                'http_retries': 0,
                'quality': {'issue_count': 0},
                'concurrency': {'reductions': 0, 'by_operation': {}},
            },
        })

        conn = store._get_conn()
        row = conn.execute(
            """SELECT quality_score_value, quality_grade, quality_score_json
               FROM tasks WHERE id=?""",
            ('abcdef123456',),
        ).fetchone()
        indexes = {
            index[1]
            for index in conn.execute("PRAGMA index_list(tasks)").fetchall()
        }

        assert row['quality_score_value'] == 100
        assert row['quality_grade'] == 'pass'
        assert json.loads(row['quality_score_json'])['grade'] == 'pass'
        assert 'idx_tasks_quality_score' in indexes
        assert 'idx_tasks_quality_score_sort' in indexes
        assert 'idx_tasks_quality_score_sort_desc' in indexes
        assert 'idx_tasks_quality_grade' in indexes

    def test_store_backfills_missing_quality_score(self, tmp_path):
        from web import store

        source = tmp_path / 'input.xlsx'
        output = tmp_path / 'out.xlsx'
        store.create('abcdef123456', 'done.xlsx', str(source))
        store.mark_done('abcdef123456', str(output), {
            'validation': {'passed': True, 'issues': []},
            'metrics': {
                'api_success_rate': 1.0,
                'http_retries': 0,
                'quality': {'issue_count': 0},
                'concurrency': {'reductions': 0, 'by_operation': {}},
            },
        })
        conn = store._get_conn()
        conn.execute(
            """UPDATE tasks
               SET quality_score_value=NULL, quality_grade=NULL, quality_score_json=NULL
               WHERE id=?""",
            ('abcdef123456',),
        )
        conn.commit()

        assert store.backfill_missing_quality_scores() == 1
        task = store.get('abcdef123456')

        assert task['quality_score']['score'] == 100
        assert conn.execute(
            "SELECT quality_grade FROM tasks WHERE id=?",
            ('abcdef123456',),
        ).fetchone()['quality_grade'] == 'pass'

    def test_tasks_list_quality_filter_and_sort(self, tmp_path):
        from web import store

        source = tmp_path / 'input.xlsx'
        output = tmp_path / 'out.xlsx'
        store.create('pass00000001', 'pass.xlsx', str(source))
        store.mark_done('pass00000001', str(output), {
            'validation': {'passed': True, 'issues': []},
            'metrics': {
                'api_success_rate': 1.0,
                'http_retries': 0,
                'quality': {'issue_count': 0},
                'concurrency': {'reductions': 0, 'by_operation': {}},
            },
        })
        store.create('sample000001', 'sample.xlsx', str(source))
        store.mark_done('sample000001', str(output), {
            'validation': {'passed': True, 'issues': []},
            'metrics': {
                'api_success_rate': 1.0,
                'http_retries': 4,
                'quality': {'issue_count': 0},
                'concurrency': {'reductions': 0, 'by_operation': {}},
            },
        })
        store.create('review000002', 'review.xlsx', str(source))
        store.mark_needs_review('review000002', str(output), {
            'validation': {'passed': False, 'issues': ['第 1 行 Bullet 存在重复内容']},
            'metrics': {
                'api_success_rate': 0.9,
                'http_retries': 2,
                'quality': {'issue_count': 1},
                'concurrency': {'reductions': 1, 'by_operation': {}},
            },
        })

        low = client.get('/api/tasks?filter=low_quality&sort=quality_asc')
        sample = client.get('/api/tasks?filter=needs_sample')
        usable = client.get('/api/tasks?filter=usable')
        sorted_resp = client.get('/api/tasks?sort=quality_asc')

        assert low.status_code == 200
        assert low.json()['filter'] == 'low_quality'
        assert [task['id'] for task in low.json()['tasks']] == ['review000002']
        assert sample.json()['tasks'][0]['id'] == 'sample000001'
        assert usable.json()['tasks'][0]['id'] == 'pass00000001'
        assert sorted_resp.json()['sort'] == 'quality_asc'
        assert [task['id'] for task in sorted_resp.json()['tasks'][:3]] == [
            'review000002',
            'sample000001',
            'pass00000001',
        ]

    def test_tasks_list_page_clamp(self):
        """page=0 和 page=-1 应被 clamp 到 page=1。"""
        for bad in [0, -1]:
            resp = client.get(f'/api/tasks?page={bad}')
            assert resp.status_code == 200
            assert resp.json()['page'] == 1

    def test_tasks_list_invalid_page(self):
        """page=abc 应返回 422。"""
        resp = client.get('/api/tasks?page=abc')
        assert resp.status_code == 422

    def test_templates_endpoint(self):
        resp = client.get('/api/templates')
        assert resp.status_code == 200
        data = resp.json()
        templates = data.get('templates', data)
        assert len(templates) > 0

    def test_stages_endpoint(self):
        resp = client.get('/api/stages')
        assert resp.status_code == 200
        data = resp.json()
        assert 'stages' in data
        assert len(data['stages']) >= 7

    def test_amazon_stages_endpoint(self):
        resp = client.get('/api/stages?pipeline=amazon')
        assert resp.status_code == 200
        assert resp.json()['stages'][0] == '读取表格'

    def test_nonexistent_task_404(self):
        resp = client.get('/api/tasks/nonexistent123')
        assert resp.status_code == 404

    def test_nonexistent_download_404(self):
        resp = client.get('/api/tasks/nonexistent123/download')
        assert resp.status_code == 404

    def test_delete_nonexistent_valid_id_404(self):
        resp = client.delete('/api/tasks/abcdef123456')
        assert resp.status_code == 404

    def test_static_files_served(self):
        """验证静态文件（HTML/JS/CSS）可访问。"""
        for path in ['/', '/app.js', '/styles.css']:
            resp = client.get(path)
            assert resp.status_code == 200, f'{path} returned {resp.status_code}'

    def test_security_headers_are_present(self):
        resp = client.get('/')
        assert resp.headers['x-frame-options'] == 'DENY'
        assert resp.headers['x-content-type-options'] == 'nosniff'
        assert "script-src 'self'" in resp.headers['content-security-policy']
        assert "frame-ancestors 'none'" in resp.headers['content-security-policy']

    def test_cross_site_write_is_rejected(self):
        resp = client.post(
            '/api/settings',
            json={'agnes_key': 'must-not-be-saved'},
            headers={
                'Origin': 'https://attacker.example',
                'Sec-Fetch-Site': 'cross-site',
                'X-CrossPilot-Request': '1',
            },
        )
        assert resp.status_code == 403

    def test_same_origin_browser_write_requires_marker(self):
        resp = client.post(
            '/api/settings',
            json={'agnes_key': 'must-not-be-saved'},
            headers={'Origin': 'http://testserver', 'Sec-Fetch-Site': 'same-origin'},
        )
        assert resp.status_code == 403

    def test_same_origin_browser_write_with_marker_reaches_endpoint(self):
        resp = client.post(
            '/api/tasks/abcdef123456/cancel',
            headers={
                'Origin': 'http://testserver',
                'Sec-Fetch-Site': 'same-origin',
                'X-CrossPilot-Request': '1',
            },
        )
        assert resp.status_code == 404

    def test_optional_basic_auth(self, monkeypatch):
        monkeypatch.setenv('CROSSPILOT_AUTH_PASSWORD', 'secret')
        assert client.get('/api/tasks').status_code == 401
        token = base64.b64encode(b'crosspilot:secret').decode('ascii')
        resp = client.get('/api/tasks', headers={'Authorization': f'Basic {token}'})
        assert resp.status_code == 200

    def test_optional_basic_auth_rejects_malformed_token(self, monkeypatch):
        monkeypatch.setenv('CROSSPILOT_AUTH_PASSWORD', 'secret')
        resp = client.get(
            '/api/tasks',
            headers={'Authorization': 'Basic !!!not-base64!!!'},
        )
        assert resp.status_code == 401

    def test_settings_validate_and_persist_keys(self):
        resp = client.post('/api/settings', json={'agnes_key': 'new-test-key'})
        assert resp.status_code == 200
        assert client.get('/api/settings').json()['agnes_key_set'] is True
        invalid = client.post('/api/settings', json={'agnes_key': ['not', 'a', 'string']})
        assert invalid.status_code == 400

    def test_settings_reject_unsupported_provider(self):
        resp = client.post('/api/settings', json={'text_provider': 'openai'})
        assert resp.status_code == 400
        assert '不支持' in resp.json()['detail']

    def test_settings_persist_exact_models_to_effective_env(self):
        resp = client.post('/api/settings', json={
            'deepseek_text_model': 'deepseek-next',
            'agnes_image_model': 'agnes-image-next',
            'agnes_image_fallback_model': 'agnes-image-stable',
        })

        assert resp.status_code == 200
        settings = client.get('/api/settings').json()
        assert settings['deepseek_text_model'] == 'deepseek-next'
        assert settings['agnes_image_model'] == 'agnes-image-next'
        assert (
            settings['agnes_image_fallback_model']
            == 'agnes-image-stable'
        )

        from crosspilot.config import get_env_path
        saved = get_env_path().read_text(encoding='utf-8')
        assert 'DEEPSEEK_TEXT_MODEL=deepseek-next' in saved
        assert 'AGNES_IMAGE_MODEL=agnes-image-next' in saved

    def test_settings_reject_model_id_with_newline(self):
        resp = client.post('/api/settings', json={
            'agnes_image_model': 'valid\nAGNES_KEY=bad',
        })

        assert resp.status_code == 400

    def test_settings_persist_agnes_fast_congestion_policy(self):
        resp = client.post('/api/settings', json={
            'agnes_503_retry_limit': 1,
            'agnes_503_backoff_min_s': 2,
            'agnes_503_backoff_max_s': 6,
            'agnes_503_circuit_threshold': 2,
            'agnes_503_circuit_cooldown_s': 90,
        })

        assert resp.status_code == 200
        settings = resp.json()['settings']
        assert settings['agnes_503_retry_limit'] == '1'
        assert settings['agnes_503_backoff_max_s'] == '6'
        assert settings['agnes_503_circuit_cooldown_s'] == '90'

    def test_settings_reject_invalid_agnes_congestion_range(self):
        resp = client.post('/api/settings', json={
            'agnes_503_backoff_min_s': 9,
            'agnes_503_backoff_max_s': 3,
        })

        assert resp.status_code == 400
        assert '等待区间' in resp.json()['detail']

    def test_settings_switch_profile_clears_exact_model_overrides(self):
        customized = client.post('/api/settings', json={
            'deepseek_text_model': 'temporary-custom-model',
        })
        assert customized.status_code == 200
        assert (
            client.get('/api/settings').json()['deepseek_text_model']
            == 'temporary-custom-model'
        )

        switched = client.post('/api/settings', json={
            'model_profile': 'test',
        })

        assert switched.status_code == 200
        settings = switched.json()['settings']
        assert settings['model_profile'] == 'test'
        assert settings['deepseek_text_model'] == 'deepseek-v4-flash'
        assert {'production', 'test'} <= set(settings['model_profiles'])

    def test_prompt_api_edits_versions_and_rolls_back(self):
        prompt_id = 'amazon.title_optimize'
        detail = client.get(f'/api/prompts/{prompt_id}')
        assert detail.status_code == 200
        original = detail.json()['content']
        assert detail.json()['source'] == 'default'

        first = client.post(
            f'/api/prompts/{prompt_id}',
            json={'content': original + '\nFirst edit'},
        )
        assert first.status_code == 200
        second = client.post(
            f'/api/prompts/{prompt_id}',
            json={'content': original + '\nSecond edit'},
        )
        assert second.status_code == 200

        history = client.get(
            f'/api/prompts/{prompt_id}/history'
        ).json()['revisions']
        assert history
        rolled_back = client.post(
            f'/api/prompts/{prompt_id}/rollback',
            json={'revision_id': history[0]['revision_id']},
        )

        assert rolled_back.status_code == 200
        assert (
            client.get(f'/api/prompts/{prompt_id}').json()['content']
            == original + '\nFirst edit'
        )

    def test_prompt_profiles_are_isolated_and_can_reset(self):
        prompt_id = 'amazon.title_optimize'
        original = client.get(f'/api/prompts/{prompt_id}').json()['content']
        saved = client.post(
            f'/api/prompts/{prompt_id}',
            json={'content': original + '\nProduction only'},
        )
        assert saved.status_code == 200

        switched = client.post('/api/settings', json={
            'prompt_profile': 'test',
        })
        assert switched.status_code == 200
        test_detail = client.get(f'/api/prompts/{prompt_id}').json()
        assert test_detail['profile'] == 'test'
        assert test_detail['content'] == original

        reset = client.delete(f'/api/prompts/{prompt_id}/override')
        assert reset.status_code == 200

    def test_prompt_api_rejects_contract_breaking_edit(self):
        prompt_id = 'amazon.title_optimize'
        resp = client.post(
            f'/api/prompts/{prompt_id}',
            json={'content': 'No required template variables'},
        )

        assert resp.status_code == 400
        assert '模板变量' in resp.json()['detail']

    def test_store_mark_done_persists_stats(self, tmp_path):
        from web import store
        store.create('stats-job', 'stats.xlsx', str(tmp_path / 'stats.xlsx'))
        store.mark_done('stats-job', str(tmp_path / 'out.xlsx'), {
            'images_reviewed': 7,
            'watermarks': 2,
            'images_generated': 2,
        })
        assert store.get('stats-job')['stats']['images_reviewed'] == 7

    def test_queued_status_is_preserved(self, tmp_path):
        from web import store
        store.create('queue-job', 'queue.xlsx', str(tmp_path / 'queue.xlsx'))
        store.update_progress('queue-job', {'status': 'queued', 'stage': '排队等待', 'percent': 0})
        assert store.get('queue-job')['status'] == 'queued'

    def test_late_progress_cannot_overwrite_terminal_state(self, tmp_path):
        from web import store
        store.create('terminal-job', 'done.xlsx', str(tmp_path / 'done.xlsx'))
        store.mark_done('terminal-job', str(tmp_path / 'out.xlsx'))
        store.update_progress(
            'terminal-job',
            {'status': 'running', 'stage': '旧进度', 'percent': 90},
        )
        assert store.get('terminal-job')['status'] == 'done'

    def test_cancelled_state_cannot_be_overwritten_by_late_progress(self, tmp_path):
        from web import store
        store.create('cancelled-job', 'cancelled.xlsx', str(tmp_path / 'cancelled.xlsx'))
        store.mark_cancelled('cancelled-job')
        store.update_progress(
            'cancelled-job',
            {'status': 'running', 'stage': '旧进度', 'percent': 90},
        )
        assert store.get('cancelled-job')['status'] == 'cancelled'

    def test_task_api_hides_internal_paths_and_cache(self, tmp_path):
        from web import store
        source = tmp_path / 'private-input.xlsx'
        output = tmp_path / 'private-output.xlsx'
        store.create('abcdeffedcba', 'public.xlsx', str(source))
        store.mark_done('abcdeffedcba', str(output), {'images_reviewed': 1})

        detail = client.get('/api/tasks/abcdeffedcba').json()
        listed = client.get('/api/tasks').json()['tasks'][0]

        for task in (detail, listed):
            assert 'input_path' not in task
            assert 'output_path' not in task
            assert 'cache' not in task

    def test_review_report_download_contains_actionable_csv(self, tmp_path):
        from web import store

        source = tmp_path / 'private-input.xlsx'
        output = tmp_path / 'private-output.xlsx'
        output.write_bytes(b'placeholder')
        store.create('abcdef123456', 'public.xlsx', str(source), pipeline='amazon')
        store.mark_needs_review(
            'abcdef123456',
            str(output),
            {
                'validation': {
                    'passed': False,
                    'issues': ['第 3 行 Bullet 存在重复内容'],
                },
                'metrics': {
                    'api_calls': 2,
                    'http_retries': 1,
                    'quality': {'issue_count': 1},
                    'concurrency': {'reductions': 1, 'by_operation': {}},
                },
            },
            '输出存在质量问题',
        )

        resp = client.get('/api/tasks/abcdef123456/review-report')

        assert resp.status_code == 200
        assert resp.content.startswith(b'\xef\xbb\xbf')
        text = resp.content.decode('utf-8-sig')
        assert 'validation,issue,review,3,Bullet' in text
        assert 'concurrency_reductions' in text
        assert str(tmp_path) not in text
        assert 'filename*=' in resp.headers['content-disposition']

    def test_review_data_returns_actionable_rows_with_context(self, tmp_path):
        from web import store

        source = tmp_path / 'amazon-input.json'
        output = tmp_path / 'amazon-output.json'
        source.write_text(json.dumps({
            '商品id': ['sku-1'],
            '产品标题': ['Original BMW Floor Mat'],
            '产品描述': ['Original factory floor mat for front seat'],
            '产品图片链接': [['https://example.com/source.jpg']],
            '变种图片链接': [[]],
        }, ensure_ascii=False), encoding='utf-8')
        output.write_text(json.dumps({
            '商品id': ['sku-1'],
            '产品标题': ['For Floor Mat Front Seat'],
            '产品描述': ['Clean floor mat for front seat'],
            '产品图片链接': [['https://example.com/output.jpg']],
            '变种图片链接': [[]],
            'Bullet Point1': ['Durable front floor mat'],
            'Bullet Point2': ['Durable front floor mat'],
            'Bullet Point3': ['Trim fit replacement mat'],
            'Bullet Point4': ['Easy clean textured surface'],
            'Bullet Point5': ['Designed for daily driving'],
            '关键词信息': ['floor mat,car liner,front seat,trim cover,auto mat,rubber mat,interior mat,replacement mat,driver mat,passenger mat'],
        }, ensure_ascii=False), encoding='utf-8')

        store.create('abcdef123456', 'amazon.json', str(source), pipeline='amazon')
        store.mark_needs_review(
            'abcdef123456',
            str(output),
            {
                'validation': {
                    'passed': False,
                    'issues': ['第 1 行 Bullet 存在重复内容'],
                    'audit': [{
                        'row': 1,
                        'stage': '标题优化',
                        'field': 'title',
                        'method': 'rule',
                        'reason': 'normalize_title',
                        'before': 'Original BMW Floor Mat',
                        'after': 'For Floor Mat Front Seat',
                    }],
                },
                'metrics': {
                    'http_retries': 1,
                    'quality': {'issue_count': 0},
                    'concurrency': {'reductions': 0, 'by_operation': {}},
                },
            },
            '输出存在质量问题',
        )

        resp = client.get('/api/tasks/abcdef123456/review-data')

        assert resp.status_code == 200
        data = resp.json()
        assert data['summary']['issue_count'] == 1
        assert data['summary']['quality_score']['score'] < 100
        assert data['summary']['validation_item_count'] == 1
        assert data['summary']['audit_item_count'] == 1
        issue = next(item for item in data['items'] if item['type'] == 'issue')
        assert issue['row'] == '1'
        assert issue['field'] == 'Bullet'
        assert '处理后: Durable front floor mat' in issue['value']
        assert '检查 5 条 Bullet' in issue['action']
        audit = next(item for item in data['items'] if item['section'] == 'audit')
        assert audit['field'] == 'title'
        assert '标题优化' in audit['message']
        assert str(tmp_path) not in json.dumps(data, ensure_ascii=False)

    def test_review_report_missing_task_404(self):
        resp = client.get('/api/tasks/nonexistent123/review-report')

        assert resp.status_code == 404

    def test_cleanup_keeps_record_when_directory_delete_fails(self, monkeypatch):
        import shutil
        from web import store

        job_dir = os.path.join(store._DATA_DIR, 'uploads', 'cleanup-job')
        os.makedirs(job_dir, exist_ok=True)
        source = os.path.join(job_dir, 'input.xlsx')
        with open(source, 'wb') as handle:
            handle.write(b'placeholder')
        store.create('cleanup-job', 'input.xlsx', source)
        store.mark_failed('cleanup-job', 'test')
        conn = store._get_conn()
        conn.execute(
            'UPDATE tasks SET updated_at = ? WHERE id = ?',
            (time.time() - 30 * 86400, 'cleanup-job'),
        )
        conn.commit()
        monkeypatch.setattr(shutil, 'rmtree', lambda _path: (_ for _ in ()).throw(OSError('locked')))

        assert store.cleanup_old_tasks(days=7) == 0
        assert store.get('cleanup-job') is not None


class TestUploadFlow:
    """上传→查询→删除 全流程（不跑真实管道）。"""

    def _make_xlsx(self, tmp_path, name='test.xlsx'):
        """创建一个有效的 eBay 格式迷你 xlsx。"""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'tiktok_chanpin_'
        # 关键表头（eBay TK 格式）
        headers = {
            1: '商品id', 2: '产品标题', 3: 'Tiktok产品描述',
            15: '价格(站点币种)', 16: '本地展示价', 17: '库存',
            18: '主图(url)地址',
            19: '附图一', 20: '附图二', 21: '附图三',
            22: '附图四', 23: '附图五', 24: '附图六',
            25: '附图七', 26: '附图八',
            27: '视频连接',
            28: '尺码表图片', 29: '变种一1图片',
            36: '来源Url',
        }
        for c, v in headers.items():
            ws.cell(1, c, v)
        # 一行数据
        ws.cell(2, 1, 'test001')
        ws.cell(2, 2, 'Test Product Title')
        ws.cell(2, 3, '<p>Product description</p>')
        ws.cell(2, 15, '9.99')
        ws.cell(2, 16, '9.99')
        ws.cell(2, 17, '100')
        ws.cell(2, 18, 'https://example.com/img.jpg')
        ws.cell(2, 36, 'https://example.com/source')

        fp = str(tmp_path / name)
        wb.save(fp)
        wb.close()
        return fp

    def _make_amazon_json(self, tmp_path, name='amazon.json'):
        path = tmp_path / name
        path.write_text(json.dumps({
            '商品id': ['amazon-1'],
            '产品标题': ['Test Amazon Product'],
            '产品描述': ['Product description'],
            '产品图片链接': [['https://img.example/main.jpg']],
            '变种图片链接': [['https://img.example/variant.jpg']],
        }, ensure_ascii=False), encoding='utf-8')
        return str(path)

    def test_upload_reject_non_xlsx(self, tmp_path):
        """非 xlsx/json 文件应被拒绝。"""
        txt = tmp_path / 'test.txt'
        txt.write_text('not an xlsx')
        resp = client.post('/api/upload/batch', files=[
            ('files', ('test.txt', open(str(txt), 'rb'), 'text/plain'))
        ])
        assert resp.status_code == 200
        results = resp.json()['results']
        assert results[0].get('error') == '只接受 .xlsx 或 Amazon .json 文件'

    def test_upload_rejects_too_many_files(self, tmp_path, monkeypatch):
        web_app = importlib.import_module('web.app')
        monkeypatch.setattr(web_app, 'MAX_BATCH_FILES', 1)
        files = [
            ('files', ('one.xlsx', b'placeholder', 'application/octet-stream')),
            ('files', ('two.xlsx', b'placeholder', 'application/octet-stream')),
        ]

        resp = client.post('/api/upload/batch', files=files)

        assert resp.status_code == 413
        assert '1' in resp.json()['detail']

    def test_upload_accepts_xlsx(self, tmp_path):
        """有效 xlsx 应被接受并返回 job_id。"""
        xlsx = self._make_xlsx(tmp_path)
        resp = client.post('/api/upload/batch', files=[
            ('files', ('test.xlsx', open(xlsx, 'rb'),
             'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'))
        ])
        assert resp.status_code == 200
        results = resp.json()['results']
        assert len(results) == 1
        assert 'job_id' in results[0]
        assert 'error' not in results[0]

    def test_upload_accepts_amazon_json(self, tmp_path):
        path = self._make_amazon_json(tmp_path)
        with open(path, 'rb') as handle:
            resp = client.post('/api/upload/batch', files=[
                ('files', ('amazon.json', handle, 'application/json')),
            ])

        assert resp.status_code == 200
        result = resp.json()['results'][0]
        assert 'job_id' in result
        assert 'error' not in result

    def test_upload_rejects_invalid_amazon_json(self, tmp_path):
        path = tmp_path / 'invalid.json'
        path.write_text(json.dumps({
            '商品id': ['amazon-1'],
            '产品标题': ['Title'],
        }, ensure_ascii=False), encoding='utf-8')
        with open(path, 'rb') as handle:
            resp = client.post('/api/upload/batch', files=[
                ('files', ('invalid.json', handle, 'application/json')),
            ])

        assert resp.status_code == 200
        assert '缺少字段' in resp.json()['results'][0]['error']

    def test_upload_rejects_fake_xlsx(self, tmp_path):
        fake = tmp_path / 'fake.xlsx'
        fake.write_bytes(b'not a zip workbook')
        with open(fake, 'rb') as fh:
            resp = client.post('/api/upload/batch', files=[
                ('files', ('fake.xlsx', fh,
                 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'))
            ])
        assert resp.status_code == 200
        assert '有效的 .xlsx' in resp.json()['results'][0]['error']

    def test_pipeline_detection_routes_amazon(self, tmp_path):
        from web import jobs
        import openpyxl
        path = tmp_path / 'amazon.xlsx'
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(1, 2, '产品标题')
        ws.cell(1, 3, '产品描述')
        wb.save(path)
        wb.close()

        assert jobs._detect_pipeline(str(path)) == 'amazon'

    def test_pipeline_detection_routes_amazon_json(self, tmp_path):
        from web import jobs
        path = self._make_amazon_json(tmp_path)

        assert jobs._detect_pipeline(path) == 'amazon'

    def test_download_preserves_amazon_json_format(self):
        import shutil
        from web import store
        web_app = importlib.import_module('web.app')

        job_id = 'abcdef123456'
        job_dir = os.path.join(web_app.UPLOAD_DIR, job_id)
        os.makedirs(job_dir, exist_ok=True)
        source = os.path.join(job_dir, '商品采集表.json')
        output = os.path.join(job_dir, '商品回填表.json')
        try:
            with open(source, 'w', encoding='utf-8') as handle:
                json.dump({
                    '商品id': ['1'],
                    '产品标题': ['Title'],
                    '产品描述': ['Description'],
                    '产品图片链接': [['https://img/main.jpg']],
                    '变种图片链接': [[]],
                }, handle, ensure_ascii=False)
            with open(output, 'w', encoding='utf-8') as handle:
                json.dump({'商品id': ['1']}, handle, ensure_ascii=False)
            store.create(job_id, '商品采集表.json', source)
            store.set_pipeline(job_id, 'amazon')
            store.mark_done(job_id, output)

            resp = client.get(f'/api/tasks/{job_id}/download')

            assert resp.status_code == 200
            assert resp.headers['content-type'].startswith('application/json')
            assert '.json' in resp.headers['content-disposition']
        finally:
            shutil.rmtree(job_dir, ignore_errors=True)

    def test_task_detail_after_upload(self, tmp_path):
        """上传后查询任务详情。"""
        xlsx = self._make_xlsx(tmp_path)
        resp = client.post('/api/upload/batch', files=[
            ('files', ('test.xlsx', open(xlsx, 'rb'),
             'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'))
        ])
        job_id = resp.json()['results'][0]['job_id']

        # 查询任务
        resp = client.get(f'/api/tasks/{job_id}')
        assert resp.status_code == 200
        task = resp.json()
        assert task['filename'] == 'test.xlsx'
        assert task['status'] in ('queued', 'running', 'failed')

    def test_delete_task(self, tmp_path):
        """上传后删除任务。"""
        xlsx = self._make_xlsx(tmp_path)
        resp = client.post('/api/upload/batch', files=[
            ('files', ('test.xlsx', open(xlsx, 'rb'),
             'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'))
        ])
        job_id = resp.json()['results'][0]['job_id']

        resp = client.delete(f'/api/tasks/{job_id}')
        assert resp.status_code == 200

        # 删除后应 404
        resp = client.get(f'/api/tasks/{job_id}')
        assert resp.status_code == 404

    def test_retry_rejects_active_task(self, tmp_path, monkeypatch):
        from web import jobs, store
        source = tmp_path / 'retry.xlsx'
        source.write_bytes(b'placeholder')
        store.create('abcdef123456', 'retry.xlsx', str(source))
        monkeypatch.setattr(jobs, 'is_active', lambda _job_id: True)

        resp = client.post('/api/tasks/abcdef123456/retry')

        assert resp.status_code == 409

    def test_retry_resumes_cache_by_default(self, tmp_path, monkeypatch):
        from web import jobs, store
        source = tmp_path / 'retry.xlsx'
        source.write_bytes(b'placeholder')
        store.create('abcdef123456', 'retry.xlsx', str(source))
        monkeypatch.setattr(jobs, 'is_active', lambda _job_id: False)
        monkeypatch.setattr(jobs, 'enqueue', lambda *_args: True)
        clear_calls = []
        monkeypatch.setattr(jobs, 'clear_cache', lambda path: clear_calls.append(path))

        resp = client.post('/api/tasks/abcdef123456/retry')

        assert resp.status_code == 200
        assert resp.json()['mode'] == 'resume'
        assert clear_calls == []

    def test_fresh_retry_clears_cache(self, tmp_path, monkeypatch):
        from web import jobs, store
        source = tmp_path / 'retry.xlsx'
        source.write_bytes(b'placeholder')
        store.create('abcdef123456', 'retry.xlsx', str(source))
        monkeypatch.setattr(jobs, 'is_active', lambda _job_id: False)
        monkeypatch.setattr(jobs, 'enqueue', lambda *_args: True)
        clear_calls = []
        monkeypatch.setattr(jobs, 'clear_cache', lambda path: clear_calls.append(path))

        resp = client.post('/api/tasks/abcdef123456/retry?fresh=true')

        assert resp.status_code == 200
        assert resp.json()['mode'] == 'fresh'
        assert clear_calls == [str(source)]

    def test_delete_rejects_uncancellable_thread(self, tmp_path, monkeypatch):
        from web import jobs, store
        source = tmp_path / 'running.xlsx'
        source.write_bytes(b'placeholder')
        store.create('abcdef123456', 'running.xlsx', str(source))
        monkeypatch.setattr(jobs, 'is_active', lambda _job_id: True)
        monkeypatch.setattr(jobs, 'cancel', lambda _job_id: False)

        resp = client.delete('/api/tasks/abcdef123456')

        assert resp.status_code == 409
        assert store.get('abcdef123456') is not None

    def test_cancel_active_task_preserves_record(self, tmp_path, monkeypatch):
        from web import jobs, store
        source = tmp_path / 'running.xlsx'
        source.write_bytes(b'placeholder')
        store.create('abcdef123456', 'running.xlsx', str(source))
        store.mark_running('abcdef123456')
        monkeypatch.setattr(jobs, 'is_active', lambda _job_id: True)
        monkeypatch.setattr(jobs, 'cancel', lambda _job_id: True)

        resp = client.post('/api/tasks/abcdef123456/cancel')

        assert resp.status_code == 200
        assert resp.json()['status'] == 'cancelled'
        assert store.get('abcdef123456') is not None


def test_static_frontend_has_no_inline_script_handlers():
    root = os.path.join(os.path.dirname(__file__), '..', 'web', 'static')
    javascript = [
        name for name in os.listdir(root)
        if name.endswith('.js')
    ]
    source = '\n'.join(
        open(os.path.join(root, name), encoding='utf-8').read()
        for name in ('index.html', *javascript)
    )
    assert 'onclick=' not in source
    assert 'onchange=' not in source
    assert 'fonts.googleapis.com' not in source
    assert 'window._' not in source


def test_static_frontend_module_graph_links():
    node = shutil.which('node')
    if not node:
        pytest.skip('Node.js is not installed')
    root = os.path.join(os.path.dirname(__file__), '..')
    result = subprocess.run(
        [
            node,
            '--no-warnings',
            '--experimental-vm-modules',
            os.path.join(root, 'tools', 'check_frontend_modules.mjs'),
        ],
        cwd=root,
        capture_output=True,
        text=True,
        timeout=20,
        check=False,
    )
    assert result.returncode == 0, result.stdout + result.stderr
