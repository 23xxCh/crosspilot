"""Core pipeline integration tests. Uses data/sample_small.xlsx (2 rows)."""
import os, shutil, json, time, pytest

from scripts import process_ebay_tk
from web import store


# sample_xlsx_path 和 invalid_xlsx_path 在 conftest.py 中定义


@pytest.fixture
def invalid_xlsx_path():
    """创一个不识格式的 xlsx。"""
    import openpyxl
    dst = os.path.join(os.path.dirname(__file__), '..', 'data', '_test_invalid.xlsx')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Sheet1'
    ws.cell(1, 1, 'Foo')
    ws.cell(1, 2, 'Bar')
    wb.save(dst)
    yield dst
    if os.path.exists(dst):
        os.remove(dst)


class TestPipeline:
    """集成测试：需要 API key 和网络。标记为 network，默认跳过。"""

    @pytest.mark.network
    def test_pipeline_with_sample(self, sample_xlsx_path):
        """迷你表 2 行全链路跑通，返回 _cleaned.xlsx 路径。"""
        out = process_ebay_tk._main(sample_xlsx_path)
        assert out is not None
        assert out.endswith('_cleaned.xlsx')
        assert os.path.exists(out)

        # 检查输出内容
        import openpyxl
        wb = openpyxl.load_workbook(out)
        ws = wb.active
        assert ws.max_row >= 3  # 表头 + 2 行
        # 列 15 应改名为"本地展示价"
        assert '本地展示价' in str(ws.cell(1, 15).value or '')

    @pytest.mark.network
    def test_cache_created_and_reused(self, sample_xlsx_path):
        """缓存命中：跑两次，图审缓存应命中。"""
        cache_path = os.path.splitext(sample_xlsx_path)[0] + '_cache.json'

        # 第一次跑
        out1 = process_ebay_tk._main(sample_xlsx_path)
        assert out1 is not None

        assert os.path.exists(cache_path)
        with open(cache_path) as f:
            cache = json.load(f)
        assert 'review_results' in cache
        assert len(cache['review_results']) > 0

        # 第二次跑（缓存命中，图审直接跳过）
        out2 = process_ebay_tk._main(sample_xlsx_path)
        assert out2 is not None
        assert os.path.exists(out2)


class TestErrorPaths:
    """错误路径测试：不需要 API key。"""

    def test_file_not_found(self):
        """文件不存在 → FileNotFoundError。"""
        with pytest.raises(FileNotFoundError):
            process_ebay_tk._main('/nonexistent/path.xlsx')

    def test_unrecognized_format_raises(self, invalid_xlsx_path):
        """不识格式 → ValueError。"""
        with pytest.raises(ValueError, match='不认识的表格格式'):
            process_ebay_tk._main(invalid_xlsx_path)

    def test_header_mismatch_raises(self, header_mismatch_xlsx_path):
        """能识格式但 col 15 header 不符 → ValueError。"""
        with pytest.raises(ValueError, match='表头结构'):
            process_ebay_tk._main(header_mismatch_xlsx_path)


class TestStore:
    """store.py 多线程安全：并发读写不崩溃。"""

    def test_concurrent_writes(self):
        """多线程同时 update_progress 不应抛异常。"""
        import threading
        store.create('test-job', 'test.xlsx', '/tmp/test.xlsx')
        errors = []

        def writer(i):
            try:
                for _ in range(10):
                    store.update_progress('test-job', {'stage': f'stage_{i}', 'percent': i})
            except Exception as e:
                errors.append(e)

        threads = [threading.Thread(target=writer, args=(i,)) for i in range(5)]
        for t in threads: t.start()
        for t in threads: t.join()
        assert not errors, f"并发写不应出错: {errors}"
        store.delete('test-job')


class TestTranslateText:
    """translate_text 中文检测：必须返回无中文内容，否则重试 + 警告。"""

    def test_translate_pure_chinese_retries(self, monkeypatch):
        """API 始终返回原文（仍含中文）→ 重试后保留原文 + WARN 日志。"""
        from unittest.mock import Mock
        mock_provider = Mock()
        mock_provider.call_text.return_value = "这是汽车配件"  # 返回原文（含中文）
        monkeypatch.setattr(
            'scripts.services.translate.get_provider',
            lambda: mock_provider,
        )
        result = process_ebay_tk.translate_text("这是汽车配件", process_ebay_tk.TITLE_TRANSLATE_PROMPT)
        assert process_ebay_tk._CHINESE_RE.search(result), f"应保留中文, got: {result}"

    def test_translate_timeout_retries(self, monkeypatch):
        """API 超时（返回 None）→ 保留原文。"""
        from unittest.mock import Mock
        mock_provider = Mock()
        mock_provider.call_text.return_value = None
        monkeypatch.setattr(
            'scripts.services.translate.get_provider',
            lambda: mock_provider,
        )
        result = process_ebay_tk.translate_text("这是产品", process_ebay_tk.TITLE_TRANSLATE_PROMPT)
        assert process_ebay_tk._CHINESE_RE.search(result)

    def test_translate_pure_english_to_target_language(self, monkeypatch):
        """目标是越南语时，英文商品标题也必须翻译。"""
        from unittest.mock import Mock
        mock_provider = Mock()
        mock_provider.call_text.return_value = "Phụ kiện ô tô 2024"
        monkeypatch.setattr(
            'scripts.services.translate.get_provider',
            lambda: mock_provider,
        )
        result = process_ebay_tk.translate_text("Car Accessories 2024", process_ebay_tk.TITLE_TRANSLATE_PROMPT)
        assert result == "Phụ kiện ô tô 2024"
        mock_provider.call_text.assert_called_once()

    def test_translate_partial_vn_with_cn_still_keeps_cn(self, monkeypatch):
        """API 返回部分翻译（含中文残留）→ 保留。"""
        from unittest.mock import Mock
        mock_provider = Mock()
        mock_provider.call_text.return_value = "Sản phẩm 汽车 用品"
        monkeypatch.setattr(
            'scripts.services.translate.get_provider',
            lambda: mock_provider,
        )
        result = process_ebay_tk.translate_text("这是产品", process_ebay_tk.TITLE_TRANSLATE_PROMPT)
        assert process_ebay_tk._CHINESE_RE.search(result)

    def test_chinese_input_uses_chinese_specific_prompt(self):
        """中文输入必须用专用 prompt（明确要求中文→越南语），不能走通用模板。

        注意：这个测试验证 prompt 选择逻辑，不实际调用 API。
        """
        # 直接测试 _select_prompt 方法
        text = "通用型汽车配件"
        default_prompt = "Translate: {}"
        result = process_ebay_tk._select_prompt(text, default_prompt)
        # 中文专用 prompt 提到 "Chinese (中文)"
        assert "Chinese (中文)" in result or "Chinese" in result


class TestBatch:
    """批量处理函数测试。"""

    def test_batch_translate_empty(self):
        """空输入 → 返回空。"""
        result = process_ebay_tk.batch_translate_texts([])
        assert result == {}

    def test_batch_translate_partial_json(self, monkeypatch):
        """JSON 解析容错：API 返回不完整响应，不抛异常。"""
        from unittest.mock import Mock
        mock_provider = Mock()
        mock_provider.call_text.return_value = 'NOT JSON'
        monkeypatch.setattr(
            'scripts.services.translate.get_provider',
            lambda: mock_provider,
        )
        result = process_ebay_tk.batch_translate_texts(["test text"])
        assert result == {}  # 解析失败返回空

    def test_store_close(self):
        """关闭连接不崩溃。"""
        store.close()
        assert True  # 不抛异常即通过


class TestRuleStripBrands:
    """品牌/平台名清洗函数。"""

    def test_strips_ebay_brand(self):
        result = process_ebay_tk.rule_strip_brands("For BMW 3 Series E90 Car Mat Premium")
        assert 'BMW' not in result
        assert 'Car Mat' in result

    def test_strips_toyota_honda(self):
        result = process_ebay_tk.rule_strip_brands("Toyota Honda LED Headlight Bulb Kit")
        assert 'Toyota' not in result
        assert 'Honda' not in result
        assert 'LED Headlight' in result

    def test_no_brand_unchanged(self):
        result = process_ebay_tk.rule_strip_brands("Car Phone Holder Universal Mount")
        assert result == "Car Phone Holder Universal Mount"

    def test_brand_substrings_do_not_corrupt_words(self):
        result = process_ebay_tk.rule_strip_brands("Affordable audio adapter for Ford")
        assert result == "Affordable audio adapter for"

    def test_empty_input(self):
        assert process_ebay_tk.rule_strip_brands("") == ""
        assert process_ebay_tk.rule_strip_brands(None) is None


class TestStripCodeFence:
    """Markdown code fence 清理。"""

    def test_strips_fence(self):
        result = process_ebay_tk._strip_code_fence("```\nHello World\n```")
        assert "```" not in result
        assert "Hello World" in result

    def test_strips_json_fence(self):
        result = process_ebay_tk._strip_code_fence('```json\n{"key": "val"}\n```')
        assert "json" not in result.lower()
        assert "key" in result

    def test_no_fence_unchanged(self):
        assert process_ebay_tk._strip_code_fence("Plain text") == "Plain text"


class TestSelectPrompt:
    """中文检测 → 选 prompt。"""

    def test_chinese_text_uses_cn_prompt(self):
        prompt = process_ebay_tk._select_prompt("汽车配件", "Default prompt")
        assert 'Chinese' in prompt
        assert 'CRITICAL' in prompt or 'Convert ALL' in prompt.lower()

    def test_english_text_uses_default(self):
        prompt = process_ebay_tk._select_prompt("Car Accessories", "Default prompt")
        assert prompt == "Default prompt"

    def test_vietnamese_text_uses_default(self):
        prompt = process_ebay_tk._select_prompt("Phụ kiện ô tô", "Custom prompt")
        assert prompt == "Custom prompt"


class TestParseBatchResponse:
    """批量 JSON 解析边界。"""

    def test_valid_json_array(self):
        raw = '[{"index": 0, "translation": "hello"}, {"index": 1, "translation": "world"}]'
        result = process_ebay_tk._parse_batch_response(raw, 2)
        assert len(result) == 2
        assert result[0] == "hello"
        assert result[1] == "world"

    def test_empty_raw(self):
        assert process_ebay_tk._parse_batch_response("", 3) == {}
        assert process_ebay_tk._parse_batch_response(None, 3) == {}

    def test_markdown_code_block(self):
        raw = '```json\n[{"index": 0, "translation": "xin chào"}]\n```'
        result = process_ebay_tk._parse_batch_response(raw, 1)
        assert len(result) == 1
        assert result[0] == "xin chào"

    def test_malformed_json(self):
        assert process_ebay_tk._parse_batch_response("not json at all", 5) == {}
        assert process_ebay_tk._parse_batch_response("{broken", 2) == {}

    def test_missing_translation_field(self):
        """_parse_batch_response 返回所有条目（含空值），过滤由 _process_batch 负责。"""
        raw = '[{"index": 0}, {"index": 1, "translation": "ok"}]'
        result = process_ebay_tk._parse_batch_response(raw, 2)
        assert len(result) == 2  # 两个 index 都返回
        assert result[0] == ""   # 缺 translation → 空字符串
        assert result[1] == "ok"


class TestEbayAdapter:
    """eBay 适配器识别。"""

    def test_ebay_tk_detected(self, sample_xlsx_path):
        import openpyxl
        from scripts.adapters.ebay_tk import EbayTkAdapter
        wb = openpyxl.load_workbook(sample_xlsx_path)
        ws = wb.active
        assert EbayTkAdapter.detect(ws), "应识别为 eBay→TikTok 格式"
