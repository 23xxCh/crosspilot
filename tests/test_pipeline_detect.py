"""pipeline.py detect_input 格式检测测试。"""
import pytest, json, os
from pathlib import Path


class TestDetectInput:
    """detect_input 格式自动检测。"""

    def test_detect_json_by_filename(self, tmp_path):
        """文件名含 'amazon' → 检测为 amazon。"""
        from crosspilot.pipeline import detect_input

        f = tmp_path / 'amazon_data.json'
        f.write_text(json.dumps({'产品标题': ['test'], '产品描述': ['desc'],
                                   '产品图片链接': [[]], '变种图片链接': [[]],
                                   '商品id': ['1']}), encoding='utf-8-sig')

        info = detect_input(str(f))
        assert info['platform'] == 'amazon'
        assert info['is_json'] is True
        assert info['row_count'] == 1

    def test_detect_json_by_content(self, tmp_path):
        """JSON 内容含 '产品标题' 字段 → amazon。"""
        from crosspilot.pipeline import detect_input

        f = tmp_path / 'unknown.json'
        f.write_text(json.dumps({'产品标题': ['test'], '产品描述': ['desc'],
                                   '产品图片链接': [[]], '变种图片链接': [[]],
                                   '商品id': ['1']}), encoding='utf-8-sig')

        info = detect_input(str(f))
        assert info['platform'] == 'amazon'

    def test_detect_xlsx_by_header(self, tmp_path):
        """XLSX 表头含 'Title' → ebay。"""
        from crosspilot.pipeline import detect_input
        import openpyxl

        f = tmp_path / 'ebay_list.xlsx'
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(1, 1).value = 'Title'
        ws.cell(2, 1).value = 'Some Product'
        wb.save(str(f))
        wb.close()

        info = detect_input(str(f))
        assert info['platform'] == 'ebay'
        assert info['is_xlsx'] is True

    def test_xlsx_defaults_to_amazon(self, tmp_path):
        """XLSX 表头无 Title → 默认 amazon。"""
        from crosspilot.pipeline import detect_input
        import openpyxl

        f = tmp_path / 'unknown.xlsx'
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(1, 1).value = 'SKU'
        wb.save(str(f))
        wb.close()

        info = detect_input(str(f))
        assert info['platform'] == 'amazon'

    def test_ebay_filename_detected(self, tmp_path):
        """文件名含 'ebay' 但 JSON 结构优先 → amazon（结构检测优先于文件名）。"""
        from crosspilot.pipeline import detect_input

        f = tmp_path / 'ebay_export.json'
        f.write_text(json.dumps({'产品标题': ['t'], '产品描述': ['d'],
                                   '产品图片链接': [[]], '变种图片链接': [[]],
                                   '商品id': ['1']}), encoding='utf-8-sig')

        info = detect_input(str(f))
        # JSON 结构检测（产品标题/产品描述 字段）优先于文件名
        assert info['platform'] == 'amazon'

    def test_caiji_filename_detected(self, tmp_path):
        """文件名含 '采集表' → amazon。"""
        from crosspilot.pipeline import detect_input

        f = tmp_path / '跨境电商自动化采集表3.json'
        f.write_text(json.dumps({'产品标题': ['t'], '产品描述': ['d'],
                                   '产品图片链接': [[]], '变种图片链接': [[]],
                                   '商品id': ['1']}), encoding='utf-8-sig')

        info = detect_input(str(f))
        assert info['platform'] == 'amazon'


class TestDetectInputErrors:
    """错误输入处理。"""

    def test_file_not_found(self, tmp_path):
        from crosspilot.pipeline import detect_input
        with pytest.raises(FileNotFoundError):
            detect_input(str(tmp_path / 'nonexistent.json'))

    def test_directory_rejected(self, tmp_path):
        from crosspilot.pipeline import detect_input
        with pytest.raises(IsADirectoryError):
            detect_input(str(tmp_path))

    def test_unsupported_format(self, tmp_path):
        from crosspilot.pipeline import detect_input
        f = tmp_path / 'test.txt'
        f.write_text('hello')
        with pytest.raises(ValueError, match='Unsupported'):
            detect_input(str(f))

    def test_invalid_json(self, tmp_path):
        from crosspilot.pipeline import detect_input
        f = tmp_path / 'bad.json'
        f.write_text('not json')
        with pytest.raises(ValueError, match='Invalid JSON'):
            detect_input(str(f))

    def test_empty_json_accepted(self, tmp_path):
        """空 JSON {} 不抛异常，返回 row_count=0。"""
        from crosspilot.pipeline import detect_input
        f = tmp_path / 'empty.json'
        f.write_text('{}')
        info = detect_input(str(f))
        assert info['row_count'] == 0
        assert info['platform'] != 'amazon'  # 无产品标题字段
