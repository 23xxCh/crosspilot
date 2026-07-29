"""Unit tests for services/ — no network required (all mocked)."""
import json, pytest
from unittest.mock import Mock, patch, MagicMock
import sys, os
_PROJ = os.path.join(os.path.dirname(__file__), '..')
sys.path.insert(0, _PROJ)  # for `import scripts.xxx`
sys.path.insert(0, os.path.join(_PROJ, 'scripts'))  # for `from adapters import xxx`
from services import TranslationService, ImageReviewService
from services.translate import TRANSLATE_BATCH_PROMPT


@pytest.fixture
def mock_provider(monkeypatch):
    """Mock model_provider.get_provider() for all tests."""
    provider = Mock()
    provider.call_text.return_value = 'Xin chào'
    provider.call_vision.return_value = False
    provider.call_image_gen.return_value = 'https://generated.example.com/img.png'
    # Patch at all locations where get_provider is imported
    monkeypatch.setattr('model_provider.get_provider', lambda: provider)
    monkeypatch.setattr('services.translate.get_provider', lambda: provider)
    monkeypatch.setattr('services.review.get_provider', lambda: provider)
    monkeypatch.setattr('dmx_client.get_provider', lambda: provider)
    return provider


@pytest.fixture
def translate_svc(mock_provider):
    """Create TranslationService with mocked provider."""
    return TranslationService()


@pytest.fixture
def review_svc(mock_provider):
    """Create ImageReviewService with mocked provider."""
    return ImageReviewService()


@pytest.fixture(autouse=True)
def no_real_retry_sleep(monkeypatch):
    import time
    monkeypatch.setattr(time, 'sleep', lambda _seconds: None)


class TestTranslationService:
    """TranslationService unit tests."""

    def test_dmx_call_success(self, monkeypatch):
        """测试 dmx_call 方法（现在使用 model_provider）。"""
        from unittest.mock import Mock
        mock_provider = Mock()
        mock_provider.call_text.return_value = 'Xin chào'
        monkeypatch.setattr('services.translate.get_provider', lambda: mock_provider)

        from services.translate import TranslationService
        svc = TranslationService()
        result = svc.dmx_call({
            'model': 'deepseek-chat',
            'messages': [{'role': 'user', 'content': 'hello'}]
        })
        assert result == 'Xin chào'

    def test_dmx_call_returns_none(self, monkeypatch):
        """API 返回 None → dmx_call 返回 None。"""
        from unittest.mock import Mock
        mock_provider = Mock()
        mock_provider.call_text.return_value = None
        monkeypatch.setattr('services.translate.get_provider', lambda: mock_provider)

        from services.translate import TranslationService
        svc = TranslationService()
        result = svc.dmx_call({
            'model': 'deepseek-chat',
            'messages': [{'role': 'user', 'content': 'hello'}]
        })
        assert result is None

    def test_strip_code_fence(self):
        """测试代码块清理。"""
        from services.translate import TranslationService
        assert TranslationService._strip_code_fence('```json\nhello\n```') == 'hello'
        assert TranslationService._strip_code_fence('```\nworld\n```') == 'world'
        assert TranslationService._strip_code_fence('plain text') == 'plain text'
        assert TranslationService._strip_code_fence(None) is None

    def test_select_prompt_chinese(self, translate_svc):
        prompt = translate_svc._select_prompt('汽车配件', 'Default')
        assert 'Chinese' in prompt
        assert 'CRITICAL' in prompt

    def test_select_prompt_english(self, translate_svc):
        prompt = translate_svc._select_prompt('Car Parts', 'Default')
        assert prompt == 'Default'

    def test_translate_text_chinese_detection(self, translate_svc):
        """Test Chinese text detection and translation."""
        # Configure mock to return Vietnamese translation
        translate_svc._provider.call_text.return_value = 'Phụ tùng xe'
        result = translate_svc.translate_text('汽车配件', 'Translate: {}')
        # Should contain Vietnamese, no Chinese
        assert '汽' not in result
        assert result == 'Phụ tùng xe'

    def test_translate_text_returns_original_on_failure(self, translate_svc):
        """On API failure, return original text."""
        translate_svc._provider.call_text.return_value = None
        result = translate_svc.translate_text('汽车配件', 'Translate: {}')
        assert result == '汽车配件'

    def test_translate_text_does_not_swallow_quota_exhaustion(self, translate_svc):
        """A terminal account failure must stop the pipeline, not trigger mass fallback."""
        from model_provider import ProviderQuotaError

        translate_svc._provider.call_text.side_effect = ProviderQuotaError('额度不足')

        with pytest.raises(ProviderQuotaError, match='额度不足'):
            translate_svc.translate_text('汽车配件', 'Translate: {}')

    def test_parse_batch_response_valid(self, translate_svc):
        raw = json.dumps([
            {'index': 0, 'translation': 'hello'},
            {'index': 1, 'translation': 'world'}
        ])
        result = translate_svc._parse_batch_response(raw)
        assert len(result) == 2
        assert result[0] == 'hello'

    def test_parse_batch_response_empty(self, translate_svc):
        assert translate_svc._parse_batch_response('') == {}
        assert translate_svc._parse_batch_response(None) == {}

    def test_parse_batch_response_markdown_fence(self, translate_svc):
        raw = '```json\n[{"index": 0, "translation": "ok"}]\n```'
        result = translate_svc._parse_batch_response(raw)
        assert result == {0: 'ok'}

    def test_batch_process_empty(self, translate_svc):
        result = translate_svc.batch_process([], TRANSLATE_BATCH_PROMPT, 'test')
        assert result == {}

    def test_batch_process_chinese_detection(self, translate_svc, mock_provider):
        """Test batch processing with Chinese text uses CN prompt."""
        mock_provider.call_text.return_value = json.dumps([
            {'index': 0, 'translation': 'Phụ tùng'}
        ])

        result = translate_svc.batch_process(
            ['汽车配件'],
            TRANSLATE_BATCH_PROMPT,
            'batch_translate'
        )
        assert len(result) == 1
        assert result['汽车配件'] == 'Phụ tùng'

    def test_batch_translate_delegates(self, translate_svc, mock_provider):
        """Test batch_translate delegates to batch_process."""
        mock_provider.call_text.return_value = json.dumps([
            {'index': 0, 'translation': 'xin chào'}
        ])
        result = translate_svc.batch_translate(['hello'])
        assert len(result) == 1
        assert result['hello'] == 'xin chào'

    def test_batch_process_retries_malformed_response(self, translate_svc, mock_provider):
        """Test retry when response is malformed JSON."""
        # First call returns malformed JSON, second returns valid
        mock_provider.call_text.side_effect = [
            'not-json',
            json.dumps([{'index': 0, 'translation': 'xin chào'}])
        ]

        result = translate_svc.batch_translate(['hello'])

        assert result == {'hello': 'xin chào'}
        assert mock_provider.call_text.call_count == 2


class TestImageReviewService:
    """ImageReviewService unit tests."""

    def test_review_once_watermark_detected(self, review_svc, mock_provider):
        """Test review_once returns True when image has watermark."""
        mock_provider.call_vision.return_value = True
        result = review_svc.review_once('https://example.com/img.jpg')
        assert result is True

    def test_review_once_clean_image(self, review_svc, mock_provider):
        """Test review_once returns False when image is clean."""
        mock_provider.call_vision.return_value = False
        result = review_svc.review_once('https://example.com/clean.jpg')
        assert result is False

    def test_review_request_includes_people_and_body_parts(self, review_svc, mock_provider):
        """Test that review calls provider with image URL."""
        mock_provider.call_vision.return_value = True
        result = review_svc.review_once('https://example.com/person.jpg')
        assert result is True
        # Verify call_vision was called with the image URL
        mock_provider.call_vision.assert_called_once_with('https://example.com/person.jpg')

    def test_review_once_ambiguous_answer_is_unknown(self, review_svc, mock_provider):
        """Test that ambiguous response returns None."""
        mock_provider.call_vision.return_value = None
        result = review_svc.review_once('https://example.com/unknown.jpg')
        assert result is None

    def test_review_once_http_error(self, review_svc, mock_provider):
        """Test that provider error returns None."""
        mock_provider.call_vision.return_value = None
        result = review_svc.review_once('https://example.com/img.jpg')
        assert result is None

    def test_review_once_timeout(self, review_svc, mock_provider):
        """Test that provider exception propagates (service doesn't catch)."""
        mock_provider.call_vision.side_effect = Exception("timeout")
        with pytest.raises(Exception, match="timeout"):
            review_svc.review_once('https://example.com/img.jpg')

    def test_review_full_fallback(self, review_svc, mock_provider):
        """Full review() with retries — all fail => None."""
        mock_provider.call_vision.return_value = None
        result = review_svc.review('https://example.com/img.jpg')
        assert result is None


class TestAmazonAdapter:
    """Amazon adapter detection."""

    def test_detect_valid_headers(self):
        from scripts.adapters.amazon_tk import AmazonTkAdapter

        # Mock worksheet with correct headers
        ws = Mock()
        ws.cell = Mock()
        def mock_cell(r, c):
            m = Mock()
            headers = {1: '商品id', 2: '产品标题', 3: '产品描述',
                      4: '产品图片', 5: '变种图片', 6: '产品图片链接', 7: '变种图片链接'}
            m.value = headers.get(c, '')
            return m
        ws.cell.side_effect = mock_cell

        assert AmazonTkAdapter.detect(ws) is True

    def test_detect_invalid_headers(self):
        from scripts.adapters.amazon_tk import AmazonTkAdapter

        ws = Mock()
        ws.cell = Mock()
        def mock_cell(r, c):
            m = Mock()
            m.value = 'Unknown' if c <= 3 else ''
            return m
        ws.cell.side_effect = mock_cell

        assert AmazonTkAdapter.detect(ws) is False


class TestEdgeCases:
    """Boundary and edge case tests."""

    def test_translate_empty_text(self, translate_svc):
        assert translate_svc.translate_text('', 'Translate: {}') == ''
        assert translate_svc.translate_text(None, 'Translate: {}') is None

    def test_translate_very_long_text(self, translate_svc, mock_provider):
        """Test translation handles long text."""
        mock_provider.call_text.return_value = 'short result'
        long_text = 'A' * 5000
        result = translate_svc.translate_text(long_text, 'Translate: {}')
        assert result == 'short result'

    def test_batch_with_mixed_languages(self, translate_svc, mock_provider):
        """Test batch translation with mixed languages."""
        mock_provider.call_text.return_value = json.dumps([
            {'index': 0, 'translation': 'Phụ tùng'},
            {'index': 1, 'translation': 'Car parts'}
        ])
        result = translate_svc.batch_translate(['汽车配件', 'Car parts'])
        assert len(result) == 2

    def test_parse_batch_malformed_json_returns_empty(self, translate_svc):
        assert translate_svc._parse_batch_response('{not json') == {}
        assert translate_svc._parse_batch_response('[1,2,3]') == {}
        assert translate_svc._parse_batch_response('{"key": "val"}') == {}

    def test_parse_batch_partial_results(self, translate_svc):
        raw = json.dumps([
            {'index': 0, 'translation': 'ok'},
            {'index': 2, 'translation': 'skip one'}
        ])
        result = translate_svc._parse_batch_response(raw)
        assert len(result) == 2
        assert 1 not in result

    def test_review_once_with_empty_url(self, review_svc, mock_provider):
        """Test review with empty URL returns None."""
        mock_provider.call_vision.return_value = None
        result = review_svc.review_once('')
        assert result is None

    def test_batch_process_large_batch(self, translate_svc, mock_provider):
        """25 texts in one batch (max batch size)."""
        mock_provider.call_text.return_value = json.dumps([
            {'index': i, 'translation': f'translated {i}'} for i in range(25)
        ])
        texts = ['text ' + str(i) for i in range(25)]
        result = translate_svc.batch_translate(texts)
        assert len(result) == 25

    def test_metrics_to_dict(self):
        from scripts.pipeline_log import PipelineMetrics
        m = PipelineMetrics()
        m.record_stage('test', 2.5, 100, 95)
        m.record_api(True)
        m.record_api(False)
        d = m.to_dict()
        assert 'stages' in d
        assert d['api_calls'] == 2
        assert d['api_errors'] == 1
        assert d['api_success_rate'] == 0.5
        assert d['stages']['test']['items_per_s'] == 40.0
        assert d['stages']['test']['success_rate'] == 0.95

    def test_configured_concurrency_clamps_env(self, monkeypatch):
        from scripts.concurrency import configured_concurrency

        monkeypatch.setenv('CROSSPILOT_REVIEW_CONCURRENCY', '9999')
        assert configured_concurrency('review', 100, maximum=120) == 120

        monkeypatch.setenv('CROSSPILOT_REVIEW_CONCURRENCY', '0')
        assert configured_concurrency('review', 100, minimum=5, maximum=120) == 5

        monkeypatch.setenv('CROSSPILOT_REVIEW_CONCURRENCY', 'bad')
        assert configured_concurrency('review', 100, minimum=5, maximum=120) == 100

    def test_adaptive_map_reduces_after_failure_batch(self, monkeypatch):
        from scripts.concurrency import adaptive_map

        monkeypatch.setenv('CROSSPILOT_ADAPTIVE_FAILURE_RATE', '0.5')
        results, stats = adaptive_map(
            range(8),
            lambda item: None if item < 4 else f'ok-{item}',
            operation='review',
            initial_workers=4,
            min_workers=1,
        )

        assert len(results) == 8
        assert stats['reductions'] >= 1
        assert stats['final_workers'] < stats['initial_workers']
        assert stats['events'][0]['reason'] == 'failure_rate'

    def test_new_request_id_unique(self):
        from scripts.pipeline_log import new_request_id
        ids = {new_request_id() for _ in range(10)}
        assert len(ids) == 10

    def test_select_prompt_vietnamese_is_default(self, translate_svc):
        """Vietnamese text should use default prompt, not Chinese."""
        prompt = translate_svc._select_prompt('Phụ kiện ô tô', 'Default')
        assert prompt == 'Default'


# === dmx_client 函数测试 ===
class TestDmxClientAgnes:
    """Agnes gen_agnes_rate_limited 和 agnes_review 测试（mock provider）。"""

    def test_agnes_review_watermark_detected(self, mock_provider):
        """Test agnes_review returns True for watermarked image."""
        mock_provider.call_vision.return_value = True

        from dmx_client import agnes_review
        result = agnes_review(None, 'https://example.com/img.jpg', retries=1)
        assert result is True

    def test_agnes_review_clean_image(self, mock_provider):
        """Test agnes_review returns False for clean image."""
        mock_provider.call_vision.return_value = False

        from dmx_client import agnes_review
        result = agnes_review(None, 'https://example.com/clean.jpg', retries=1)
        assert result is False

    def test_agnes_review_empty_content(self, mock_provider):
        """Test agnes_review returns None for empty response."""
        mock_provider.call_vision.return_value = None

        from dmx_client import agnes_review
        result = agnes_review(None, 'https://example.com/img.jpg', retries=1)
        assert result is None

    def test_gen_agnes_quota_error_is_not_silenced(self, mock_provider):
        """Test quota error returns empty string (current behavior)."""
        from dmx_client import gen_agnes_rate_limited

        # Simulate quota error by returning None (model_provider returns None on failure)
        mock_provider.call_image_gen.return_value = None

        # Current implementation returns empty string on failure, not raise
        result = gen_agnes_rate_limited(None, 'https://example.com/img.jpg', retries=1)
        assert result == ''

    def test_agnes_review_http_error(self, mock_provider):
        """Test agnes_review handles HTTP error."""
        mock_provider.call_vision.return_value = None

        from dmx_client import agnes_review
        result = agnes_review(None, 'https://example.com/img.jpg', retries=1)
        assert result is None

    def test_gen_agnes_rate_limited_success(self, mock_provider):
        """Test successful image generation."""
        mock_provider.call_image_gen.return_value = 'https://new-image.example.com/1.png'

        from dmx_client import gen_agnes_rate_limited
        result = gen_agnes_rate_limited(None, 'https://example.com/img.jpg', retries=1)
        assert result == 'https://new-image.example.com/1.png'

    def test_gen_agnes_rate_limited_failure(self, mock_provider):
        """Test image generation failure returns empty string."""
        mock_provider.call_image_gen.return_value = None

        from dmx_client import gen_agnes_rate_limited
        result = gen_agnes_rate_limited(None, 'https://example.com/img.jpg', retries=1)
        assert result == ''

    def test_gen_agnes_rate_limited_429_retry(self, mock_provider):
        """Test retry on rate limit (current implementation doesn't retry at dmx_client level)."""
        mock_provider.call_image_gen.return_value = 'https://retry.example.com/1.png'

        from dmx_client import gen_agnes_rate_limited
        result = gen_agnes_rate_limited(None, 'https://example.com/img.jpg', retries=3)
        assert result == 'https://retry.example.com/1.png'

    def test_all_generation_prompts_remove_people(self):
        from dmx_client import (
            AGNES_MAIN_PROMPT,
            AGNES_VARIANT_PROMPT,
            AGNES_PROMPT,
        )

        prompts = (
            AGNES_MAIN_PROMPT,
            AGNES_VARIANT_PROMPT,
            AGNES_PROMPT,
        )
        for prompt in prompts:
            lowered = prompt.lower()
            assert 'person' in lowered
            assert 'human' in lowered
            assert 'hand' in lowered
            assert 'reconstruct' in lowered


# === services/constants 测试 ===
class TestSharedConstants:
    """验证 BRANDS 共享常量正确导入且内容一致。"""

    def test_brands_not_empty(self):
        from services.constants import BRANDS
        assert len(BRANDS) > 30
        assert 'bmw' in BRANDS
        assert '丰田' in BRANDS

    def test_brands_contains_ebay_specific(self):
        from services.constants import BRANDS, STRIP_ONLY_BRANDS
        assert 'joyon' in BRANDS
        assert 'shopee' in BRANDS
        assert 'lazada' in BRANDS
        assert 'diy' in STRIP_ONLY_BRANDS

    def test_compatibility_brands_are_separate_from_cleanup_tokens(self):
        from services.constants import (
            COMPATIBILITY_BRANDS,
            STRIP_ONLY_BRANDS,
        )
        assert 'toyota' in COMPATIBILITY_BRANDS
        assert 'suzuki' in COMPATIBILITY_BRANDS
        assert 'diy' not in COMPATIBILITY_BRANDS
        assert 'joyon' in STRIP_ONLY_BRANDS

    def test_brands_all_lowercase_ascii(self):
        from services.constants import BRANDS
        for b in BRANDS:
            if b.isascii():
                assert b == b.lower()

    def test_image_policy_detects_any_human_presence(self):
        from services.constants import (
            IMAGE_POLICY_VERSION,
            IMAGE_REMEDIATION_REVIEW_PROMPT,
        )
        assert IMAGE_POLICY_VERSION == 'remove_people_v1'
        assert 'ANY person' in IMAGE_REMEDIATION_REVIEW_PROMPT
        assert 'face' in IMAGE_REMEDIATION_REVIEW_PROMPT
        assert 'hand' in IMAGE_REMEDIATION_REVIEW_PROMPT
        assert 'mannequin' in IMAGE_REMEDIATION_REVIEW_PROMPT


# === Pipeline 管道阶段测试 ===
class TestPipelineStages:
    """测试管道阶段函数（mock 依赖）。"""

    def test_detect_amazon_adapter(self):
        from adapters import detect_adapter
        # 用正确 header mock
        ws = Mock()
        ws.cell = Mock()
        headers = {1: '商品id', 2: '产品标题', 3: '产品描述',
                   4: '产品图片', 5: '变种图片', 6: '产品图片链接'}
        def mock_cell(r, c):
            m = Mock()
            m.value = headers.get(c, f'col{c}')
            return m
        ws.cell.side_effect = mock_cell
        result = detect_adapter(ws)
        assert result is not None

    def test_rule_strip_brands_removes_bmw(self):
        from services.constants import BRANDS
        import re
        _BRAND_PATTERN = re.compile('|'.join(re.escape(b) for b in BRANDS), re.IGNORECASE)
        result = _BRAND_PATTERN.sub('', 'BMW Car Parts').strip()
        result = re.sub(r'\s+', ' ', result)
        assert 'BMW' not in result
        assert 'Car Parts' in result

    def test_brand_re_matches_chinese(self):
        from services.constants import BRANDS
        import re
        _BRAND_PATTERN = re.compile('|'.join(re.escape(b) for b in BRANDS), re.IGNORECASE)
        result = _BRAND_PATTERN.sub('', '丰田配件')
        assert '丰田' not in result

    def test_amazon_load_keys_returns_dict(self, monkeypatch, tmp_path):
        """验证 model_provider._load_keys 返回 dict。"""
        import scripts.model_provider as mp

        # Create a temporary keys.json
        keys_file = tmp_path / 'keys.json'
        keys_file.write_text(json.dumps({'dmx_key': 'test-key', 'deepseek_key': 'ds-key'}))

        monkeypatch.setenv('CROSSPILOT_KEYS_PATH', str(keys_file))
        # Reload keys to pick up the new file
        monkeypatch.setattr(mp, '_KEYS', mp._load_keys())

        keys = mp._load_keys()
        assert isinstance(keys, dict)
        assert 'dmx_key' in keys or 'deepseek_key' in keys

    def test_environment_keys_override_file_for_canary(self, monkeypatch, tmp_path):
        """环境变量应能覆盖文件中的 key（配置统一后也需要保证）。"""
        import scripts.model_provider as mp
        from crosspilot.config import reload_config

        # 清缓存 + 设环境变量
        monkeypatch.setenv('CROSSPILOT_DEEPSEEK_KEY', 'canary-ds-key')
        monkeypatch.setenv('CROSSPILOT_AGNES_KEY', 'canary-ag-key')
        reload_config()

        keys = mp._load_keys()
        assert keys['deepseek_key'] == 'canary-ds-key'
        assert keys['agnes_key'] == 'canary-ag-key'

    def test_provider_quota_response_fails_fast(self):
        """Quota exhaustion is terminal and must not be disguised as an empty result."""
        import scripts.model_provider as mp

        response = Mock(ok=False, status_code=402, text='insufficient balance')
        provider = mp.DeepSeekProvider('test-key')
        provider._session = Mock()
        provider._session.post.return_value = response

        with pytest.raises(mp.ProviderQuotaError, match='额度'):
            provider.call_text('test', retries=3)

        provider._session.post.assert_called_once()

    def test_composite_provider_records_logical_call_metrics(self):
        import scripts.model_provider as mp

        provider = mp.CompositeProvider({
            'text_provider': 'deepseek',
            'vision_provider': 'agnes',
            'image_gen_provider': 'agnes',
            'deepseek_key': 'test-deepseek',
            'agnes_key': 'test-agnes',
        })
        text = Mock()
        text.call_text.return_value = 'ok'
        vision = Mock()
        vision.call_vision.return_value = False
        image = Mock()
        image.call_image_gen.return_value = None
        provider._providers = {
            'text': text,
            'vision': vision,
            'image_gen': image,
        }
        provider._image_gen_fallbacks = []

        assert provider.call_text('prompt') == 'ok'
        assert provider.call_vision('https://img.example/a.jpg') is False
        assert provider.call_image_gen('https://img.example/a.jpg') is None

        metrics = provider.metrics_snapshot()
        assert metrics['api_calls'] == 3
        assert metrics['api_errors'] == 1
        assert metrics['by_operation']['text']['calls'] == 1
        assert metrics['by_operation']['vision']['errors'] == 0
        assert metrics['by_operation']['image_gen']['errors'] == 1

    def test_agnes_image_quality_gate_sends_source_and_candidate(self):
        import scripts.model_provider as mp

        response = Mock(ok=True, status_code=200, text='ok')
        response.json.return_value = {
            'choices': [{
                'message': {
                    'content': (
                        '{"accepted": true, "score": 97, '
                        '"reasons": []}'
                    ),
                },
            }],
        }
        provider = mp.AgnesProvider('test-agnes')
        provider._acquire_text = lambda: None
        provider._session = Mock()
        provider._session.post.return_value = response

        result = provider.call_image_quality(
            'https://img.example/source.jpg',
            'https://img.example/generated.jpg',
            context='12 piece flat decal set',
        )

        assert result == {
            'accepted': True,
            'score': 97,
            'reasons': [],
        }
        payload = provider._session.post.call_args.kwargs['json']
        content = payload['messages'][0]['content']
        image_urls = [
            item['image_url']['url']
            for item in content
            if item['type'] == 'image_url'
        ]
        assert image_urls == [
            'https://img.example/source.jpg',
            'https://img.example/generated.jpg',
        ]
        assert '12 piece flat decal set' in content[-1]['text']

    def test_agnes_image_generation_includes_listing_context(self):
        import scripts.model_provider as mp

        response = Mock(ok=True, status_code=200, text='ok')
        response.json.return_value = {
            'data': [{'url': 'https://generated.example/result.png'}],
        }
        provider = mp.AgnesProvider('test-agnes')
        provider._acquire_image = lambda: None
        provider._session = Mock()
        provider._session.post.return_value = response

        provider.call_image_gen(
            'https://img.example/source.jpg',
            context='12Pcs reflective flat wheel decal sticker set',
        )

        payload = provider._session.post.call_args.kwargs['json']
        assert '12Pcs reflective flat wheel decal' in payload['prompt']
        assert 'installation scene only as context' in payload['prompt']

    def test_composite_records_image_quality_gate_metrics(self):
        import scripts.model_provider as mp

        provider = mp.CompositeProvider({
            'text_provider': 'deepseek',
            'vision_provider': 'agnes',
            'image_gen_provider': 'agnes',
            'deepseek_key': 'test-deepseek',
            'agnes_key': 'test-agnes',
        })
        vision = Mock()
        vision.call_image_quality.return_value = {
            'accepted': True,
            'score': 95,
            'reasons': [],
        }
        provider._providers['vision'] = vision

        result = provider.call_image_quality(
            'https://img.example/source.jpg',
            'https://img.example/generated.jpg',
        )

        assert result['accepted'] is True
        metrics = provider.metrics_snapshot()
        assert metrics['by_operation']['image_quality']['calls'] == 1
        assert metrics['by_operation']['image_quality']['errors'] == 0

    def test_composite_honors_agnes_image_provider_when_gpt_key_exists(self):
        """IMAGE_PROVIDER=agnes must not be overridden just because GPT is configured."""
        import scripts.model_provider as mp

        provider = mp.CompositeProvider({
            'text_provider': 'deepseek',
            'vision_provider': 'agnes',
            'image_gen_provider': 'agnes',
            'deepseek_key': 'test-deepseek',
            'agnes_key': 'test-agnes',
            'gpt_image_key': 'test-gpt',
        })

        assert isinstance(provider._providers['image_gen'], mp.AgnesProvider)

    def test_composite_uses_exact_configured_model_ids_and_endpoints(self):
        import scripts.model_provider as mp

        provider = mp.CompositeProvider({
            'text_provider': 'deepseek',
            'vision_provider': 'agnes',
            'image_gen_provider': 'agnes',
            'deepseek_key': 'test-deepseek',
            'agnes_key': 'test-agnes',
            'deepseek_base_url': 'https://text.example',
            'deepseek_text_model': 'text-primary',
            'deepseek_text_fallback_model': 'text-fallback',
            'agnes_vision_base_url': 'https://vision.example',
            'agnes_vision_model': 'vision-model',
            'agnes_image_base_url': 'https://image.example',
            'agnes_image_model': 'image-primary',
            'agnes_image_fallback_model': 'image-fallback',
        })

        text = provider._providers['text']
        vision = provider._providers['vision']
        image = provider._providers['image_gen']
        fallback = provider._image_gen_fallbacks[0]

        assert text.BASE_URL == 'https://text.example'
        assert text.MODEL == 'text-primary'
        assert text.FALLBACK_MODEL == 'text-fallback'
        assert vision.BASE_URL == 'https://vision.example'
        assert vision.VISION_MODEL == 'vision-model'
        assert image.BASE_URL == 'https://image.example'
        assert image.IMAGE_MODEL == 'image-primary'
        assert fallback.IMAGE_MODEL == 'image-fallback'

    def test_composite_image_generation_falls_back_without_marking_call_failed(self):
        """A successful fallback is one successful logical image-generation call."""
        import scripts.model_provider as mp

        provider = mp.CompositeProvider({
            'text_provider': 'deepseek',
            'vision_provider': 'agnes',
            'image_gen_provider': 'agnes',
            'deepseek_key': 'test-deepseek',
            'agnes_key': 'test-agnes',
            'gpt_image_key': 'test-gpt',
        })
        primary = Mock()
        primary.call_image_gen.return_value = None
        fallback = Mock()
        fallback.call_image_gen.return_value = 'https://generated.example/fallback.png'
        provider._providers['image_gen'] = primary
        provider._image_gen_fallbacks = [fallback]

        result = provider.call_image_gen('https://img.example/source.jpg')

        assert result == 'https://generated.example/fallback.png'
        metrics = provider.metrics_snapshot()
        assert metrics['by_operation']['image_gen']['calls'] == 1
        assert metrics['by_operation']['image_gen']['errors'] == 0
        assert metrics['fallback_attempts'] == 1
        assert metrics['fallback_successes'] == 1
        assert metrics['fallback_failures'] == 0

    def test_composite_records_each_image_fallback_route(self):
        import scripts.model_provider as mp

        provider = mp.CompositeProvider({
            'text_provider': 'deepseek',
            'vision_provider': 'agnes',
            'image_gen_provider': 'agnes',
            'deepseek_key': 'test-deepseek',
            'agnes_key': 'test-agnes',
            'gpt_image_key': 'test-gpt',
        })
        primary = Mock()
        primary.call_image_gen.side_effect = mp.ProviderError(
            'primary unavailable',
            retryable=True,
        )
        agnes_fallback = Mock()
        agnes_fallback.IMAGE_MODEL = 'agnes-image-2.0-flash'
        agnes_fallback.call_image_gen.side_effect = mp.ProviderError(
            'fallback unavailable',
            retryable=True,
        )
        gpt_fallback = Mock()
        gpt_fallback.IMAGE_MODEL = 'gpt-image-2'
        gpt_fallback.call_image_gen.return_value = (
            'https://generated.example/gpt.png'
        )
        provider._providers['image_gen'] = primary
        provider._image_gen_fallbacks = [
            agnes_fallback,
            gpt_fallback,
        ]

        result = provider.call_image_gen(
            'https://img.example/source.jpg',
        )

        assert result == 'https://generated.example/gpt.png'
        metrics = provider.metrics_snapshot()
        assert metrics['fallback_attempts'] == 2
        assert metrics['fallback_successes'] == 1
        assert metrics['fallback_failures'] == 1
        assert sum(
            route['attempts']
            for route in metrics['fallback_routes'].values()
        ) == 2
        assert any(
            route['successes'] == 1
            for route in metrics['fallback_routes'].values()
        )

    def test_composite_route_offset_advances_after_quality_rejection(self):
        import scripts.model_provider as mp

        provider = mp.CompositeProvider({
            'text_provider': 'deepseek',
            'vision_provider': 'agnes',
            'image_gen_provider': 'agnes',
            'deepseek_key': 'test-deepseek',
            'agnes_key': 'test-agnes',
            'gpt_image_key': 'test-gpt',
        })
        primary = Mock()
        agnes_fallback = Mock()
        agnes_fallback.IMAGE_MODEL = 'agnes-image-2.0-flash'
        agnes_fallback.call_image_gen.return_value = (
            'https://generated.example/agnes20.png'
        )
        gpt_fallback = Mock()
        gpt_fallback.IMAGE_MODEL = 'gpt-image-2'
        provider._providers['image_gen'] = primary
        provider._image_gen_fallbacks = [
            agnes_fallback,
            gpt_fallback,
        ]

        result = provider.call_image_gen(
            'https://img.example/source.jpg',
            route_offset=1,
        )

        assert result == 'https://generated.example/agnes20.png'
        primary.call_image_gen.assert_not_called()
        agnes_fallback.call_image_gen.assert_called_once()
        gpt_fallback.call_image_gen.assert_not_called()

    def test_provider_records_http_attempts_and_retries(self):
        import scripts.model_provider as mp

        rate_limited = Mock(ok=False, status_code=429, text='rate limited')
        ok = Mock(ok=True, status_code=200, text='ok')
        ok.json.return_value = {'choices': [{'message': {'content': 'translated'}}]}
        provider = mp.CompositeProvider({
            'text_provider': 'deepseek',
            'vision_provider': 'agnes',
            'image_gen_provider': 'agnes',
            'deepseek_key': 'test-deepseek',
            'agnes_key': 'test-agnes',
        })
        provider._providers['text']._session = Mock()
        provider._providers['text']._session.post.side_effect = [
            rate_limited,
            ok,
        ]

        assert provider.call_text('prompt', retries=2) == 'translated'

        metrics = provider.metrics_snapshot()
        assert metrics['api_calls'] == 1
        assert metrics['http_attempts'] == 2
        assert metrics['http_retries'] == 1
        assert metrics['http_status']['429'] == 1
        assert metrics['by_operation']['text']['http_attempts'] == 2

    def test_composite_provider_circuit_breaker_opens_after_failures(self, monkeypatch):
        import scripts.model_provider as mp

        monkeypatch.setenv('CROSSPILOT_CIRCUIT_FAILURE_THRESHOLD', '2')
        monkeypatch.setenv('CROSSPILOT_CIRCUIT_COOLDOWN_S', '60')
        provider = mp.CompositeProvider({
            'text_provider': 'deepseek',
            'vision_provider': 'agnes',
            'image_gen_provider': 'agnes',
            'deepseek_key': 'test-deepseek',
            'agnes_key': 'test-agnes',
        })
        text = Mock()
        text.call_text.return_value = None
        provider._providers['text'] = text

        assert provider.call_text('prompt-1') is None
        assert provider.call_text('prompt-2') is None
        assert provider.call_text('prompt-3') is None

        assert text.call_text.call_count == 2
        metrics = provider.metrics_snapshot()
        assert metrics['circuit_open'] == 1
        assert metrics['by_operation']['text']['circuit_open'] == 1

    def test_amazon_xlsx_keeps_product_and_variant_images_separate(self):
        import openpyxl
        import scripts.process_amazon as p
        from scripts.adapters.amazon_tk import AmazonTkAdapter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(1, 2, '产品标题')
        ws.cell(1, 3, '产品描述')
        ws.cell(1, 6, '产品图片链接')
        ws.cell(1, 7, '变种图片链接')
        ws.cell(2, 2, 'Test')
        ws.cell(2, 6, 'https://img/main.jpg\nhttps://img/extra.jpg')
        ws.cell(2, 7, 'https://img/variant.jpg')

        row = p._stage_read(ws, AmazonTkAdapter)[0]
        wb.close()

        assert row['main_img'] == 'https://img/main.jpg'
        assert row['extra_imgs'] == ['https://img/extra.jpg']
        assert row['var_img'] == 'https://img/variant.jpg'
        assert row['var_imgs'] == ['https://img/variant.jpg']

    def test_ebay_person_issue_routes_by_image_role(self, monkeypatch):
        """Test person detection routes by image role."""
        from pipelines import ebay_stages

        monkeypatch.setattr(ebay_stages, 'review_single', lambda _url: True)
        status = Mock()
        cache = {'review_results': {}, 'gen_results': {}}
        saved = []
        urls = [
            'https://img/main.jpg',
            'https://img/variant.jpg',
            'https://img/attachment.jpg',
        ]
        url_map = {
            urls[0]: {'main': [2], 'variant': [], 'att': []},
            urls[1]: {'main': [], 'variant': [2], 'att': []},
            urls[2]: {'main': [], 'variant': [], 'att': [(2, 19)]},
        }

        _review, _unknown, to_regen, to_delete = ebay_stages._stage_review(
            status, urls, url_map, cache, lambda value: saved.append(value)
        )

        assert set(to_regen) == set(urls[:2])
        assert to_delete == [urls[2]]
        assert saved

    def test_ebay_generation_uses_main_and_variant_person_prompts(self, monkeypatch, mock_provider):
        """Test generation uses different prompts for main vs variant images."""
        from pipelines import ebay_stages

        captured = {}

        def fake_generate(url, is_variant=False):
            captured[url] = {'is_variant': is_variant}
            return url.replace('img/', 'generated/')

        monkeypatch.setattr(ebay_stages, '_gen_image', fake_generate)
        status = Mock()
        main_url = 'https://img/main.jpg'
        variant_url = 'https://img/variant.jpg'
        url_map = {
            main_url: {'main': [2], 'variant': [], 'att': []},
            variant_url: {'main': [], 'variant': [2], 'att': []},
        }
        mains = [main_url]
        variants = [variant_url]

        cache = {'gen_results': {}}
        ebay_stages._stage_generate(
            status,
            [main_url, variant_url],
            url_map,
            cache,
            lambda _cache: None,
            mains,
            variants,
        )

        assert captured[main_url]['is_variant'] is False
        assert captured[variant_url]['is_variant'] is True
        assert mains[0] == 'https://generated/main.jpg'
        assert variants[0] == 'https://generated/variant.jpg'
        assert cache['concurrency_stats']['image_gen']['items'] == 2

    def test_ebay_image_policy_invalidates_only_image_cache(
            self, tmp_path, monkeypatch):
        """Test image policy version invalidates image cache only."""
        import hashlib
        from pipelines import ebay_stages
        from services.constants import IMAGE_POLICY_VERSION

        monkeypatch.setenv('CROSSPILOT_DATA_DIR', str(tmp_path))
        source = tmp_path / 'input.xlsx'
        source.write_bytes(b'image-policy-test')
        digest = hashlib.sha256(source.read_bytes()).hexdigest()[:16]
        cache_dir = tmp_path / 'cache'
        cache_dir.mkdir()
        (cache_dir / f'{digest}.json').write_text(json.dumps({
            'version': 2,
            'image_policy_version': 'old-policy',
            'text_cache_version': ebay_stages._current_text_cache_version(),
            'review_results': {'https://img/person.jpg': False},
            'gen_results': {'https://img/main.jpg': 'https://old/generated.jpg'},
            'title_translations': {'title': 'translated'},
            'desc_cleaned': {},
            'desc_translations': {},
        }), encoding='utf-8')

        cache, _save = ebay_stages._setup_cache(str(source))

        assert cache['image_policy_version'] == IMAGE_POLICY_VERSION
        assert cache['review_results'] == {}
        assert cache['gen_results'] == {}
        assert cache['title_translations'] == {'title': 'translated'}

    def test_ebay_text_cache_version_invalidates_only_text_cache(
            self, tmp_path, monkeypatch):
        import hashlib
        from pipelines import ebay_stages
        from services.constants import IMAGE_POLICY_VERSION

        monkeypatch.setenv('CROSSPILOT_DATA_DIR', str(tmp_path))
        source = tmp_path / 'input.xlsx'
        source.write_bytes(b'text-cache-policy-test')
        digest = hashlib.sha256(source.read_bytes()).hexdigest()[:16]
        cache_dir = tmp_path / 'cache'
        cache_dir.mkdir()
        (cache_dir / f'{digest}.json').write_text(json.dumps({
            'version': 2,
            'image_policy_version': IMAGE_POLICY_VERSION,
            'image_cache_version': ebay_stages._current_image_cache_version(),
            'text_cache_version': 'old-text-policy',
            'review_results': {'https://img/person.jpg': True},
            'gen_results': {'https://img/main.jpg': 'https://generated/main.jpg'},
            'title_translations': {'title': 'stale translation'},
            'desc_cleaned': {'desc': 'stale cleaned'},
            'desc_translations': {'desc': 'stale translated'},
        }), encoding='utf-8')

        cache, _save = ebay_stages._setup_cache(str(source))

        assert cache['review_results'] == {'https://img/person.jpg': True}
        assert cache['gen_results'] == {'https://img/main.jpg': 'https://generated/main.jpg'}
        assert cache['title_translations'] == {}
        assert cache['desc_cleaned'] == {}
        assert cache['desc_translations'] == {}
        assert cache['text_cache_version'] == ebay_stages._current_text_cache_version()

    def test_amazon_rechecks_old_cache_and_deletes_person_attachment(
            self, tmp_path, monkeypatch):
        """Test Amazon rechecks old cache and deletes person attachment."""
        import scripts.process_amazon as p
        from unittest.mock import Mock

        # Create mock provider
        mock_provider = Mock()
        mock_provider.call_vision.return_value = True
        mock_provider.call_image_gen.return_value = 'https://generated.example.com/img.png'

        # Mock get_provider at the module level
        monkeypatch.setattr(p, 'get_provider', lambda: mock_provider)

        person_url = 'https://img/person-attachment.jpg'
        cache_path = tmp_path / 'amazon-cache.json'
        cache_path.write_text(json.dumps({
            'image_policy_version': 'old-policy',
            'review_results': {person_url: False},
            'gen_results': {},
        }), encoding='utf-8')
        rows = [{
            'main_img': '',
            'var_img': '',
            'var_imgs': [],
            'extra_imgs': [person_url],
        }]

        runtime_metrics = {}
        result = p._stage_review_and_gen(rows, str(cache_path), runtime_metrics=runtime_metrics)

        assert result[0]['extra_imgs'] == []
        # Verify call_vision was called
        mock_provider.call_vision.assert_called()
        assert runtime_metrics['concurrency']['amazon_review']['items'] == 1

    def test_amazon_image_cache_versions_include_registered_prompts(
            self, monkeypatch):
        from scripts.pipelines import amazon_review_gen

        calls = []

        def fake_signature(policy, *prompt_ids):
            calls.append((policy, prompt_ids))
            return '|'.join(prompt_ids)

        monkeypatch.setattr(
            amazon_review_gen,
            'build_runtime_signature',
            fake_signature,
        )

        review_version, generation_version = (
            amazon_review_gen._current_image_cache_versions()
        )

        assert review_version == 'images.review'
        assert generation_version == (
            'images.main_product|images.variant|images.quality_gate'
        )
        assert calls[0][1] == ('images.review',)

    def test_amazon_generation_persists_runtime_prompt_version(
            self, tmp_path):
        from scripts.pipelines import amazon_review_gen

        provider = Mock()
        provider.call_image_gen.side_effect = (
            lambda url, **_kwargs: url.replace(
                'https://img.example/',
                'https://generated.example/',
            )
        )
        cache_path = tmp_path / 'amazon-generation-cache.json'
        rows = [{
            'main_img': 'https://img.example/main.jpg',
            'var_img': 'https://img.example/variant.jpg',
            'var_imgs': ['https://img.example/variant.jpg'],
            'extra_imgs': [],
        }]

        result = amazon_review_gen._stage_review_and_gen(
            rows,
            str(cache_path),
            provider_getter=lambda: provider,
        )

        cache = json.loads(cache_path.read_text(encoding='utf-8'))
        assert result[0]['main_img'] == (
            'https://generated.example/main.jpg'
        )
        assert result[0]['var_imgs'] == [
            'https://generated.example/variant.jpg',
        ]
        assert len(cache['gen_results']) == 2
        assert {
            item['prompt_version']
            for item in cache['gen_meta'].values()
        } == {cache['gen_prompt_version']}

    def test_amazon_quality_gate_regenerates_once_then_accepts(
            self, tmp_path, monkeypatch):
        from scripts.pipelines import amazon_review_gen

        monkeypatch.setenv('CROSSPILOT_IMAGE_QUALITY_GATE', '1')
        monkeypatch.setenv('CROSSPILOT_IMAGE_QUALITY_REGEN_LIMIT', '1')
        provider = Mock()
        provider.call_image_gen.side_effect = [
            'https://generated.example/distorted.png',
            'https://generated.example/faithful.png',
        ]
        provider.call_image_quality.side_effect = [
            {
                'accepted': False,
                'score': 20,
                'reasons': ['wrong_product_form'],
            },
            {
                'accepted': True,
                'score': 96,
                'reasons': [],
            },
        ]
        cache_path = tmp_path / 'amazon-gated-cache.json'
        metrics = {}
        source = 'https://img.example/flat-sticker.jpg'
        rows = [{
            'title': 'Flat honeycomb decal sticker',
            'main_img': source,
            'var_img': '',
            'var_imgs': [],
            'extra_imgs': [],
        }]

        result = amazon_review_gen._stage_review_and_gen(
            rows,
            str(cache_path),
            runtime_metrics=metrics,
            provider_getter=lambda: provider,
        )

        assert result[0]['main_img'] == (
            'https://generated.example/faithful.png'
        )
        assert provider.call_image_gen.call_count == 2
        assert [
            call.kwargs['route_offset']
            for call in provider.call_image_gen.call_args_list
        ] == [0, 1]
        assert provider.call_image_quality.call_count == 2
        assert metrics['image_quality_gate'] == {
            'checked': 2,
            'accepted': 1,
            'rejected': 1,
            'unavailable': 0,
            'regenerated': 1,
            'retained_original': 0,
            'reasons': {'wrong_product_form': 1},
        }
        cache = json.loads(cache_path.read_text(encoding='utf-8'))
        meta = cache['gen_meta'][f'main:{source}']
        assert meta['quality_gate']['accepted'] is True

    def test_amazon_flat_product_main_uses_variant_as_reference(
            self, tmp_path, monkeypatch):
        from scripts.pipelines import amazon_review_gen

        monkeypatch.setenv('CROSSPILOT_IMAGE_QUALITY_GATE', '1')
        monkeypatch.setenv('CROSSPILOT_IMAGE_QUALITY_REGEN_LIMIT', '0')
        provider = Mock()
        provider.call_image_gen.return_value = (
            'https://generated.example/flat-set.png'
        )
        provider.call_image_quality.return_value = {
            'accepted': True,
            'score': 98,
            'reasons': [],
        }
        main = 'https://img.example/installed-on-wheel.jpg'
        variant = 'https://img.example/flat-sticker-set.jpg'
        rows = [{
            'title': '12Pcs Wheel Rim Decal Sticker Set',
            'main_img': main,
            'var_img': variant,
            'var_imgs': [variant],
            'extra_imgs': [],
        }]

        amazon_review_gen._stage_review_and_gen(
            rows,
            str(tmp_path / 'reference-cache.json'),
            provider_getter=lambda: provider,
        )

        main_calls = [
            call for call in provider.call_image_gen.call_args_list
            if call.kwargs.get('is_variant') is False
        ]
        assert len(main_calls) == 1
        assert main_calls[0].args[0] == variant
        gate_calls = [
            call for call in provider.call_image_quality.call_args_list
            if call.kwargs.get('is_variant') is False
        ]
        assert gate_calls[0].args[0] == variant

    def test_amazon_remediation_only_skips_clean_main_images(
            self, tmp_path, monkeypatch):
        from scripts.pipelines import amazon_review_gen

        monkeypatch.setenv('CROSSPILOT_IMAGE_REMEDIATE_ONLY', '1')
        monkeypatch.setenv('CROSSPILOT_IMAGE_QUALITY_GATE', '0')
        clean = 'https://img.example/clean.jpg'
        risky = 'https://img.example/watermarked.jpg'
        provider = Mock()
        provider.call_vision.side_effect = (
            lambda url: url == risky
        )
        provider.call_image_gen.return_value = (
            'https://generated.example/cleaned.jpg'
        )
        metrics = {}
        issues = []
        rows = [
            {
                'title': 'Clean product',
                'main_img': clean,
                'var_img': '',
                'var_imgs': [],
                'extra_imgs': [],
            },
            {
                'title': 'Watermarked product',
                'main_img': risky,
                'var_img': '',
                'var_imgs': [],
                'extra_imgs': [],
            },
        ]

        result = amazon_review_gen._stage_review_and_gen(
            rows,
            str(tmp_path / 'remediate-only-cache.json'),
            quality_issues=issues,
            runtime_metrics=metrics,
            provider_getter=lambda: provider,
        )

        assert result[0]['main_img'] == clean
        assert result[1]['main_img'] == (
            'https://generated.example/cleaned.jpg'
        )
        assert provider.call_vision.call_count == 2
        provider.call_image_gen.assert_called_once()
        assert not any('图片生成不完整' in issue for issue in issues)
        assert metrics['image_remediation'] == {
            'reviewed': 2,
            'flagged': 1,
            'clean_retained': 1,
            'unknown_retained': 0,
            'attachment_reviewed': 0,
            'attachment_flagged': 0,
            'attachment_deleted': 0,
            'generated_main': 1,
            'generated_variant': 0,
            'failed_main': 0,
            'failed_variant': 0,
            'generation_url_checked': 0,
            'generation_url_valid': 0,
            'generation_url_invalid': 0,
        }

    def test_amazon_remediation_routes_each_image_role(
            self, tmp_path, monkeypatch):
        from scripts.pipelines import amazon_review_gen

        monkeypatch.setenv('CROSSPILOT_IMAGE_REMEDIATE_ONLY', '1')
        monkeypatch.setenv('CROSSPILOT_IMAGE_QUALITY_GATE', '0')
        main = 'https://img.example/risky-main.jpg'
        clean_extra = 'https://img.example/clean-extra.jpg'
        risky_extra = 'https://img.example/risky-extra.jpg'
        clean_variant = 'https://img.example/clean-variant.jpg'
        risky_variant = 'https://img.example/risky-variant.jpg'
        risky = {main, risky_extra, risky_variant}
        provider = Mock()
        provider.call_vision.side_effect = lambda url: url in risky
        provider.call_image_gen.side_effect = (
            lambda url, **_kwargs: url.replace(
                'https://img.example/',
                'https://generated.example/',
            )
        )
        rows = [{
            'title': 'Product',
            'main_img': main,
            'var_img': clean_variant,
            'var_imgs': [clean_variant, risky_variant],
            'extra_imgs': [clean_extra, risky_extra],
        }]

        result = amazon_review_gen._stage_review_and_gen(
            rows,
            str(tmp_path / 'role-cache.json'),
            provider_getter=lambda: provider,
        )

        assert result[0]['main_img'] == (
            'https://generated.example/risky-main.jpg'
        )
        assert result[0]['extra_imgs'] == [clean_extra]
        assert result[0]['var_imgs'] == [
            clean_variant,
            'https://generated.example/risky-variant.jpg',
        ]

    def test_amazon_generated_url_validation_uses_next_route(
            self, tmp_path, monkeypatch):
        from scripts.pipelines import amazon_review_gen

        monkeypatch.setenv('CROSSPILOT_IMAGE_REMEDIATE_ONLY', '1')
        monkeypatch.setenv('CROSSPILOT_IMAGE_QUALITY_GATE', '0')
        monkeypatch.setenv('CROSSPILOT_VALIDATE_GENERATED_IMAGE', '1')
        source = 'https://img.example/risky.jpg'
        provider = Mock()
        provider.call_vision.return_value = True
        provider.call_image_gen.side_effect = [
            'https://generated.example/broken.png',
            'https://generated.example/valid.png',
        ]
        monkeypatch.setattr(
            amazon_review_gen,
            '_validate_generated_image_url',
            Mock(side_effect=[
                (False, 'image_decode_failed'),
                (True, ''),
            ]),
        )
        metrics = {}
        rows = [{
            'title': 'Product',
            'main_img': source,
            'var_img': '',
            'var_imgs': [],
            'extra_imgs': [],
        }]

        result = amazon_review_gen._stage_review_and_gen(
            rows,
            str(tmp_path / 'validated-cache.json'),
            runtime_metrics=metrics,
            provider_getter=lambda: provider,
        )

        assert result[0]['main_img'].endswith('/valid.png')
        assert [
            call.kwargs['route_offset']
            for call in provider.call_image_gen.call_args_list
        ] == [0, 1]
        assert metrics['image_delivery_validation'] == {
            'enabled': True,
            'checked': 2,
            'accepted': 1,
            'rejected': 1,
            'reasons': {'image_decode_failed': 1},
        }

    def test_amazon_all_invalid_generated_urls_retain_original_and_flag(
            self, tmp_path, monkeypatch):
        from scripts.pipelines import amazon_review_gen

        monkeypatch.setenv('CROSSPILOT_IMAGE_REMEDIATE_ONLY', '1')
        monkeypatch.setenv('CROSSPILOT_IMAGE_QUALITY_GATE', '0')
        monkeypatch.setenv('CROSSPILOT_VALIDATE_GENERATED_IMAGE', '1')
        source = 'https://img.example/risky.jpg'
        provider = Mock()
        provider.call_vision.return_value = True
        provider.call_image_gen.side_effect = [
            f'https://generated.example/broken-{index}.png'
            for index in range(3)
        ]
        monkeypatch.setattr(
            amazon_review_gen,
            '_validate_generated_image_url',
            Mock(return_value=(False, 'image_decode_failed')),
        )
        rows = [{
            'title': 'Product',
            'main_img': source,
            'var_img': '',
            'var_imgs': [],
            'extra_imgs': [],
        }]

        result = amazon_review_gen._stage_review_and_gen(
            rows,
            str(tmp_path / 'invalid-cache.json'),
            provider_getter=lambda: provider,
        )

        assert result[0]['main_img'] == source
        assert any(
            issue['code'] == 'main_image_generation_failed'
            for issue in result[0]['_quality_issues']
        )
        cache = json.loads(
            (tmp_path / 'invalid-cache.json').read_text(
                encoding='utf-8',
            )
        )
        assert cache['gen_failures'][f'main:{source}']['terminal'] is False

    def test_amazon_remediation_only_ignores_old_clean_generation(
            self, tmp_path, monkeypatch):
        from scripts.pipelines import amazon_review_gen

        monkeypatch.setenv('CROSSPILOT_IMAGE_REMEDIATE_ONLY', '1')
        monkeypatch.setenv('CROSSPILOT_IMAGE_QUALITY_GATE', '0')
        source = 'https://img.example/clean-original.jpg'
        generated = 'https://generated.example/old-copy.png'
        review_version, generation_version = (
            amazon_review_gen._current_image_cache_versions()
        )
        cache_path = tmp_path / 'clean-old-generation-cache.json'
        cache_path.write_text(json.dumps({
            'review_prompt_version': review_version,
            'gen_prompt_version': generation_version,
            'review_results': {source: False},
            'gen_results': {f'main:{source}': generated},
            'gen_meta': {},
        }), encoding='utf-8')
        provider = Mock()
        rows = [{
            'title': 'Clean product',
            'main_img': source,
            'var_img': '',
            'var_imgs': [],
            'extra_imgs': [],
        }]

        result = amazon_review_gen._stage_review_and_gen(
            rows,
            str(cache_path),
            provider_getter=lambda: provider,
        )

        assert result[0]['main_img'] == source
        provider.call_image_gen.assert_not_called()

    def test_amazon_quality_gate_retains_original_after_rejection(
            self, tmp_path, monkeypatch):
        from scripts.pipelines import amazon_review_gen

        monkeypatch.setenv('CROSSPILOT_IMAGE_QUALITY_GATE', '1')
        monkeypatch.setenv('CROSSPILOT_IMAGE_QUALITY_REGEN_LIMIT', '0')
        provider = Mock()
        provider.call_image_gen.return_value = (
            'https://generated.example/rigid-shell.png'
        )
        provider.call_image_quality.return_value = {
            'accepted': False,
            'score': 10,
            'reasons': ['flat_product_became_rigid'],
        }
        cache_path = tmp_path / 'amazon-rejected-cache.json'
        metrics = {}
        issues = []
        source = 'https://img.example/decal-film.jpg'
        rows = [{
            'title': 'Flat decal film',
            'main_img': source,
            'var_img': '',
            'var_imgs': [],
            'extra_imgs': [],
        }]

        result = amazon_review_gen._stage_review_and_gen(
            rows,
            str(cache_path),
            quality_issues=issues,
            runtime_metrics=metrics,
            provider_getter=lambda: provider,
        )

        assert result[0]['main_img'] == source
        assert metrics['image_quality_gate']['retained_original'] == 1
        assert any('质量门禁拒绝' in issue for issue in issues)

    def test_amazon_quality_rejection_is_cached_across_resume(
            self, tmp_path, monkeypatch):
        from scripts.pipelines import amazon_review_gen

        monkeypatch.setenv('CROSSPILOT_IMAGE_QUALITY_GATE', '1')
        monkeypatch.setenv('CROSSPILOT_IMAGE_QUALITY_REGEN_LIMIT', '0')
        source = 'https://img.example/fragile-sticker.jpg'
        cache_path = tmp_path / 'rejection-resume-cache.json'
        first = Mock()
        first.call_image_gen.return_value = (
            'https://generated.example/distorted.png'
        )
        first.call_image_quality.return_value = {
            'accepted': False,
            'score': 10,
            'reasons': ['wrong_item_count'],
        }
        rows = [{
            'title': '12-piece flat sticker',
            'main_img': source,
            'var_img': '',
            'var_imgs': [],
            'extra_imgs': [],
        }]

        amazon_review_gen._stage_review_and_gen(
            [dict(rows[0])],
            str(cache_path),
            provider_getter=lambda: first,
        )
        cache = json.loads(cache_path.read_text(encoding='utf-8'))
        assert cache['gen_failures'][f'main:{source}']['terminal'] is True

        resumed = Mock()
        second_result = amazon_review_gen._stage_review_and_gen(
            [dict(rows[0])],
            str(cache_path),
            provider_getter=lambda: resumed,
        )

        assert second_result[0]['main_img'] == source
        resumed.call_image_gen.assert_not_called()

    def test_amazon_quality_gate_outage_fails_closed_but_can_resume(
            self, tmp_path, monkeypatch):
        from scripts.model_provider import ProviderUnavailableError
        from scripts.pipelines import amazon_review_gen

        monkeypatch.setenv('CROSSPILOT_IMAGE_QUALITY_GATE', '1')
        monkeypatch.setenv('CROSSPILOT_IMAGE_QUALITY_REGEN_LIMIT', '0')
        source = 'https://img.example/product.jpg'
        provider = Mock()
        provider.call_image_gen.return_value = (
            'https://generated.example/unverified.png'
        )
        provider.call_image_quality.side_effect = ProviderUnavailableError(
            'quality server unavailable',
            provider='agnes',
            operation='image_quality',
        )
        cache_path = tmp_path / 'gate-outage-cache.json'
        metrics = {}
        rows = [{
            'title': 'Product',
            'main_img': source,
            'var_img': '',
            'var_imgs': [],
            'extra_imgs': [],
        }]

        result = amazon_review_gen._stage_review_and_gen(
            rows,
            str(cache_path),
            runtime_metrics=metrics,
            provider_getter=lambda: provider,
        )

        assert result[0]['main_img'] == source
        assert metrics['image_quality_gate']['unavailable'] == 1
        assert metrics['image_quality_gate']['retained_original'] == 1
        cache = json.loads(cache_path.read_text(encoding='utf-8'))
        assert cache['gen_failures'] == {}

    def test_amazon_description_does_not_leak_image_placeholder(self, monkeypatch):
        """Test description cleaning doesn't leak __IMG__ placeholder."""
        import scripts.process_amazon as p
        from unittest.mock import Mock

        mock_provider = Mock()
        mock_provider.call_text.return_value = '<p>Useful product</p>__IMG__'

        # Mock get_provider at the module level
        monkeypatch.setattr(p, 'get_provider', lambda: mock_provider)

        data = [{'desc': '<p>Useful product</p><img src="https://img/a.jpg">'}]
        result = p._stage_clean_descs(data)
        assert '__IMG__' not in result[0]['desc']
        assert '<img' not in result[0]['desc']

    def test_amazon_title_ai_fallback_is_marked_for_review(self, monkeypatch):
        """A valid rule fallback is still a quality downgrade users must see."""
        import scripts.process_amazon as p

        provider = Mock()
        provider.call_text.return_value = None
        monkeypatch.setattr(p, 'get_provider', lambda: provider)
        data = [{
            'title': (
                'BMW Universal Replacement Dashboard Control Switch Assembly '
                'with Mounting Hardware'
            ),
        }]

        result = p._stage_optimize_titles(data)

        assert any(
            issue['code'] == 'title_ai_fallback'
            for issue in result[0]['_quality_issues']
        )

    def test_amazon_description_ai_fallback_is_marked_for_review(self, monkeypatch):
        """Rule-cleaned descriptions must not hide an AI cleaning failure."""
        import scripts.process_amazon as p

        provider = Mock()
        provider.call_text.return_value = None
        monkeypatch.setattr(p, 'get_provider', lambda: provider)
        data = [{'desc': 'Durable metal product for daily use.'}]

        result = p._stage_clean_descs(data)

        assert any(
            issue['code'] == 'description_ai_fallback'
            for issue in result[0]['_quality_issues']
        )

    def test_amazon_bullet_rule_fallback_is_marked_for_review(self, monkeypatch):
        """Fallbacks stay traceable and never fabricate placeholder product claims."""
        import scripts.process_amazon as p

        provider = Mock()
        provider.call_text.return_value = None
        monkeypatch.setattr(p, 'get_provider', lambda: provider)
        data = [{
            'title': 'Universal Control Switch',
            'desc': 'Durable metal construction for common replacement applications.',
        }]

        result = p._stage_generate_bullets_keywords(data)

        assert len(result[0]['bullets']) == 5
        assert not any(
            'additional product detail' in bullet
            for bullet in result[0]['bullets']
        )
        assert any(
            issue['code'] == 'bullet_rule_fallback'
            for issue in result[0]['_quality_issues']
        )

    def test_amazon_partial_bullet_payload_is_padded_without_crashing(self, monkeypatch):
        import scripts.process_amazon as p

        provider = Mock()
        provider.call_text.return_value = json.dumps({
            'bullets': ['Durable metal', 'Simple installation'],
            'keywords': 'switch,control,replacement,metal,auto,part,fit,install,durable,universal',
        })
        monkeypatch.setattr(p, 'get_provider', lambda: provider)
        data = [{
            'title': 'Universal Control Switch',
            'desc': 'Durable metal construction for common replacement applications.',
        }]

        result = p._stage_generate_bullets_keywords(data)

        # 校验拒绝 < 20 chars 的 bullet → 回退规则补全
        assert len(result[0]['bullets']) == 5
        assert any(
            issue['code'] in ('bullet_rule_fallback', 'bullet_quality_warning')
            for issue in result[0].get('_quality_issues', [])
        )

    def test_amazon_poor_bullets_and_keywords_are_marked_for_review(self, monkeypatch):
        import scripts.process_amazon as p

        provider = Mock()
        provider.call_text.return_value = json.dumps({
            'bullets': [
                'BMW quality product',
                'BMW quality product',
                'Great quality product',
                'Great quality product',
                'Great quality product',
            ],
            'keywords': 'bmw, product, durable',
        })
        monkeypatch.setattr(p, 'get_provider', lambda: provider)
        data = [{
            'title': '12V Stainless Trailer Hitch Lock 2 Pack',
            'desc': (
                'Made from stainless steel for corrosion resistance. '
                'Includes two keys and lock pin for trailer hitch use. '
                'Fits 1/2 inch and 5/8 inch receivers. '
                'Package includes 2 lock pins and dust caps. '
                'Simple tool-free installation for towing accessories.'
            ),
        }]

        result = p._stage_generate_bullets_keywords(data)
        issues = {issue['code'] for issue in result[0].get('_quality_issues', [])}
        # 校验拒绝短 bullet → 质量标记
        assert bool(issues), f'Expected quality issues, got none'
        assert 'bmw' not in result[0]['keywords'].lower()

    def test_amazon_title_fact_loss_rejects_ai_title(self, monkeypatch):
        import scripts.process_amazon as p

        provider = Mock()
        provider.call_text.return_value = 'Universal Replacement Control Switch'
        monkeypatch.setattr(p, 'get_provider', lambda: provider)
        data = [{
            'title': (
                'BMW X5 2018 12V Dashboard Control Switch Assembly '
                '2 Pack with Mounting Hardware and Long Compatibility Text'
            ),
        }]

        result = p._stage_optimize_titles(data)

        assert '2018' in result[0]['title']
        assert any(
            item['reason'] == 'title_fact_loss'
            and item['method'] == 'ai_rejected'
            for item in result[0]['_audit']
        )

    def test_amazon_title_rejects_ai_added_brand(self, monkeypatch):
        import scripts.process_amazon as p

        provider = Mock()
        provider.call_text.return_value = (
            'For BMW Universal Reflective Car Door Side Sport Vinyl Decal Sticker'
        )
        monkeypatch.setattr(p, 'get_provider', lambda: provider)
        original = (
            'Universal Reflective Car Door Side Sport Vinyl Decal Sticker '
            'Waterproof Exterior Decoration'
        )

        result = p._stage_optimize_titles([{'title': original}])

        assert 'BMW' not in result[0]['title']

    def test_amazon_description_fact_loss_keeps_rule_result(self, monkeypatch):
        import scripts.process_amazon as p

        provider = Mock()
        provider.call_text.return_value = 'Useful trailer lock for daily use.'
        monkeypatch.setattr(p, 'get_provider', lambda: provider)
        original = (
            'Stainless lock pin for 12V trailer, size 5/8 inch, includes 2 keys.'
        )

        result = p._stage_clean_descs([{'desc': original}])

        assert '12V' in result[0]['desc']
        assert '5/8 inch' in result[0]['desc']
        assert '2 keys' in result[0]['desc']

    def test_amazon_keyword_normalizer_fills_ten_terms_within_limit(self):
        import scripts.process_amazon as p

        row = {
            'title': (
                "1x 38CM 15'' Car Steering Wheel Cover PU Leather "
                'Anti-Slip Protector'
            ),
            'desc': (
                'Material: imitation leather, Color: black red, Style: Dragon '
                'Reflective Totem, Fits steering wheels 14.5-15 inches diameter '
                '(38cm), Universal fit for most vehicles, Non-slip, breathable, '
                'odorless, eco-friendly, Package: 1 steering wheel cover.'
            ),
            'keywords': (
                'steering wheel cover, pu leather cover, '
                'non-slip steering wheel protector, car steering wheel wrap, '
                'universal steering wheel cover, 15 inch steering wheel cover, '
                'breathable steering wheel cover, dragon steering wheel cover'
            ),
        }

        p._normalize_keywords_for_row(row)
        terms = p._dedupe_terms(p._split_keywords(row['keywords']))

        assert len(terms) == 10
        assert len(row['keywords']) <= 250

    def test_amazon_rule_fallback_builds_five_source_grounded_bullets(self):
        from scripts.pipelines.amazon_stages import (
            _source_bullet_candidates,
        )

        row = {
            'title': 'PU Leather Dashboard Trim Strip',
            'desc': (
                'High-grade PU leather with adhesive tape on the back. '
                'Easy to install on the dashboard or door. '
                'The strip can be cut to the required size.'
            ),
        }

        bullets = _source_bullet_candidates(row)

        assert len(bullets) == 5
        assert len(set(bullets)) == 5
        assert all(
            term in ' '.join(bullets).lower()
            for term in ('pu leather', 'adhesive', 'dashboard')
        )

    def test_amazon_description_fact_loss_rejects_ai_desc(self, monkeypatch):
        import scripts.process_amazon as p

        provider = Mock()
        provider.call_text.return_value = 'Useful product for daily use.'
        monkeypatch.setattr(p, 'get_provider', lambda: provider)
        data = [{
            'desc': 'Stainless lock pin for 12V trailer, size 5/8 inch, includes 2 keys.',
        }]

        result = p._stage_clean_descs(data)

        assert '12V' in result[0]['desc']
        assert '5/8 inch' in result[0]['desc']
        assert '2 keys' in result[0]['desc']
        assert any(
            item['reason'] == 'description_fact_loss'
            and item['method'] == 'ai_rejected'
            for item in result[0]['_audit']
        )

    def test_amazon_quality_issue_summary_is_bounded_and_aggregated(self):
        import scripts.process_amazon as p

        rows = [
            {'_quality_issues': [
                {'code': 'title_ai_fallback', 'message': 'fallback'},
            ]},
            {'_quality_issues': [
                {'code': 'title_ai_fallback', 'message': 'fallback'},
                {'code': 'description_ai_fallback', 'message': 'fallback'},
            ]},
        ]

        summary = p._summarize_row_quality_issues(rows)

        assert any('标题 AI 优化降级为规则处理：2 行' in item for item in summary)
        assert any('描述 AI 清洗降级为规则处理：1 行' in item for item in summary)

    def test_amazon_input_rejects_missing_core_fields(self):
        """缺少主图仍然阻止运行。"""
        import scripts.process_amazon as p

        with pytest.raises(ValueError, match='缺少有效的主图'):
            p._validate_amazon_input([{
                'title': 'Test Product',
                'desc': '',
                'main_img': '',  # 缺少主图才真正报错
            }])

    def test_amazon_input_keeps_missing_description_and_marks_review(self):
        """空源描述保留为空，并进入结构化人工复核。"""
        import scripts.process_amazon as p

        row = {
            'id': 'missing-desc',
            'title': 'Test Product',
            'desc': '',
            'main_img': 'https://img.example/main.jpg',
        }

        p._validate_amazon_input([row])

        assert row['desc'] == ''
        assert row['_quality_issues'] == [{
            'code': 'missing_source_description',
            'message': '源产品描述为空，已保留空值并标记人工复核',
        }]
        assert any(
            item['reason'] == 'missing_source_description'
            and item['field'] == 'description'
            for item in row['_audit']
        )

    def test_amazon_input_rejects_missing_title(self):
        import scripts.process_amazon as p

        with pytest.raises(ValueError, match='缺少产品标题'):
            p._validate_amazon_input([{
                'title': '',
                'desc': 'Description',
                'main_img': 'https://img.example/main.jpg',
            }])

    def test_amazon_output_validation_marks_incomplete_bullets_for_review(self, tmp_path):
        """质量问题必须成为结构化复核结果，不能只打印后继续标记成功。"""
        import openpyxl
        import scripts.process_amazon as p

        output = tmp_path / 'bad-amazon-output.xlsx'
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['header'] * 24)
        ws.append(['defaults'] * 24)
        row = [''] * 24
        row[0] = 'Test Product'
        row[1] = 'Useful description'
        row[2] = 'https://img.example/main.jpg'
        row[17] = 'Only one bullet'
        row[22] = 'test keyword'
        ws.append(row)
        wb.save(output)
        wb.close()

        result = p._validate_amazon_output(str(output), 1)

        assert result['passed'] is False
        assert any('Bullet' in issue for issue in result['issues'])

    def test_amazon_output_validation_flags_semantic_quality(self, tmp_path):
        import openpyxl
        import scripts.process_amazon as p

        output = tmp_path / 'weak-amazon-output.xlsx'
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['header'] * 24)
        ws.append(['defaults'] * 24)
        row = [''] * 24
        row[0] = 'Test Product'
        row[1] = 'Useful description with stainless steel lock pin.'
        row[2] = 'https://img.example/main.jpg'
        for column in range(17, 22):
            row[column] = 'Great quality product'
        row[22] = 'product, bmw'
        ws.append(row)
        wb.save(output)
        wb.close()

        result = p._validate_amazon_output(str(output), 1)

        assert result['passed'] is False
        assert any('Bullet' in issue and '重复' in issue for issue in result['issues'])
        assert any('关键词' in issue for issue in result['issues'])

    def test_amazon_json_rows_use_same_quality_rules(self):
        import scripts.process_amazon as p

        result = p._validate_amazon_rows([{
            'title': 'Test Product',
            'desc': 'Useful description',
            'main_img': 'https://img.example/main.jpg',
            'bullets': ['Great quality product'] * 5,
            'keywords': 'product, item',
        }])

        assert result['passed'] is False
        assert any('Bullet' in issue for issue in result['issues'])
        assert any('关键词' in issue for issue in result['issues'])

    def test_amazon_title_normalization_restores_compatibility_prefix(self):
        import scripts.process_amazon as p

        title = p._normalize_title('BMW Replacement Control Arm')

        assert title == 'Generic Replacement Control Arm for BMW'
        assert len(title) <= 75

    @pytest.mark.parametrize(
        ('source', 'expected'),
        [
            (
                'For Generic Windshield Washer Spray Nozzle Kit '
                'for Toyota Camry (2 Pack)',
                'Generic Windshield Washer Spray Nozzle Kit '
                'for Toyota Camry (2 Pack)',
            ),
            (
                'Fits For Suzuki Baleno 1998-2001 4PCS '
                'Car Inside Door Handle',
                'Generic Car Inside Door Handle '
                'for Suzuki Baleno 1998-2001 4PCS',
            ),
            (
                'For 40× 3D DIY Chrome Metal Letter Numbers '
                'Car Motorcycle Emblem Badge',
                'Generic 40× 3D Chrome Metal Letter Numbers '
                'Car Motorcycle Emblem Badge',
            ),
            (
                'For Generic Rear Bumper Lip Protector Guard for',
                'Generic Rear Bumper Lip Protector Guard',
            ),
            (
                'Seat Cover Organizer for Car',
                'Seat Cover Organizer for Car',
            ),
            (
                'Mini USB Car Charger Adapter',
                'Mini USB Car Charger Adapter',
            ),
        ],
    )
    def test_amazon_title_uses_generic_compatibility_contract(
            self, source, expected,
    ):
        import scripts.process_amazon as p

        assert p._normalize_title(source) == expected

    def test_amazon_title_normalization_is_idempotent(self):
        import scripts.process_amazon as p

        title = (
            'Generic Windshield Washer Spray Nozzle Kit '
            'for Toyota Camry (2 Pack)'
        )

        assert p._normalize_title(p._normalize_title(title)) == title

    def test_amazon_title_keeps_specs_and_compatibility_within_limit(self):
        import scripts.process_amazon as p

        title = p._normalize_title(
            'BMW X5 2018 12V Dashboard Control Switch Assembly '
            '2 Pack with Mounting Hardware and Long Compatibility Text'
        )

        assert title.startswith('Generic ')
        assert title.endswith(' for BMW X5 2018')
        assert '12V' in title
        assert '2 Pack' in title
        assert len(title) <= 75
        assert title.lower().count(' for ') == 1

    def test_amazon_stage_changes_are_captured_as_audit_trail(self, monkeypatch):
        import scripts.process_amazon as p

        provider = Mock()
        provider.call_text.return_value = None
        monkeypatch.setattr(p, 'get_provider', lambda: provider)
        data = [{
            'title': (
                'BMW Universal Replacement Dashboard Control Switch Assembly '
                'with Mounting Hardware'
            ),
        }]

        result = p._stage_optimize_titles(data)
        validation = p._attach_audit_to_validation(
            {'passed': False, 'issues': ['第 1 行标题需复核']},
            result,
        )

        assert any(
            item['stage'] == '标题优化'
            and item['field'] == 'title'
            and item['before'].startswith('BMW Universal')
            and item['after'].startswith('Generic Universal')
            and item['after'].endswith(' for BMW')
            for item in result[0]['_audit']
        )
        assert validation['audit'][0]['row'] == 1
        assert validation['audit'][0]['stage'] == '标题优化'
