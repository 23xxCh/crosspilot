"""FastAPI app：上传/历史/详情/SSE/下载/设置。复用 process_ebay_tk 管道。"""
from contextlib import asynccontextmanager
import base64, binascii, errno, os, json, uuid, shutil, time, re, secrets, zipfile, glob as _glob
from urllib.parse import quote
from scripts.pipeline_log import log as _log

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

from fastapi import FastAPI, UploadFile, File, HTTPException, Request, Query
from fastapi.responses import FileResponse, StreamingResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from starlette.middleware.trustedhost import TrustedHostMiddleware
from web import store, jobs
from web.api_v1 import router as api_v1_router
from web.routes.dashboard import router as dashboard_router
from web.routes.settings import router as settings_router
from web.routes.tasks import router as tasks_router
from web.review_report import (
    build_quality_score,
    build_review_data,
    build_review_report_csv,
)


@asynccontextmanager
async def _lifespan(_app):
    try:
        retention_days = int(os.environ.get('CROSSPILOT_RETENTION_DAYS', '7'))
    except ValueError:
        retention_days = 7
    if retention_days > 0:
        try:
            cleaned = store.cleanup_old_tasks(retention_days)
            if cleaned:
                print(f"[cleanup] 已清理 {cleaned} 个过期任务", flush=True)
        except Exception as exc:
            _log.warn("过期任务清理失败", error=str(exc))
    jobs.start_monitor()
    try:
        yield
    finally:
        jobs.stop_monitor(cancel_running=True)


app = FastAPI(title="CrossPilot", lifespan=_lifespan)

app.include_router(dashboard_router)
app.include_router(settings_router)
app.include_router(tasks_router)
app.include_router(api_v1_router)

_ALLOWED_HOSTS = [
    item.strip()
    for item in os.environ.get(
        'CROSSPILOT_ALLOWED_HOSTS',
        '127.0.0.1,localhost,[::1],::1,testserver',
    ).split(',')
    if item.strip()
]
_ALLOWED_ORIGINS = {
    item.strip().rstrip('/')
    for item in os.environ.get('CROSSPILOT_ALLOWED_ORIGINS', '').split(',')
    if item.strip()
}
app.add_middleware(TrustedHostMiddleware, allowed_hosts=_ALLOWED_HOSTS)

_SECURITY_HEADERS = {
    'Content-Security-Policy': (
        "default-src 'self'; script-src 'self'; style-src 'self' 'unsafe-inline'; "
        "img-src 'self' data: https:; connect-src 'self'; font-src 'self' data:; "
        "object-src 'none'; base-uri 'none'; frame-ancestors 'none'; form-action 'self'"
    ),
    'X-Frame-Options': 'DENY',
    'X-Content-Type-Options': 'nosniff',
    'Referrer-Policy': 'no-referrer',
    'Permissions-Policy': 'camera=(), microphone=(), geolocation=()',
}


def _secure_response(response):
    for name, value in _SECURITY_HEADERS.items():
        response.headers[name] = value
    return response

@app.middleware("http")
async def _optional_basic_auth(request, call_next):
    """设置 CROSSPILOT_AUTH_PASSWORD 后，为整个 Web 应用启用 HTTP Basic。"""
    expected_password = os.environ.get('CROSSPILOT_AUTH_PASSWORD', '')
    if expected_password:
        expected_user = os.environ.get('CROSSPILOT_AUTH_USER', 'crosspilot')
        auth = request.headers.get('Authorization', '')
        valid = False
        if auth.startswith('Basic '):
            try:
                user, password = base64.b64decode(
                    auth[6:], validate=True
                ).decode('utf-8').split(':', 1)
                valid = (
                    secrets.compare_digest(user, expected_user)
                    and secrets.compare_digest(password, expected_password)
                )
            except (binascii.Error, ValueError, UnicodeDecodeError):
                pass
        if not valid:
            return JSONResponse(
                {'detail': 'Authentication required'},
                status_code=401,
                headers={'WWW-Authenticate': 'Basic realm="CrossPilot"'},
            )
    return await call_next(request)

# 强制禁止浏览器缓存静态文件（HTML/JS 更新后前端立刻生效）
@app.middleware("http")
async def _no_cache_static(request, call_next):
    resp = await call_next(request)
    path = request.url.path
    if path.endswith(('.html', '.js', '.css')) or path == '/' or path == '':
        resp.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        resp.headers['Pragma'] = 'no-cache'
        resp.headers['Expires'] = '0'
    return resp


@app.middleware("http")
async def _browser_security(request: Request, call_next):
    """阻止跨站写操作，并为本地管理界面补齐浏览器安全边界。"""
    if request.method in {'POST', 'PUT', 'PATCH', 'DELETE'}:
        origin = (request.headers.get('Origin') or '').rstrip('/')
        fetch_site = (request.headers.get('Sec-Fetch-Site') or '').lower()
        host = request.headers.get('Host', '')
        allowed_origins = {f'http://{host}', f'https://{host}'} | _ALLOWED_ORIGINS
        if fetch_site == 'cross-site' or (origin and origin not in allowed_origins):
            return _secure_response(
                JSONResponse({'detail': 'Cross-site request blocked'}, status_code=403)
            )
        if origin and request.headers.get('X-CrossPilot-Request') != '1':
            return _secure_response(
                JSONResponse(
                    {'detail': 'Missing same-origin request marker'},
                    status_code=403,
                )
            )

    response = await call_next(request)
    return _secure_response(response)

_WIN_RESERVED = {n.upper() for n in ['CON','PRN','AUX','NUL'] + [f'COM{i}' for i in range(1,10)] + [f'LPT{i}' for i in range(1,10)]}

def _safe_filename(filename):
    """过滤 path traversal 和危险字符。"""
    name = os.path.basename(filename or '')
    name = re.sub(r'[<>:"/\\|?*\x00-\x1f]', '_', name).rstrip(' .')
    extension = os.path.splitext(name)[1].lower()
    if extension not in {'.xlsx', '.json'}:
        return None
    stem = os.path.splitext(name)[0].rstrip(' .')
    if not name or not stem or stem.split('.')[0].upper() in _WIN_RESERVED:
        return None
    return stem[:180] + extension
# 支持 PyInstaller 打包后 APPDATA 路径
DATA_DIR = os.environ.get('CROSSPILOT_DATA_DIR') or os.path.join(ROOT, 'data')
KEYS_PATH = os.environ.get('CROSSPILOT_KEYS_PATH') or os.path.join(ROOT, 'keys.json')
UPLOAD_DIR = os.path.join(DATA_DIR, 'uploads')
os.makedirs(UPLOAD_DIR, exist_ok=True)
KEYS_EXAMPLE = os.path.join(ROOT, 'keys.example.json') if os.path.exists(os.path.join(ROOT, 'keys.example.json')) else None


# ===== 仪表盘 =====


# ===== 批量上传 =====
MAX_UPLOAD_SIZE = max(
    1, int(os.environ.get('CROSSPILOT_MAX_UPLOAD_MB', '50'))
) * 1024 * 1024
MAX_BATCH_FILES = max(
    1, min(int(os.environ.get('CROSSPILOT_MAX_BATCH_FILES', '20')), 100)
)

_PUBLIC_TASK_FIELDS = (
    'id', 'filename', 'pipeline', 'status', 'stage', 'stage_index',
    'stage_total', 'current', 'total', 'percent', 'eta_s', 'error',
    'created_at', 'updated_at', 'stats', 'total_elapsed_s', 'quality_score',
)
_PUBLIC_PROGRESS_FIELDS = (
    'status', 'stage', 'stage_index', 'stage_total', 'current', 'total',
    'percent', 'eta_s', 'error', 'total_elapsed_s', 'updated_at',
)


def _public_task(task):
    """只向浏览器返回展示字段，隐藏本地路径和内部缓存。"""
    public = {
        key: task.get(key)
        for key in _PUBLIC_TASK_FIELDS
        if key in task
    }
    if not public.get('quality_score'):
        public['quality_score'] = build_quality_score(public)
    return public


def _public_progress(progress):
    return {
        key: progress.get(key)
        for key in _PUBLIC_PROGRESS_FIELDS
        if key in progress
    }


def _review_value(value, limit=240):
    """Short display-only value for review context; never include local paths."""
    if isinstance(value, (list, tuple)):
        text = ' | '.join(str(item or '').strip() for item in value if str(item or '').strip())
    else:
        text = str(value or '')
    text = re.sub(r'\s+', ' ', text).strip()
    return text if len(text) <= limit else text[:limit - 1] + '…'


def _first_image(value):
    if isinstance(value, list):
        values = value
    else:
        values = str(value or '').replace('\r', '').split('\n')
    for item in values:
        text = str(item or '').strip()
        if text.startswith(('http://', 'https://')):
            return text
    return ''


def _review_field(current):
    current = _review_value(current)
    return {'current': current} if current else {}


def _read_amazon_json_review_rows(path, *, output=False):
    from scripts.services.amazon_json import load_columnar_json
    payload = load_columnar_json(path)
    titles = payload.get('产品标题', [])
    descriptions = payload.get('产品描述', [])
    product_images = payload.get('产品图片链接', [])
    rows = []
    for index, title in enumerate(titles, start=1):
        fields = {
            'title': _review_field(title),
            'description': _review_field(descriptions[index - 1] if index - 1 < len(descriptions) else ''),
            'main_image': _review_field(_first_image(product_images[index - 1] if index - 1 < len(product_images) else '')),
        }
        if output:
            bullets = [
                (payload.get(f'Bullet Point{bullet_index}', []) or [''])[index - 1]
                if index - 1 < len(payload.get(f'Bullet Point{bullet_index}', []))
                else ''
                for bullet_index in range(1, 6)
            ]
            fields['Bullet'] = _review_field([bullet for bullet in bullets if bullet])
            keywords = payload.get('关键词信息', [])
            fields['keywords'] = _review_field(keywords[index - 1] if index - 1 < len(keywords) else '')
        rows.append({
            'data_row': index,
            'row': index,
            'output_row' if output else 'source_row': index,
            'fields': {key: value for key, value in fields.items() if value},
        })
    return rows


def _read_source_xlsx_review_rows(path):
    import openpyxl
    from scripts.adapters import detect_adapter

    wb = None
    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        adapter = detect_adapter(wb.active)
        ws = wb[adapter.sheet_name] if adapter and adapter.sheet_name and adapter.sheet_name in wb.sheetnames else wb.active
        cols = adapter.cols if adapter else {}
        rows = []
        for sheet_row in range(2, ws.max_row + 1):
            fields = {
                'title': _review_field(ws.cell(sheet_row, cols.get('title', 2)).value),
                'description': _review_field(ws.cell(sheet_row, cols.get('desc', 3)).value),
            }
            main_col = cols.get('main_image')
            if main_col:
                fields['main_image'] = _review_field(_first_image(ws.cell(sheet_row, main_col).value))
            rows.append({
                'data_row': sheet_row - 1,
                'row': sheet_row,
                'source_row': sheet_row,
                'fields': {key: value for key, value in fields.items() if value},
            })
        return rows
    finally:
        if wb is not None:
            wb.close()


def _read_output_xlsx_review_rows(path):
    import openpyxl

    wb = None
    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        ws = wb.active
        rows = []
        for sheet_row in range(3, ws.max_row + 1):
            bullets = [ws.cell(sheet_row, col).value for col in range(18, 23)]
            fields = {
                'title': _review_field(ws.cell(sheet_row, 1).value),
                'description': _review_field(ws.cell(sheet_row, 2).value),
                'main_image': _review_field(_first_image(ws.cell(sheet_row, 3).value)),
                'Bullet': _review_field([bullet for bullet in bullets if bullet]),
                'keywords': _review_field(ws.cell(sheet_row, 23).value),
            }
            rows.append({
                'data_row': sheet_row - 2,
                'row': sheet_row,
                'output_row': sheet_row,
                'fields': {key: value for key, value in fields.items() if value},
            })
        return rows
    finally:
        if wb is not None:
            wb.close()


def _merge_review_context(source_rows, output_rows):
    merged = {}
    for item in source_rows or []:
        data_row = item.get('data_row')
        context = merged.setdefault(data_row, {'data_row': data_row, 'fields': {}})
        context['source_row'] = item.get('source_row') or item.get('row')
        context.setdefault('row', item.get('row'))
        for field, value in (item.get('fields') or {}).items():
            current = _review_value(value.get('current') if isinstance(value, dict) else value)
            if current:
                context['fields'].setdefault(field, {})['original'] = current
    for item in output_rows or []:
        data_row = item.get('data_row')
        context = merged.setdefault(data_row, {'data_row': data_row, 'fields': {}})
        context['output_row'] = item.get('output_row') or item.get('row')
        context['row'] = item.get('row')
        for field, value in (item.get('fields') or {}).items():
            current = _review_value(value.get('current') if isinstance(value, dict) else value)
            if current:
                context['fields'].setdefault(field, {})['processed'] = current
    for context in merged.values():
        title_field = context.get('fields', {}).get('title', {})
        context['title'] = title_field.get('processed') or title_field.get('original') or ''
    return [merged[key] for key in sorted(key for key in merged if key is not None)]


def _review_context_rows(task):
    """Best-effort row context for review UI/CSV; all filesystem details stay server-side."""
    if task.get('pipeline') != 'amazon':
        return []
    input_path = task.get('input_path') or ''
    output_path = task.get('output_path') or ''
    source_rows = []
    output_rows = []
    try:
        if input_path and os.path.exists(input_path):
            if input_path.lower().endswith('.json'):
                source_rows = _read_amazon_json_review_rows(input_path)
            elif input_path.lower().endswith('.xlsx'):
                source_rows = _read_source_xlsx_review_rows(input_path)
    except Exception as exc:
        _log.warn('复核上下文读取输入失败', error=str(exc))
    try:
        if output_path and os.path.exists(output_path):
            if output_path.lower().endswith('.json'):
                output_rows = _read_amazon_json_review_rows(output_path, output=True)
            elif output_path.lower().endswith('.xlsx'):
                output_rows = _read_output_xlsx_review_rows(output_path)
    except Exception as exc:
        _log.warn('复核上下文读取输出失败', error=str(exc))
    return _merge_review_context(source_rows, output_rows)


async def _save_upload(file: UploadFile, destination: str):
    """流式保存并验证 XLSX 或 Amazon 列式 JSON。"""
    written = 0
    try:
        with open(destination, 'wb') as out:
            while chunk := await file.read(1024 * 1024):
                written += len(chunk)
                if written > MAX_UPLOAD_SIZE:
                    raise ValueError(f'文件超过 {MAX_UPLOAD_SIZE // 1024 // 1024}MB 上限')
                out.write(chunk)
    except Exception:
        try:
            os.remove(destination)
        except OSError:
            pass
        raise
    extension = os.path.splitext(destination)[1].lower()
    validation_error = None
    if written == 0:
        validation_error = '文件为空'
    elif extension == '.xlsx' and not zipfile.is_zipfile(destination):
        validation_error = '文件不是有效的 .xlsx 工作簿'
    elif extension == '.json':
        try:
            from scripts.services.amazon_json import load_columnar_json
            max_rows = max(
                1, int(os.environ.get('CROSSPILOT_MAX_ROWS', '10000'))
            )
            load_columnar_json(destination, max_rows=max_rows)
        except ValueError as exc:
            validation_error = str(exc)

    if validation_error:
        try:
            os.remove(destination)
        except OSError:
            pass
        raise ValueError(validation_error)
    return written


@app.post("/api/upload/batch")
async def upload_batch(files: list[UploadFile] = File(...)):
    if len(files) > MAX_BATCH_FILES:
        raise HTTPException(413, f'单次最多上传 {MAX_BATCH_FILES} 个文件')
    results = []
    for file in files:
        extension = os.path.splitext(file.filename or '')[1].lower()
        if extension not in {'.xlsx', '.json'}:
            results.append({
                'filename': file.filename,
                'error': '只接受 .xlsx 或 Amazon .json 文件',
            })
            continue
        safe_name = _safe_filename(file.filename)
        if not safe_name:
            results.append({'filename': file.filename, 'error': '文件名无效'})
            continue
        job_id = uuid.uuid4().hex[:12]
        job_dir = os.path.join(UPLOAD_DIR, job_id)
        os.makedirs(job_dir, exist_ok=True)
        in_path = os.path.join(job_dir, safe_name)
        try:
            await _save_upload(file, in_path)
        except ValueError as e:
            shutil.rmtree(job_dir, ignore_errors=True)
            results.append({'filename': file.filename, 'error': str(e)})
            continue
        except OSError as e:
            shutil.rmtree(job_dir, ignore_errors=True)
            results.append({'filename': file.filename, 'error': f'文件保存失败: {e}'})
            continue
        store.create(job_id, safe_name, in_path)
        if jobs.enqueue(job_id, in_path):
            results.append({'job_id': job_id, 'filename': file.filename})
        else:
            state = store.get(job_id) or {}
            results.append({
                'job_id': job_id,
                'filename': file.filename,
                'error': state.get('error') or '任务无法入队',
            })
    return {'results': results, **jobs.queue_snapshot()}


# ===== 任务每行水印/生图明细 =====


# ===== 模板列表（来源适配器注册表） =====


# ===== 管道阶段（前端动态读取，避免硬编码不同步） =====


# ===== 版本 & 更新 =====


# ===== 单文件上传 =====
@app.post("/api/upload")
async def upload(file: UploadFile = File(...)):
    extension = os.path.splitext(file.filename or '')[1].lower()
    if extension not in {'.xlsx', '.json'}:
        raise HTTPException(400, "只接受 .xlsx 或 Amazon .json")
    safe_name = _safe_filename(file.filename)
    if not safe_name:
        raise HTTPException(400, "文件名无效")
    job_id = uuid.uuid4().hex[:12]
    job_dir = os.path.join(UPLOAD_DIR, job_id)
    os.makedirs(job_dir, exist_ok=True)
    in_path = os.path.join(job_dir, safe_name)
    try:
        await _save_upload(file, in_path)
    except ValueError as e:
        shutil.rmtree(job_dir, ignore_errors=True)
        status = 413 if '超过' in str(e) else 400
        raise HTTPException(status, str(e))
    except OSError as e:
        shutil.rmtree(job_dir, ignore_errors=True)
        raise HTTPException(500, f"文件保存失败: {e}")
    store.create(job_id, safe_name, in_path)
    if not jobs.enqueue(job_id, in_path):
        state = store.get(job_id) or {}
        raise HTTPException(503, state.get('error') or "任务无法入队")
    return {'job_id': job_id}






def _sync_progress(job_id, input_path):
    """从 status.json 同步真实进度到 SQLite（公开 helper，task_detail/task_events 复用）。"""
    if not input_path:
        return None
    st = jobs.read_status(input_path)
    if st and st.get('status') not in ('done', 'needs_review', 'failed', 'cancelled'):
        store.update_progress(job_id, st)
    return st


app.mount("/", StaticFiles(directory=os.path.join(os.path.dirname(__file__), 'static'),
                          html=True), name="static")
