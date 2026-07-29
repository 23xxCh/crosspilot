"""Task routes — 列表、详情、SSE、下载、审核报告、重试、取消。"""
import os, json, time
from queue import Empty
from fastapi import APIRouter, HTTPException, Request, Query
from fastapi.responses import FileResponse, StreamingResponse
from web import store, jobs
from web.context import ctx
from web.routes._helpers import public_task, public_progress, sync_progress

router = APIRouter(tags=["tasks"])


def _review_ctx(task):
    """惰性导入 _review_context_rows（依赖链复杂，未迁移至 _helpers）。"""
    from web.app import _review_context_rows
    return _review_context_rows(task)


@router.get("/api/tasks/{job_id}/rows")
def task_rows(job_id: str):
    t = store.get(job_id)
    if not t:
        raise HTTPException(404, "任务不存在")
    input_path = t.get('input_path', '')
    c = jobs.read_cache(input_path)
    review = c.get('review_results', {})
    generated = c.get('gen_results', {})

    rows = []
    wb = None
    try:
        import openpyxl
        from scripts.adapters import detect_adapter
        wb = openpyxl.load_workbook(input_path, data_only=True)
        adapter = detect_adapter(wb.active)
        ws = wb[adapter.sheet_name] if adapter and adapter.sheet_name and adapter.sheet_name in wb.sheetnames else wb.active
        cols = adapter.cols if adapter else {
            'title': 2, 'main_image': 18, 'attachments': list(range(19, 27)),
            'variant': 29, 'size_image': 28,
        }
        title_col = cols['title']
        main_col = cols['main_image']
        att_cols = cols.get('attachments', [])
        variant_col = cols.get('variant')
        size_col = cols.get('size_image')

        for r in range(2, ws.max_row + 1):
            title = str(ws.cell(r, title_col).value or '').strip()
            img_urls = []
            main_raw = str(ws.cell(r, main_col).value or '').strip()
            main_values = [
                u.strip() for u in main_raw.replace('\r', '').split('\n')
                if u.strip().startswith('http')
            ]
            main_url = main_values[0] if main_values else ''
            if main_url:
                img_urls.append(('main', main_url))
            img_urls.extend(('attachment', u) for u in main_values[1:])
            for ac in att_cols:
                au = str(ws.cell(r, ac).value or '').strip()
                if au and au.startswith('http'):
                    img_urls.append(('attachment', au))
            if variant_col:
                variant_raw = str(ws.cell(r, variant_col).value or '').strip()
                variant_values = [
                    u.strip() for u in variant_raw.replace('\r', '').split('\n')
                    if u.strip().startswith('http')
                ]
                img_urls.extend(('variant', u) for u in variant_values)
            if size_col:
                su = str(ws.cell(r, size_col).value or '').strip()
                if su and su.startswith('http'):
                    img_urls.append(('size', su))

            wm_count = sum(1 for _, u in img_urls if review.get(u) is True)

            def _generated_url(kind, url):
                return generated.get(url) or generated.get(f'{kind}:{url}')

            gen_count = sum(1 for kind, u in img_urls if _generated_url(kind, u))
            wm_main = None
            if main_url and review.get(main_url) is True:
                wm_main = main_url
            gen_main = _generated_url('main', main_url) if main_url else None

            rows.append({
                'row': r - 1,
                'title': title[:120],
                'main_image': main_url,
                'wm_main': wm_main,
                'gen_main': gen_main,
                'image_count': len(img_urls),
                'watermark_count': wm_count,
                'generated_count': gen_count,
                'all_clean': wm_count == 0,
            })
    except Exception as e:
        print(f"[WARN] xlsx row parse failed: {e}")
    finally:
        if wb is not None:
            wb.close()
    return {'rows': rows, 'total': len(rows), 'filename': t['filename']}


@router.get("/api/tasks")
def list_tasks(
        page: int = 1, limit: int = 20,
        task_filter: str = Query('all', alias='filter'),
        sort: str = 'created_desc',
):
    page = max(1, page)
    limit = max(1, min(limit, 100))
    task_filter = store.normalize_task_filter(task_filter)
    task_sort = store.normalize_task_sort(sort)
    offset = (page - 1) * limit
    total = store.count_tasks(task_filter=task_filter)
    tasks = [
        public_task(task)
        for task in store.list_tasks(limit=limit, offset=offset,
                                      task_filter=task_filter, task_sort=task_sort)
    ]
    return {
        'tasks': tasks, 'total': total, 'page': page,
        'filter': task_filter, 'sort': task_sort,
        'pages': max(1, (total + limit - 1) // limit),
    }


@router.get("/api/tasks/{job_id}")
def task_detail(job_id: str):
    t = store.get(job_id)
    if not t:
        raise HTTPException(404, "任务不存在")
    if t.get('status') == 'running':
        st = sync_progress(job_id, t.get('input_path'))
        if st:
            t.update({k: st.get(k) for k in [
                'stage', 'stage_index', 'stage_total', 'current', 'total',
                'percent', 'eta_s'
            ] if k in st})
        if t.get('created_at'):
            t['total_elapsed_s'] = max(0, int(time.time() - t['created_at']))
    return public_task(t)


@router.get("/api/tasks/{job_id}/events")
def task_events(job_id: str):
    q = jobs.subscribe(job_id)

    def gen():
        _SSE_MAX_S = 1800
        _started = time.time()
        try:
            while time.time() - _started < _SSE_MAX_S:
                t = store.get(job_id)
                if not t:
                    break
                status = t.get('status')
                if status in ('done', 'needs_review', 'failed', 'cancelled'):
                    event = {'type': t['status'], 'data': public_task(t)}
                    yield f"data: {json.dumps(event, ensure_ascii=False)}\n\n"
                    break
                if status == 'queued':
                    event = {'type': 'queued', 'data': public_task(t)}
                    yield f"data: {json.dumps(event, ensure_ascii=False)}\n\n"
                else:
                    st = jobs.read_status(t['input_path'])
                    if not st:
                        st = t
                    if t.get('created_at'):
                        st['total_elapsed_s'] = max(0, int(time.time() - t['created_at']))
                    if st.get('status') not in ('done', 'needs_review', 'failed', 'cancelled'):
                        store.update_progress(job_id, st)
                    event = {'type': 'progress', 'data': public_progress(st)}
                    yield f"data: {json.dumps(event, ensure_ascii=False)}\n\n"
                try:
                    ev = q.get(timeout=3)
                except Empty:
                    ev = None
                if ev:
                    event_type = ev.get('type')
                    event_data = ev.get('data') if isinstance(ev.get('data'), dict) else {}
                    if event_type in ('done', 'needs_review', 'failed', 'cancelled'):
                        current = store.get(job_id)
                        event_data = public_task(current) if current else {}
                    else:
                        event_data = public_progress(event_data)
                    event = {'type': event_type, 'data': event_data}
                    yield f"data: {json.dumps(event, ensure_ascii=False)}\n\n"
                    if event_type in ('done', 'needs_review', 'failed', 'cancelled'):
                        break
        finally:
            jobs.unsubscribe(job_id, q)

    return StreamingResponse(gen(), media_type="text/event-stream", headers={
        'Cache-Control': 'no-cache', 'X-Accel-Buffering': 'no'})


@router.get("/api/tasks/{job_id}/download")
def download(job_id: str):
    t = store.get(job_id)
    if not t or not t.get('output_path') or not os.path.exists(t['output_path']):
        raise HTTPException(404, "输出文件不存在")
    real_out = os.path.realpath(t['output_path'])
    real_uploads = os.path.realpath(ctx.upload_dir)
    if not real_out.startswith(real_uploads + os.sep):
        raise HTTPException(403, "非法文件路径")
    output_extension = os.path.splitext(real_out)[1].lower()
    if t.get('pipeline') == 'amazon':
        if output_extension == '.json':
            stem = os.path.splitext(t.get('filename', 'output'))[0]
            download_name = (
                stem.replace('采集表', '回填表', 1)
                if '采集表' in stem
                else stem + '_回填'
            ) + '.json'
            return FileResponse(real_out, filename=download_name)
        suffix = '_回填.xlsx'
    else:
        suffix = '_cleaned.xlsx'
    download_name = os.path.splitext(t.get('filename', 'output'))[0] + suffix
    return FileResponse(real_out, filename=download_name)


@router.get("/api/tasks/{job_id}/review-report")
def review_report(job_id: str):
    import re as _re
    from urllib.parse import quote
    from web.review_report import build_review_report_csv
    t = store.get(job_id)
    if not t:
        raise HTTPException(404, "任务不存在")
    pt = public_task(t)
    report = '﻿' + build_review_report_csv(pt,
        row_contexts=_review_ctx(t),
    )
    stem = os.path.splitext(t.get('filename', 'output'))[0] or 'output'
    stem = _re.sub(r'[<>:"/\\|?*\x00-\x1f]', '_', stem).strip(' .') or 'output'
    download_name = stem[:120] + '_复核报告.csv'
    return StreamingResponse(
        iter([report.encode('utf-8')]),
        media_type='text/csv; charset=utf-8',
        headers={
            'Content-Disposition': (
                "attachment; filename*=UTF-8''" + quote(download_name)
            ),
        },
    )


@router.get("/api/tasks/{job_id}/review-data")
def review_data(job_id: str):
    from web.review_report import build_review_data
    t = store.get(job_id)
    if not t:
        raise HTTPException(404, "任务不存在")
    return build_review_data(
        public_task(t),
        row_contexts=_review_ctx(t),
    )


@router.post("/api/tasks/{job_id}/retry")
async def retry_task(job_id: str, request: Request):
    params = dict(request.query_params)
    fresh = params.get('fresh', '').lower() in ('1', 'true', 'yes')
    t = store.get(job_id)
    if not t:
        raise HTTPException(404, "任务不存在")
    if jobs.is_active(job_id):
        raise HTTPException(409, "任务正在处理中")

    mode = 'fresh' if fresh else 'resume'
    if fresh:
        jobs.clear_cache(t['input_path'])
    store.prepare_retry(job_id)
    jobs.enqueue(job_id, t['input_path'])
    return {'ok': True, 'mode': mode}


@router.post("/api/tasks/{job_id}/cancel")
async def cancel_task(job_id: str):
    t = store.get(job_id)
    if not t:
        raise HTTPException(404, "任务不存在")
    if t['status'] in ('done', 'failed', 'cancelled'):
        raise HTTPException(400, f"任务已结束 ({t['status']})")
    if not jobs.cancel(job_id):
        raise HTTPException(409, '任务正在启动，暂时无法安全取消')
    store.mark_cancelled(job_id)
    return public_task(store.get(job_id))


@router.delete("/api/tasks/{job_id}")
def delete_task(job_id: str):
    import re as _re, shutil
    if not job_id or not _re.match(r'^[0-9a-f]{12}$', job_id):
        raise HTTPException(400, "无效的任务 ID")
    t = store.get(job_id)
    if not t:
        raise HTTPException(404, "任务不存在")
    job_dir = os.path.join(ctx.upload_dir, job_id)
    real_dir = os.path.realpath(job_dir)
    real_upload = os.path.realpath(ctx.upload_dir) + os.sep
    if not (real_dir + os.sep).startswith(real_upload):
        raise HTTPException(400, "路径非法")
    if jobs.is_active(job_id) and not jobs.cancel(job_id):
        raise HTTPException(409, "该任务正在进程内处理，暂时无法安全删除")
    if os.path.exists(job_dir):
        try:
            shutil.rmtree(job_dir)
        except OSError as e:
            raise HTTPException(409, f"任务文件暂时无法删除: {e}")
    store.delete(job_id)
    return {'ok': True}
